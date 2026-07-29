#!/usr/bin/env python3
import openpyxl
from pathlib import Path

file_path = Path(r"C:\Users\user\Downloads\exemplo.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("\n" + "="*100)
print("ANALISE DO ARQUIVO EXEMPLO")
print("="*100)

print(f"\nTotal de colunas: {ws.max_column}")
print(f"Total de linhas: {ws.max_row}")

print("\n[HEADERS]")
print("="*100)

for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    print(f"  Col {col:2d}: {header}")

print("\n[PRIMEIRAS 5 LINHAS DE DADOS]")
print("="*100)

for row in range(2, min(7, ws.max_row + 1)):
    print(f"\nLinha {row}:")

    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        value = ws.cell(row, col).value

        if value is not None:
            valor_str = str(value)[:50]
            print(f"  {header:25s}: {valor_str}")

print("\n" + "="*100)
print("ESTRUTURA ENTENDIDA")
print("="*100)
