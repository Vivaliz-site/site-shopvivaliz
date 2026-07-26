#!/usr/bin/env python3
"""
Análise da planilha de anúncios do Mercado Livre
"""
import openpyxl
from openpyxl.utils import get_column_letter
import sys

def analyze_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        print(f"📄 Arquivo: {file_path}")
        print(f"📊 Planilha ativa: {ws.title}")
        print(f"📈 Dimensões: {ws.dimensions}")

        # Cabeçalho
        header = []
        max_col = ws.max_column
        print(f"\n🔍 Colunas encontradas ({max_col}):")
        for col_idx in range(1, max_col + 1):
            cell_value = ws.cell(1, col_idx).value
            header.append(cell_value)
            print(f"  [{col_idx:2d}] {cell_value}")

        # Primeiras linhas de dados
        print(f"\n📋 Primeiras 5 linhas de dados:")
        for row_idx in range(2, min(7, ws.max_row + 1)):
            print(f"\n  Linha {row_idx}:")
            for col_idx, col_name in enumerate(header, 1):
                cell_value = ws.cell(row_idx, col_idx).value
                # Limita a exibição de valores longos
                if isinstance(cell_value, str) and len(str(cell_value)) > 50:
                    display_value = str(cell_value)[:47] + "..."
                else:
                    display_value = cell_value
                print(f"    {col_name}: {display_value}")

        print(f"\n✅ Total de linhas de dados: {ws.max_row - 1}")

        # Identifica colunas potencialmente úteis
        print("\n🎯 Colunas identificadas:")
        for col_idx, col_name in enumerate(header, 1):
            name_lower = str(col_name).lower()
            if 'preço' in name_lower or 'price' in name_lower or 'valor' in name_lower:
                print(f"  [PREÇO] Coluna {col_idx}: {col_name}")
            elif 'sku' in name_lower or 'código' in name_lower or 'id' in name_lower:
                print(f"  [ID] Coluna {col_idx}: {col_name}")
            elif 'descrição' in name_lower or 'título' in name_lower or 'name' in name_lower:
                print(f"  [DESC] Coluna {col_idx}: {col_name}")
            elif 'ean' in name_lower or 'gtin' in name_lower or 'barcode' in name_lower:
                print(f"  [EAN] Coluna {col_idx}: {col_name}")

        wb.close()
        return True

    except Exception as e:
        print(f"❌ Erro ao analisar: {e}", file=sys.stderr)
        return False

if __name__ == "__main__":
    file_path = "/mnt/data/Anuncios-2026_07_26-09_32(1).xlsx"
    analyze_excel(file_path)
