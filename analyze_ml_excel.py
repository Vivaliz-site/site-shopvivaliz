#!/usr/bin/env python3
"""
Análise da planilha de anúncios do Mercado Livre
"""
import openpyxl
from pathlib import Path

def analyze_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        print(f"📄 Arquivo: {Path(file_path).name}")
        print(f"📊 Planilha ativa: {ws.title}")
        print(f"📈 Dimensões: Linhas={ws.max_row}, Colunas={ws.max_column}")

        # Cabeçalho
        header = []
        print(f"\n🔍 COLUNAS ({ws.max_column}):")
        print("─" * 100)
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(1, col_idx).value
            header.append(cell_value)
            print(f"  [{col_idx:2d}] {cell_value}")

        # Primeiras 3 linhas de dados completas
        print(f"\n📋 PRIMEIRAS 3 LINHAS DE DADOS:")
        print("─" * 100)
        for row_idx in range(2, min(5, ws.max_row + 1)):
            print(f"\n  >>>>>>>>>> LINHA {row_idx} <<<<<<<<<<")
            for col_idx, col_name in enumerate(header, 1):
                cell_value = ws.cell(row_idx, col_idx).value
                # Limita a exibição
                display_value = str(cell_value)[:80] if isinstance(cell_value, str) else cell_value
                print(f"    {col_name}: {display_value}")

        print(f"\n✅ TOTAL DE ANÚNCIOS: {ws.max_row - 1}")

        # Identifica colunas-chave
        print("\n🎯 COLUNAS IDENTIFICADAS:")
        print("─" * 100)
        price_cols = []
        id_cols = []
        desc_cols = []

        for col_idx, col_name in enumerate(header, 1):
            name_lower = str(col_name).lower() if col_name else ""

            if any(x in name_lower for x in ['preço', 'price', 'valor']):
                price_cols.append((col_idx, col_name))
                print(f"  💰 [PREÇO] Col {col_idx}: {col_name}")
            elif any(x in name_lower for x in ['sku', 'código', 'code', 'ean', 'gtin']):
                id_cols.append((col_idx, col_name))
                print(f"  🆔 [ID] Col {col_idx}: {col_name}")
            elif any(x in name_lower for x in ['descrição', 'título', 'name', 'product', 'description']):
                desc_cols.append((col_idx, col_name))
                print(f"  📝 [DESC] Col {col_idx}: {col_name}")

        if not price_cols:
            print("  ⚠️  Nenhuma coluna de preço identificada automaticamente")
        if not id_cols:
            print("  ⚠️  Nenhuma coluna de ID identificada automaticamente")

        wb.close()

        print("\n" + "─" * 100)
        print(f"✅ Análise concluída. Total de linhas: {ws.max_row - 1}")
        return {
            'total_rows': ws.max_row - 1,
            'total_cols': ws.max_column,
            'header': header,
            'price_cols': price_cols,
            'id_cols': id_cols,
            'desc_cols': desc_cols
        }

    except Exception as e:
        print(f"❌ Erro ao analisar: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    file_path = r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx"
    analyze_excel(file_path)
