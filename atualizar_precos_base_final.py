#!/usr/bin/env python3
"""
ESTRATÉGIA CORRIGIDA:
As tabelas de preços são CALCULADAS automaticamente pelo Tiny baseado no acréscimo
PDV (982): cadastro * 1.45
Shopee (989): cadastro * 1.10
Amazon (991): cadastro * 1.00
TikTok (990): cadastro * 1.10

Para atualizar um preço em uma tabela, atualizar o preço CADASTRO
e o Tiny recalcula todas as tabelas automaticamente!
"""
import os
import time
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("ATUALIZAR PRECOS BASE (Tiny calcula automaticamente tabelas)")
print("="*100)

# Mapeamento: coluna -> (nome_tabela, acrescimo)
MAPA_COLUNAS = {
    7: ("Cadastro", 1.00),
    8: ("PDV", 1.45),
    9: ("Shopee", 1.10),
    10: ("Amazon", 1.00),
    11: ("TikTok", 1.10),
}

# 1. LER RELATORIO E IDENTIFICAR PRECOS EM VERMELHO
print("\n[1] Lendo relatorio...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

atualizacoes = {}  # {id_tiny: novo_preco_cadastro}

for row_idx in range(2, ws.max_row + 1):
    id_tiny = ws.cell(row_idx, 3).value
    preco_ml = ws.cell(row_idx, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    novo_preco = round(preco_ml + 0.01, 2)

    # Verificar cada coluna (até achar a primeira com vermelho)
    # Se Cadastro (7) é vermelho, atualizar Cadastro
    # Se PDV (8) é vermelho, calcular Cadastro inverso = Tabela / 1.45
    # E assim por diante

    for col_idx in [7, 8, 9, 10, 11]:
        cell = ws.cell(row_idx, col_idx)

        if cell.fill and cell.fill.start_color:
            color_str = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color_str:
                local_nome, acrescimo = MAPA_COLUNAS[col_idx]

                # Calcular o preco base do Cadastro que resultaria neste preco na tabela
                preco_cadastro_necessario = round(novo_preco / acrescimo, 2)

                if id_tiny in atualizacoes:
                    # Se ja tem, manter o maior
                    atualizacoes[id_tiny] = max(atualizacoes[id_tiny], preco_cadastro_necessario)
                else:
                    atualizacoes[id_tiny] = preco_cadastro_necessario

                break  # Processar só primeira coluna vermelha

wb.close()

print(f"    Identificados {len(atualizacoes)} produtos pra atualizar")
print(f"    Exemplos (ID: novo_preco_cadastro):")

for id_tiny, preco in list(atualizacoes.items())[:3]:
    print(f"      ID {id_tiny}: R$ {preco:.2f}")

# 2. CARREGAR DADOS DOS PRODUTOS
print("\n[2] Carregando dados dos produtos...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

produtos_map = {}

for prod in resp.json().get("itens", []):
    id_prod = prod.get("id")
    produtos_map[id_prod] = {
        'sku': prod.get("sku"),
        'descricao': prod.get("descricao"),
        'preco_atual': prod.get("precos", {}).get("preco")
    }

print(f"    Carregados {len(produtos_map)} produtos")

# 3. ATUALIZAR PRECOS CADASTRO VIA API PUT
print("\n[3] Atualizando precos no Cadastro...")

atualizados = 0
erros = 0

for idx, (id_tiny, preco_novo) in enumerate(atualizacoes.items(), 1):

    if id_tiny not in produtos_map:
        print(f"    [{idx}] ID {id_tiny}: nao carregado")
        erros += 1
        continue

    prod_data = produtos_map[id_tiny]

    try:
        payload = {
            "descricao": prod_data['descricao'],
            "sku": prod_data['sku'],
            "precos": {"preco": preco_novo}
        }

        url = f"https://api.tiny.com.br/public-api/v3/produtos/{id_tiny}"
        resp = requests.put(url, headers=headers, json=payload, timeout=20)

        if resp.status_code in [200, 201, 204]:
            atualizados += 1

            if atualizados % 5 == 0:
                print(f"    [{atualizados}] Atualizados")
        else:
            print(f"    [{idx}] ID {id_tiny}: erro {resp.status_code}")
            erros += 1

        time.sleep(0.3)

    except Exception as e:
        print(f"    [{idx}] ID {id_tiny}: {str(e)[:50]}")
        erros += 1

# 4. VALIDAR
print(f"\n[4] Validando (primeiros 10)...")

validacoes = []

for id_tiny, preco_esperado in list(atualizacoes.items())[:10]:

    try:
        url = f"https://api.tiny.com.br/public-api/v3/produtos/{id_tiny}"
        resp = requests.get(url, headers=headers, timeout=10)

        if resp.status_code == 200:
            prod = resp.json()
            preco_atual = prod.get("precos", {}).get("preco")

            ok = abs(float(preco_atual) - preco_esperado) < 0.01 if preco_atual else False

            validacoes.append({'id': id_tiny, 'ok': ok, 'preco': preco_atual})

            status = "OK" if ok else "XX"
            print(f"    {status} ID {id_tiny}: esperado R$ {preco_esperado:.2f}, obteve R$ {preco_atual}")

        time.sleep(0.2)

    except Exception as e:
        print(f"    XX ID {id_tiny}: erro")

# RELATORIO FINAL
print(f"\n" + "="*100)
print("RELATORIO FINAL")
print("="*100)

print(f"\n[RESUMO]")
print(f"  Total identificado: {len(atualizacoes)}")
print(f"  Atualizados: {atualizados}")
print(f"  Erros: {erros}")

print(f"\n[VALIDACAO]")
validadas_ok = sum(1 for v in validacoes if v['ok'])
print(f"  Confirmadas: {validadas_ok}/{len(validacoes)}")

print(f"\n✅ Processo concluido!")
