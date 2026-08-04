#!/usr/bin/env python3
import os
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\nCarregando tabelas com acrescimo...")

TABELAS = {"PDV": 982, "Shopee": 989, "Amazon": 991, "TikTok": 990}
tabelas_config = {}

for nome_tabela, id_tabela in TABELAS.items():
    url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}"
    resp = requests.get(url, headers=headers, timeout=20)

    if resp.status_code == 200:
        data = resp.json()
        acrescimo = data.get("acrescimoDesconto", 0)
        tabelas_config[nome_tabela] = acrescimo

print(f"PDV: {tabelas_config.get('PDV')}%")
print(f"Shopee: {tabelas_config.get('Shopee')}%")
print(f"Amazon: {tabelas_config.get('Amazon')}%")
print(f"TikTok: {tabelas_config.get('TikTok')}%")

print("\nAtualizando relatorio...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_FINAL_6_PRECOS.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

for row_idx in range(2, ws.max_row + 1):
    preco_cadastro = ws.cell(row_idx, 7).value

    if not preco_cadastro:
        continue

    try:
        preco_cadastro = float(preco_cadastro)
    except:
        continue

    # PDV (col 8): +45%
    ws.cell(row_idx, 8, round(preco_cadastro * 1.45, 2))
    ws.cell(row_idx, 8).number_format = 'R$ #,##0.00'

    # Shopee (col 9): +10%
    ws.cell(row_idx, 9, round(preco_cadastro * 1.10, 2))
    ws.cell(row_idx, 9).number_format = 'R$ #,##0.00'

    # Amazon (col 10): +0%
    ws.cell(row_idx, 10, round(preco_cadastro * 1.00, 2))
    ws.cell(row_idx, 10).number_format = 'R$ #,##0.00'

    # TikTok (col 11): +10%
    ws.cell(row_idx, 11, round(preco_cadastro * 1.10, 2))
    ws.cell(row_idx, 11).number_format = 'R$ #,##0.00'

wb.save(file_path)
wb.close()

print(f"\nSalvo: {file_path.name}")
print("Pronto para download!")
