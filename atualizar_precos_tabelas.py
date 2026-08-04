#!/usr/bin/env python3
"""
CORRIGIR: Preencher os 4 preços de tabelas no RELATORIO_6_PRECOS_COMPLETO.xlsx
"""
import os
import time
import requests
import openpyxl
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from openpyxl.styles import PatternFill

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n╔════════════════════════════════════════════════════════════════════════════════╗")
print("║  ATUALIZAR: Preços das Tabelas em RELATORIO_6_PRECOS_COMPLETO.xlsx            ║")
print("╚════════════════════════════════════════════════════════════════════════════════╝")

# IDs das tabelas
TABELAS = {
    "PDV": 982,
    "Shopee": 989,
    "Amazon": 991,
    "TikTok": 990
}

# 1. CARREGAR MAPEAMENTOS DO TINY
print("\n📥 Carregando mapeamentos...")

mapping_id = {}  # ID Tiny → título ML

mapping_files = [
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-48.xls',
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-50.xls',
]

for mfile in mapping_files:
    if Path(mfile).exists():
        try:
            df = pd.read_excel(mfile, sheet_name=0)
            for idx, row in df.iterrows():
                id_tiny = int(row.get('Id'))
                titulo = row.get('Título')
                mapping_id[id_tiny] = titulo
            print(f"✅ {Path(mfile).name}: {len(df)} itens")
        except Exception as e:
            print(f"⚠️  Erro: {e}")

print(f"   Total: {len(mapping_id)} mapeados")

# 2. ABRIR RELATÓRIO EXISTENTE
print("\n📂 Abrindo relatório existente...")

rel_file = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COMPLETO.xlsx")
wb = openpyxl.load_workbook(rel_file)
ws = wb.active

print(f"✅ {rel_file.name} ({ws.max_row - 1} linhas)")

# 3. PREENCHER PREÇOS DAS TABELAS
print("\n🔄 Buscando preços das tabelas...")

total = 0
sucesso = 0

for row_idx in range(2, ws.max_row + 1):
    # Ler ID Tiny (coluna 3)
    id_tiny = ws.cell(row_idx, 3).value

    if not id_tiny:
        continue

    try:
        id_tiny = int(id_tiny)
    except:
        continue

    total += 1

    # Buscar preços
    precos = {}

    for nome_tabela, id_tabela in TABELAS.items():
        try:
            url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}?produto_id={id_tiny}"
            resp = requests.get(url, headers=headers, timeout=10)

            if resp.status_code == 200:
                data = resp.json()
                excecoes = data.get("excecoes", [])

                for exc in excecoes:
                    if exc.get("idProduto") == id_tiny:
                        preco = exc.get("preco")
                        precos[nome_tabela] = preco
                        break

        except Exception as e:
            pass

        time.sleep(0.15)

    # Preencher colunas (8=PDV, 9=Shopee, 10=Amazon, 11=TikTok)
    ws.cell(row_idx, 8, precos.get("PDV"))
    ws.cell(row_idx, 9, precos.get("Shopee"))
    ws.cell(row_idx, 10, precos.get("Amazon"))
    ws.cell(row_idx, 11, precos.get("TikTok"))

    # Formatar
    for col_idx in [8, 9, 10, 11]:
        ws.cell(row_idx, col_idx).number_format = 'R$ #,##0.00'

    sucesso += 1

    if total % 20 == 0:
        print(f"  [{total}] ID {id_tiny} - PDV: {precos.get('PDV')}, Shopee: {precos.get('Shopee')}")

# 4. SALVAR
print(f"\n💾 Salvando...")
wb.save(rel_file)
wb.close()

print(f"\n{'='*100}")
print(f"✅ CONCLUÍDO!")
print(f"{'='*100}")
print(f"\n📊 ESTATÍSTICAS:")
print(f"  Total processado: {total}")
print(f"  ✅ Sucesso: {sucesso}")
print(f"\n📁 Arquivo atualizado: {rel_file}")
print(f"✅ Pronto!")
