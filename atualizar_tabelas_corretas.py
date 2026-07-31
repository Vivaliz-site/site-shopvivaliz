#!/usr/bin/env python3
"""
CORRIGIDO: Atualizar APENAS a tabela/cadastro correspondente a cada celula vermelha
Coluna 7 (Cadastro) -> atualizar /produtos/{id}
Coluna 8 (PDV 982) -> atualizar /listas-precos/982
Coluna 9 (Shopee 989) -> atualizar /listas-precos/989
Coluna 10 (Amazon 991) -> atualizar /listas-precos/991
Coluna 11 (TikTok 990) -> atualizar /listas-precos/990
"""
import os
import time
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("ATUALIZAR TABELAS CORRETAS (cada celula vermelha em seu local)")
print("="*100)

# Mapeamento: coluna -> (nome_tabela, id_tabela_ou_cadastro)
MAPA_COLUNAS = {
    7: ("Cadastro", None),  # None = usar PUT /produtos/{id}
    8: ("PDV", 982),
    9: ("Shopee", 989),
    10: ("Amazon", 991),
    11: ("TikTok", 990),
}

# 1. LER RELATORIO E IDENTIFICAR PRECOS EM VERMELHO
print("\n[1] Lendo relatorio...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

atualizacoes = []  # Lista de {id_tiny, local, tabela_id, preco_novo}

for row_idx in range(2, ws.max_row + 1):
    id_tiny = ws.cell(row_idx, 3).value
    preco_ml = ws.cell(row_idx, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    # Verificar cada coluna
    for col_idx, (local_nome, tabela_id) in MAPA_COLUNAS.items():
        cell = ws.cell(row_idx, col_idx)

        if cell.fill and cell.fill.start_color:
            color_str = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color_str:
                novo_preco = round(preco_ml + 0.01, 2)

                atualizacoes.append({
                    'id_tiny': id_tiny,
                    'local': local_nome,
                    'tabela_id': tabela_id,
                    'preco_novo': novo_preco
                })

wb.close()

print(f"    Identificadas {len(atualizacoes)} atualizacoes")

# Contar por local
from collections import Counter
contagem = Counter(a['local'] for a in atualizacoes)

for local, count in contagem.items():
    print(f"      {local:15s}: {count}")

# 2. CARREGAR DADOS DOS PRODUTOS
print("\n[2] Carregando dados dos produtos...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

produtos_map = {}

for prod in resp.json().get("itens", []):
    id_prod = prod.get("id")
    produtos_map[id_prod] = {
        'sku': prod.get("sku"),
        'descricao': prod.get("descricao")
    }

print(f"    Carregados {len(produtos_map)} produtos")

# 3. ATUALIZAR PRECOS NOS LOCAIS CORRETOS
print("\n[3] Atualizando precos...")

atualizados = 0
erros = 0

for idx, atua in enumerate(atualizacoes, 1):
    id_tiny = atua['id_tiny']
    local = atua['local']
    tabela_id = atua['tabela_id']
    preco_novo = atua['preco_novo']

    if id_tiny not in produtos_map:
        print(f"    [{idx}] ID {id_tiny} {local:15s}: nao carregado")
        erros += 1
        continue

    prod_data = produtos_map[id_tiny]

    try:
        if local == "Cadastro":
            # PUT /produtos/{id}
            payload = {
                "descricao": prod_data['descricao'],
                "sku": prod_data['sku'],
                "precos": {"preco": preco_novo}
            }

            url = f"https://api.tiny.com.br/public-api/v3/produtos/{id_tiny}"
            resp = requests.put(url, headers=headers, json=payload, timeout=20)

        else:
            # PUT /listas-precos/{tabela_id}/produtos/{id_tiny}
            # Ou POST /listas-precos/{tabela_id} com lista de precos
            payload = {
                "produtos": [
                    {
                        "idProduto": id_tiny,
                        "preco": preco_novo
                    }
                ]
            }

            url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{tabela_id}"
            resp = requests.post(url, headers=headers, json=payload, timeout=20)

        if resp.status_code in [200, 201, 204]:
            atualizados += 1

            if atualizados % 5 == 0:
                print(f"    [{atualizados}] Atualizados com sucesso")
        else:
            print(f"    [{idx}] ID {id_tiny} {local:15s}: erro {resp.status_code}")
            erros += 1

        time.sleep(0.3)

    except Exception as e:
        print(f"    [{idx}] ID {id_tiny} {local:15s}: {str(e)[:50]}")
        erros += 1

# 4. VALIDAR
print(f"\n[4] Validando (primeiros 5)...")

validacoes = []

for atua in atualizacoes[:5]:
    id_tiny = atua['id_tiny']
    local = atua['local']
    tabela_id = atua['tabela_id']
    preco_novo = atua['preco_novo']

    try:
        url = f"https://api.tiny.com.br/public-api/v3/produtos/{id_tiny}"
        resp = requests.get(url, headers=headers, timeout=10)

        if resp.status_code == 200:
            prod = resp.json()
            preco_atual = prod.get("precos", {}).get("preco")

            ok = abs(float(preco_atual) - preco_novo) < 0.01 if preco_atual else False

            validacoes.append({'id': id_tiny, 'local': local, 'ok': ok, 'preco': preco_atual})

            status = "✓" if ok else "✗"
            print(f"    {status} ID {id_tiny} ({local:15s}): R$ {preco_atual}")

        time.sleep(0.2)

    except Exception as e:
        print(f"    ✗ ID {id_tiny}: erro")

# RELATORIO FINAL
print(f"\n" + "="*100)
print("RELATORIO FINAL")
print("="*100)

print(f"\n[RESUMO]")
print(f"  Total identificado: {len(atualizacoes)}")
print(f"  Atualizados: {atualizados}")
print(f"  Erros: {erros}")

print(f"\n[VALIDACAO]")
validadas_ok = sum(1 for v in validacoes if v['ok'])
print(f"  Confirmadas: {validadas_ok}/{len(validacoes)}")

print(f"\n✅ Processo concluido!")
