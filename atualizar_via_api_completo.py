#!/usr/bin/env python3
"""
ATUALIZAR E VALIDAR TUDO VIA API:
1. Ler precos em vermelho do relatorio (Excel)
2. Atualizar NO TINY via API para ML + 0.01
3. Extrair dados do Tiny via API para validar se salvou
"""
import os
import time
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("ATUALIZAR PRECOS VIA API E VALIDAR")
print("="*100)

# 1. LER RELATORIO E IDENTIFICAR PRECOS EM VERMELHO
print("\n[1] Lendo relatorio para identificar precos em vermelho...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

atualizacoes_pendentes = []

for row_idx in range(2, ws.max_row + 1):
    id_tiny = ws.cell(row_idx, 3).value
    titulo_ml = ws.cell(row_idx, 2).value
    preco_ml = ws.cell(row_idx, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    # Verificar se tem vermelho nesta linha
    tem_vermelho = False

    for col_idx in [7, 8, 9, 10, 11]:
        cell = ws.cell(row_idx, col_idx)

        if cell.fill and cell.fill.start_color:
            color_str = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color_str:
                tem_vermelho = True
                break

    if tem_vermelho:
        novo_preco = round(preco_ml + 0.01, 2)

        atualizacoes_pendentes.append({
            'id_tiny': id_tiny,
            'titulo': titulo_ml,
            'preco_ml': preco_ml,
            'novo_preco': novo_preco
        })

wb.close()

print(f"    Identificados {len(atualizacoes_pendentes)} produtos com vermelho")

if atualizacoes_pendentes:
    print(f"\n    Primeiros 3 exemplos:")
    for atua in atualizacoes_pendentes[:3]:
        print(f"      ID {atua['id_tiny']:10d}: ML R$ {atua['preco_ml']:7.2f} -> Novo R$ {atua['novo_preco']:7.2f}")

# 2. CARREGAR DADOS DOS PRODUTOS VIA API
print("\n[2] Carregando dados dos produtos via API...")

print(f"    GET /produtos para extrair sku e descricao...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

if resp.status_code != 200:
    print(f"    Erro: {resp.status_code}")
    exit(1)

produtos_map = {}

for prod in resp.json().get("itens", []):
    id_prod = prod.get("id")
    sku = prod.get("sku")
    desc = prod.get("descricao")

    produtos_map[id_prod] = {
        'sku': sku,
        'descricao': desc
    }

print(f"    Carregados {len(produtos_map)} produtos")

# 3. ATUALIZAR PRECOS VIA API PUT
print("\n[3] Atualizando precos via API PUT...")

atualizacoes_ok = []
atualizacoes_erro = []

for idx, atua in enumerate(atualizacoes_pendentes, 1):
    id_tiny = atua['id_tiny']

    if id_tiny not in produtos_map:
        print(f"    [{idx}] ID {id_tiny}: ERRO - nao carregado")
        atualizacoes_erro.append(atua)
        continue

    prod_data = produtos_map[id_tiny]

    # Preparar payload
    payload = {
        "descricao": prod_data['descricao'],
        "sku": prod_data['sku'],
        "precos": {
            "preco": atua['novo_preco']
        }
    }

    # PUT
    url = f"https://api.tiny.com.br/public-api/v3/produtos/{id_tiny}"

    try:
        resp = requests.put(url, headers=headers, json=payload, timeout=20)

        if resp.status_code in [200, 201, 204]:  # 204 = No Content (sucesso)
            atualizacoes_ok.append(atua)

            if len(atualizacoes_ok) % 10 == 0:
                print(f"    [{len(atualizacoes_ok)}] Atualizado com sucesso")
        else:
            print(f"    [{idx}] ID {id_tiny}: Erro {resp.status_code}")
            atualizacoes_erro.append(atua)

        time.sleep(0.5)

    except Exception as e:
        print(f"    [{idx}] ID {id_tiny}: {str(e)[:50]}")
        atualizacoes_erro.append(atua)

# 4. VALIDAR VIA API GET
print(f"\n[4] Validando via API GET (primeiros 10 produtos)...")

validacoes = []

for atua in atualizacoes_ok[:10]:
    id_tiny = atua['id_tiny']

    url = f"https://api.tiny.com.br/public-api/v3/produtos/{id_tiny}"

    try:
        resp = requests.get(url, headers=headers, timeout=10)

        if resp.status_code == 200:
            prod = resp.json()
            preco_atual = prod.get("precos", {}).get("preco")

            ok = abs(float(preco_atual) - atua['novo_preco']) < 0.01 if preco_atual else False

            validacoes.append({
                'id': id_tiny,
                'esperado': atua['novo_preco'],
                'atual': preco_atual,
                'ok': ok
            })

            status = "✓" if ok else "✗"
            print(f"    {status} ID {id_tiny:10d}: esperado R$ {atua['novo_preco']:.2f}, atual R$ {preco_atual}")
        else:
            print(f"    ✗ ID {id_tiny}: erro {resp.status_code}")

        time.sleep(0.2)

    except Exception as e:
        print(f"    ✗ ID {id_tiny}: {str(e)[:50]}")

# RELATORIO FINAL
print(f"\n" + "="*100)
print("RELATORIO FINAL")
print("="*100)

print(f"\n[RESUMO]")
print(f"  Total de produtos com vermelho identificado: {len(atualizacoes_pendentes)}")
print(f"  Atualizados com sucesso: {len(atualizacoes_ok)}")
print(f"  Erros na atualizacao: {len(atualizacoes_erro)}")

print(f"\n[VALIDACAO]")
validacoes_ok = sum(1 for v in validacoes if v['ok'])
print(f"  Validadas: {len(validacoes)}")
print(f"  Confirmadas (salvo corretamente): {validacoes_ok}")

if validacoes_erro := [v for v in validacoes if not v['ok']]:
    print(f"  Falhas na validacao: {len(validacoes_erro)}")

    for v in validacoes_erro:
        print(f"    ID {v['id']}: esperado R$ {v['esperado']}, obteve R$ {v['atual']}")

print(f"\n✅ Processo concluido!")
