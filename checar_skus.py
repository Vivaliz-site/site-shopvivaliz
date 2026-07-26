#!/usr/bin/env python3
import openpyxl
from pathlib import Path

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("\n\nPrimeiros SKUs do RELATORIO:")
print("="*60)

for row in range(2, min(15, ws.max_row + 1)):
    sku = ws.cell(row, 5).value
    id_tiny = ws.cell(row, 3).value
    titulo = ws.cell(row, 2).value

    if id_tiny and titulo:
        print(f"  ID {id_tiny}: SKU={sku} | {str(titulo)[:40]}")
