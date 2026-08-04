#!/usr/bin/env python3
"""
CORRIGIR: Buscar precos das tabelas SEM filtrar por produto especifico
Se cada tabela tem acrescimo/desconto, talvez o preco base seja: preco_cadastro + acrescimo
"""
import os
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("CORRIGIR: Buscar precos das tabelas com acrescimo/desconto")
print("="*100)

# Carregar tabelas
print("\n[1] Carregando tabelas com acrescimo/desconto...")

TABELAS = {"PDV": 982, "Shopee": 989, "Amazon": 991, "TikTok": 990}
tabelas_config = {}

for nome_tabela, id_tabela in TABELAS.items():
    url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}"

    resp = requests.get(url, headers=headers, timeout=20)

    if resp.status_code == 200:
        data = resp.json()
        acrescimo = data.get("acrescimoDesconto", 0)  # % de acrescimo/desconto
        descricao = data.get("descricao", "")

        tabelas_config[nome_tabela] = {
            'id': id_tabela,
            'acrescimo': acrescimo,
            'descricao': descricao
        }

        print(f"    {nome_tabela:15s}: acrescimo={acrescimo:6.1f}%")

# Abrir relatorio
print("\n[2] Abrindo relatorio para atualizar...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COMPLETO_ATUALIZADO.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"    {file_path.name}")

# Calcular precos baseado em acrescimo
print("\n[3] Calculando precos com acrescimo/desconto...\n")

from openpyxl.styles import PatternFill

for row_idx in range(2, ws.max_row + 1):
    preco_cadastro = ws.cell(row_idx, 7).value

    if not preco_cadastro or preco_cadastro == 0:
        continue

    # Converter para float
    try:
        preco_cadastro = float(preco_cadastro)
    except:
        continue

    # Calcular para cada tabela
    for col_idx, nome_tabela in enumerate([("PDV", 8), ("Shopee", 9), ("Amazon", 10), ("TikTok", 11)], 1):
        tabela_nome, col = nome_tabela

        if tabela_nome in tabelas_config:
            acrescimo_pct = tabelas_config[tabela_nome]['acrescimo']

            # Calcular preco: preco_cadastro * (1 + acrescimo/100)
            preco_tabela = preco_cadastro * (1 + acrescimo_pct / 100)

            # Arredondar para 2 casas
            preco_tabela = round(preco_tabela, 2)

            # Preencher
            ws.cell(row_idx, col, preco_tabela)
            ws.cell(row_idx, col).number_format = 'R$ #,##0.00'

    if row_idx % 20 == 0:
        titulo = ws.cell(row_idx, 2).value
        print(f"  [{row_idx-1}] {str(titulo)[:50]}")

# Salvar
print(f"\n[4] Salvando...")
wb.save(file_path)
wb.close()

print(f"\n" + "="*100)
print("SUCESSO!")
print("="*100)
print(f"\nArquivo atualizado: {file_path}")
