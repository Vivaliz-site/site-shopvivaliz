#!/usr/bin/env python3
import pandas as pd
from pathlib import Path
import openpyxl

# Carregar titulos do mapeamento
print("MAPEAMENTO (Anuncios):")
mapping_titles = {}

for mfile in [
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-48.xls',
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-50.xls',
]:
    if Path(mfile).exists():
        df = pd.read_excel(mfile, sheet_name=0)
        for idx, row in df.iterrows():
            titulo = str(row['Título']).strip()
            id_tiny = int(row['Id'])
            mapping_titles[titulo] = id_tiny

print(f"  Total: {len(mapping_titles)} titulos unicos")
print(f"  Exemplos:")
for t in list(mapping_titles.keys())[:3]:
    print(f"    - {t[:60]}")

# Carregar titulos da planilha base ML
print("\nPLANILHA BASE ML:")

ml_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
wb = openpyxl.load_workbook(ml_file, data_only=True)
ws = wb.active

# Encontrar coluna TITLE
title_col = None
for col_idx in range(1, ws.max_column + 1):
    if ws.cell(1, col_idx).value == "TITLE":
        title_col = col_idx
        break

ml_titles = {}
for row_idx in range(6, ws.max_row + 1):
    title = ws.cell(row_idx, title_col).value
    if title:
        titulo = str(title).strip()
        ml_titles[titulo] = row_idx

print(f"  Total: {len(ml_titles)} titulos")
print(f"  Exemplos:")
for t in list(ml_titles.keys())[:3]:
    print(f"    - {t[:60]}")

wb.close()

# Comparar
print("\nCOMPARACAO:")
encontrados = 0
nao_encontrados_list = []

for ml_titulo in ml_titles.keys():
    if ml_titulo in mapping_titles:
        encontrados += 1
    else:
        nao_encontrados_list.append(ml_titulo)

print(f"  Encontrados no mapeamento: {encontrados}/{len(ml_titles)}")
print(f"  Nao encontrados: {len(nao_encontrados_list)}")

if nao_encontrados_list:
    print(f"\n  Exemplos nao encontrados no mapeamento:")
    for titulo in nao_encontrados_list[:10]:
        print(f"    - {titulo[:60]}")

# Procurar por similaridade
print(f"\n  Tentando buscar por similaridade...")

from difflib import SequenceMatcher

encontrados_sim = 0

for ml_titulo in nao_encontrados_list[:5]:
    print(f"\n    Procurando: {ml_titulo[:50]}")

    best_match = None
    best_score = 0

    for map_titulo in mapping_titles.keys():
        score = SequenceMatcher(None, ml_titulo.lower(), map_titulo.lower()).ratio()

        if score > best_score:
            best_score = score
            best_match = map_titulo

    if best_score > 0.5:
        print(f"      Encontrado: {best_match[:50]} (score: {best_score:.2f})")
        encontrados_sim += 1
    else:
        print(f"      Nao encontrado")

print(f"\n  Encontrados por similaridade: {encontrados_sim}")
