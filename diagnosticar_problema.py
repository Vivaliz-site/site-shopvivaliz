#!/usr/bin/env python3
"""
Diagnosticar: quais IDs foram realmente atualizados vs quais faltam
"""
import os
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv
from collections import defaultdict

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("DIAGNOSTICAR PROBLEMA DE ATUALIZACAO")
print("="*100)

# 1. LER RELATORIO ORIGINAL
print("\n[1] Lendo relatorio original...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Mapeamento de colunas
MAPA_COLUNAS = {
    7: "Cadastro",
    8: "PDV",
    9: "Shopee",
    10: "Amazon",
    11: "TikTok",
}

vermelho_por_id = defaultdict(list)  # {id: [(row, col, preco_ml), ...]}

for row_idx in range(2, ws.max_row + 1):
    id_tiny = ws.cell(row_idx, 3).value
    preco_ml = ws.cell(row_idx, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    # Verificar cada coluna
    for col_idx in [7, 8, 9, 10, 11]:
        cell = ws.cell(row_idx, col_idx)

        if cell.fill and cell.fill.start_color:
            color_str = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color_str:
                vermelho_por_id[id_tiny].append({
                    'row': row_idx,
                    'col': col_idx,
                    'col_nome': MAPA_COLUNAS[col_idx],
                    'preco_ml': preco_ml
                })

wb.close()

print(f"    Identificados {len(vermelho_por_id)} IDs com vermelho")
print(f"    Total de células vermelhas: {sum(len(v) for v in vermelho_por_id.values())}")

# 2. VERIFICAR ATUAL NO TINY
print("\n[2] Carregando dados atuais do Tiny...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

produtos_tiny = {}

for prod in resp.json().get("itens", []):
    id_prod = prod.get("id")
    produtos_tiny[id_prod] = prod.get("precos", {}).get("preco")

print(f"    Carregados {len(produtos_tiny)} produtos")

# 3. ANALISAR CADA ID
print(f"\n[3] Analisando cada ID...")

atualizado_ok = 0
nao_atualizado = 0
erro_404 = 0

for id_tiny, vermelhos in sorted(vermelho_por_id.items()):
    if id_tiny not in produtos_tiny:
        print(f"\n    ID {id_tiny:10d}: ERRO 404")
        print(f"      Vermelhos em: {', '.join(v['col_nome'] for v in vermelhos)}")
        erro_404 += 1
        continue

    preco_atual = float(produtos_tiny[id_tiny])

    # Qual é o preco esperado?
    # Comparar com o ML price (deve ser ML + 0.01)
    preco_ml = vermelhos[0]['preco_ml']
    preco_esperado = round(preco_ml + 0.01, 2)

    # Verificar se foi atualizado
    if abs(preco_atual - preco_esperado) < 0.01:
        print(f"\n    ID {id_tiny:10d}: OK")
        print(f"      Preco esperado: R$ {preco_esperado:.2f}")
        print(f"      Preco atual:    R$ {preco_atual:.2f}")
        atualizado_ok += 1
    else:
        print(f"\n    ID {id_tiny:10d}: NAO ATUALIZADO")
        print(f"      Preco esperado: R$ {preco_esperado:.2f}")
        print(f"      Preco atual:    R$ {preco_atual:.2f}")
        print(f"      Vermelhos em: {', '.join(v['col_nome'] for v in vermelhos)}")
        nao_atualizado += 1

# RESUMO
print(f"\n" + "="*100)
print("RESUMO")
print("="*100)

print(f"\n  IDs com vermelho identificados: {len(vermelho_por_id)}")
print(f"  Atualizados OK: {atualizado_ok}")
print(f"  Nao atualizados: {nao_atualizado}")
print(f"  Erro 404: {erro_404}")

if nao_atualizado > 0:
    print(f"\n⚠️ PROBLEMA: {nao_atualizado} IDs NÃO foram atualizados!")
    print(f"Motivo provável: o script atualizar_precos_base_final.py processou só os primeiros 8 IDs únicos")
    print(f"e não reprocessou o restante.")
