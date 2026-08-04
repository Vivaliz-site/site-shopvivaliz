#!/usr/bin/env python3
"""
Diagnostico completo: por que tantos ainda falhando?
"""
import os
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv
from collections import defaultdict

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("DIAGNOSTICO COMPLETO")
print("="*100)

# 1. LER RELATORIO
print("\n[1] Lendo relatorio...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

ids_no_relatorio = set()
precos_ml_por_id = {}

for row_idx in range(2, ws.max_row + 1):
    id_tiny = ws.cell(row_idx, 3).value
    preco_ml = ws.cell(row_idx, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    ids_no_relatorio.add(id_tiny)
    if id_tiny not in precos_ml_por_id:
        precos_ml_por_id[id_tiny] = preco_ml

wb.close()

print(f"    IDs no relatorio: {len(ids_no_relatorio)}")

# 2. CARREGAR DADOS DO TINY (pagina 1)
print("\n[2] Carregando dados do Tiny (pagina 1)...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

ids_no_tiny_pag1 = set()
precos_tiny = {}

for prod in resp.json().get("itens", []):
    id_prod = prod.get("id")
    ids_no_tiny_pag1.add(id_prod)
    precos_tiny[id_prod] = prod.get("precos", {}).get("preco")

print(f"    IDs no Tiny (pagina 1): {len(ids_no_tiny_pag1)}")

# 3. ANALISAR DIFERENCA
print("\n[3] Analisando diferenca...")

ids_faltando = ids_no_relatorio - ids_no_tiny_pag1
ids_encontrados = ids_no_relatorio & ids_no_tiny_pag1

print(f"\n    IDs no relatorio: {len(ids_no_relatorio)}")
print(f"    IDs encontrados no Tiny: {len(ids_encontrados)}")
print(f"    IDs FALTANDO no Tiny: {len(ids_faltando)}")

if ids_faltando:
    print(f"\n    Primeiros IDs faltando:")
    for id_faltando in list(ids_faltando)[:5]:
        print(f"      {id_faltando}")

# 4. VALIDAR OS QUE FORAM ENCONTRADOS
print(f"\n[4] Validando os {len(ids_encontrados)} IDs encontrados...")

ok_count = 0
erro_count = 0
erros = []

for id_tiny in ids_encontrados:
    preco_ml = precos_ml_por_id[id_tiny]
    preco_esperado = round(preco_ml + 0.01, 2)
    preco_atual = precos_tiny[id_tiny]

    if abs(float(preco_atual) - preco_esperado) < 0.01:
        ok_count += 1
    else:
        erro_count += 1
        erros.append({
            'id': id_tiny,
            'preco_ml': preco_ml,
            'preco_esperado': preco_esperado,
            'preco_atual': preco_atual
        })

print(f"    OK: {ok_count}")
print(f"    ERRO: {erro_count}")

if erros:
    print(f"\n    Primeiros erros:")
    for erro in erros[:5]:
        print(f"      ID {erro['id']}: ML R$ {erro['preco_ml']:.2f} -> esperado R$ {erro['preco_esperado']:.2f}, obteve R$ {erro['preco_atual']}")

# 5. RESUMO
print(f"\n" + "="*100)
print("RESUMO DIAGNOSTICO")
print("="*100)

print(f"\n  Total de IDs no relatorio: {len(ids_no_relatorio)}")
print(f"  IDs encontrados no Tiny (pag 1): {len(ids_encontrados)}")
print(f"  IDs NAO encontrados: {len(ids_faltando)}")
print(f"\n  Dos encontrados:")
print(f"    OK (atualizado): {ok_count}")
print(f"    NAO OK: {erro_count}")

if len(ids_encontrados) > 0:
    taxa_sucesso_geral = 100 * ok_count / len(ids_no_relatorio)
    taxa_sucesso_encontrados = 100 * ok_count / len(ids_encontrados)

    print(f"\n  Taxa de sucesso (geral): {taxa_sucesso_geral:.1f}%")
    print(f"  Taxa de sucesso (encontrados): {taxa_sucesso_encontrados:.1f}%")

print(f"\n⚠️ RAZAO POSSIVEL DA TAXA BAIXA:")
print(f"   - {len(ids_faltando)} IDs ({100*len(ids_faltando)/len(ids_no_relatorio):.1f}%) não estão no GET /produtos")
print(f"   - Isso pode ser porque GET limit=100 só retorna os primeiros 100 produtos")
print(f"   - Precisaríamos de paginacao (offset) pra buscar TODOS os produtos")
