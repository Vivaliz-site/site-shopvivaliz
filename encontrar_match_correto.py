#!/usr/bin/env python3
"""
Encontrar o match correto entre:
- ID do Tiny (do relatorio)
- ID do anuncio (do arquivo anuncios_*.xls)
"""
import openpyxl
from pathlib import Path

try:
    import xlrd
except:
    import os
    os.system("pip install xlrd -q")
    import xlrd

print("\n" + "="*100)
print("ENTENDER O MATCH ENTRE RELATORIO E ANUNCIOS")
print("="*100)

# 1. LER RELATORIO
print("\n[1] Lendo relatorio...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb_orig = openpyxl.load_workbook(file_path)
ws_orig = wb_orig.active

relatorio_dados = []

for row in range(2, min(10, ws_orig.max_row + 1)):
    titulo_ml = ws_orig.cell(row, 2).value
    id_tiny = ws_orig.cell(row, 3).value

    if id_tiny and titulo_ml:
        try:
            id_tiny = int(id_tiny)
            relatorio_dados.append({
                'id_tiny': id_tiny,
                'titulo': str(titulo_ml)[:50]
            })
        except:
            pass

print(f"    Primeiros dados do relatorio:")
for d in relatorio_dados[:5]:
    print(f"      ID Tiny: {d['id_tiny']:12d} | Titulo: {d['titulo']}")

# 2. LER ANUNCIOS
print("\n[2] Lendo anuncios...")

downloads_path = Path(r"C:\Users\user\Downloads")
anuncios_files = sorted(downloads_path.glob("anuncios_2026-07-26-12-*.xls"))

anuncios_dados = []

for file_path in anuncios_files:
    try:
        workbook = xlrd.open_workbook(str(file_path))
        worksheet = workbook.sheet_by_index(0)

        for row in range(1, min(10, worksheet.nrows)):
            id_anuncio = worksheet.cell_value(row, 0)
            integracao = worksheet.cell_value(row, 1)
            identificador = worksheet.cell_value(row, 2)
            titulo = worksheet.cell_value(row, 3)
            sku = worksheet.cell_value(row, 4)

            anuncios_dados.append({
                'id_anuncio': id_anuncio,
                'integracao': integracao,
                'identificador': identificador,
                'titulo': str(titulo)[:50] if titulo else '',
                'sku': sku
            })

    except Exception as e:
        if "Workbook corruption" not in str(e):
            pass

print(f"    Primeiros dados dos anuncios:")
for d in anuncios_dados[:5]:
    print(f"      ID Anuncio: {d['id_anuncio']} | Integracao: {d['integracao']} | Titulo: {d['titulo']}")

# 3. PROCURAR MATCH
print("\n[3] Procurando MATCH por TITULO...")
print("="*100)

import difflib

for rel in relatorio_dados[:3]:
    print(f"\nRelatorio - ID Tiny {rel['id_tiny']}: '{rel['titulo']}'")
    print("  Possíveis matches nos anuncios:")

    for anunc in anuncios_dados:
        # Comparar titulos
        ratio = difflib.SequenceMatcher(None, rel['titulo'].lower(), anunc['titulo'].lower()).ratio()

        if ratio > 0.6:  # 60% de similaridade
            print(f"    {ratio:.1%} - ID Anunc {anunc['id_anuncio']:10} | {anunc['integracao']:15s} | {anunc['identificador']:20s} | {anunc['titulo']}")

print("\n" + "="*100)
print("CONCLUSAO")
print("="*100)
print("""
O match é feito por SIMILARIDADE DE TITULO!
- ID Tiny do relatorio NÃO bate com ID dos anuncios
- Usar TITULO como chave para encontrar o anúncio correspondente
- Então extrair: Integração, Identificador (ID do anúncio na plataforma), SKU
""")
