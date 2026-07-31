#!/usr/bin/env python3
"""
Extrair dados dos anuncios_*.xls e preencher a planilha final com:
- Id
- Integracao (extraido dos dados)
- Identificacao (ID Tiny)
- Titulo
- Produto (SKU Tiny)
- Preco de custo
- Preco de venda (Cadastro Tiny)
- Preco prom (ML + 0.01)
- Status
"""
import os
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv
import requests

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("EXTRAIR DADOS DOS ANUNCIOS E GERAR PLANILHA FINAL")
print("="*100)

# 1. LISTAR ARQUIVOS ANUNCIOS
print("\n[1] Procurando arquivos anuncios_*.xls...")

downloads_path = Path(r"C:\Users\user\Downloads")
anuncios_files = sorted(downloads_path.glob("anuncios_2026-07-26-*.xls"))

print(f"    Encontrados {len(anuncios_files)} arquivos")

# 2. EXTRAIR DADOS DOS ANUNCIOS
print("\n[2] Extraindo dados dos anuncios...")

dados_anuncios = {}  # {id_tiny: {integracao, titulo, sku, ...}}

for file_path in anuncios_files:
    print(f"    Lendo: {file_path.name}...")

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Procurar headers
        headers_arquivo = {}

        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value

            if header:
                headers_arquivo[header] = col

        # Processar linhas
        for row in range(2, ws.max_row + 1):
            # Procurar ID do Tiny (pode estar em coluna "Id", "Identificacao", etc)
            id_tiny = None

            for col_name in ["Id", "ID", "Identificacao", "Id produto Tiny", "ID Tiny"]:
                if col_name in headers_arquivo:
                    id_tiny = ws.cell(row, headers_arquivo[col_name]).value

                    if id_tiny:
                        try:
                            id_tiny = int(id_tiny)
                            break
                        except:
                            id_tiny = None

            if not id_tiny:
                continue

            # Extrair dados
            integracao = None
            titulo = None
            sku = None

            # Integracao
            for col_name in ["Integracao", "Canal", "Marketplace", "Plataforma"]:
                if col_name in headers_arquivo:
                    integracao = ws.cell(row, headers_arquivo[col_name]).value
                    if integracao:
                        break

            # Titulo
            for col_name in ["Titulo", "Title", "TITLE", "Nome do anuncio"]:
                if col_name in headers_arquivo:
                    titulo = ws.cell(row, headers_arquivo[col_name]).value
                    if titulo:
                        break

            # SKU
            for col_name in ["Produto", "SKU", "Sku", "PRODUCT_NUMBER"]:
                if col_name in headers_arquivo:
                    sku = ws.cell(row, headers_arquivo[col_name]).value
                    if sku:
                        break

            # Guardar
            if id_tiny not in dados_anuncios:
                dados_anuncios[id_tiny] = {
                    'integracao': integracao,
                    'titulo': titulo,
                    'sku': sku
                }

        wb.close()

    except Exception as e:
        print(f"      Erro ao ler {file_path.name}: {str(e)[:50]}")

print(f"    Extraidos dados de {len(dados_anuncios)} produtos")

# 3. LER RELATORIO ORIGINAL
print("\n[3] Lendo relatorio original...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb_orig = openpyxl.load_workbook(file_path)
ws_orig = wb_orig.active

# 4. CARREGAR DADOS DO TINY
print("\n[4] Carregando dados do Tiny...")

produtos_tiny = {}
offset = 0

while True:
    url_produtos = f"https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset={offset}"
    resp = requests.get(url_produtos, headers=headers, timeout=30)

    itens = resp.json().get("itens", [])

    if not itens:
        break

    for prod in itens:
        id_prod = prod.get("id")
        produtos_tiny[id_prod] = {
            'sku': prod.get("sku"),
            'descricao': prod.get("descricao"),
            'preco': prod.get("precos", {}).get("preco")
        }

    offset += 100

print(f"    Carregados {len(produtos_tiny)} produtos")

# 5. CRIAR PLANILHA FINAL
print("\n[5] Criando planilha final...")

wb_new = openpyxl.Workbook()
ws_new = wb_new.active

# Headers
headers_corretos = [
    "Id",
    "Integracao",
    "Identificacao",
    "Titulo",
    "Produto",
    "Preco de custo",
    "Preco de venda",
    "Preco prom",
    "Status"
]

# Escrever headers
style_header = Font(bold=True, color="FFFFFF", size=11)
fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
alignment_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, header in enumerate(headers_corretos, 1):
    cell = ws_new.cell(1, col_idx)
    cell.value = header
    cell.font = style_header
    cell.fill = fill_header
    cell.alignment = alignment_header

# Preencher dados
linha_saida = 2
row_count = 0
total_ok = 0
total_erro = 0

MAPA_INTEGRACAO_COL = {
    8: "PDV",
    9: "Shopee",
    10: "Amazon",
    11: "TikTok"
}

# Processar cada linha
for row_orig in range(2, ws_orig.max_row + 1):
    titulo_ml = ws_orig.cell(row_orig, 2).value
    id_tiny = ws_orig.cell(row_orig, 3).value
    preco_ml = ws_orig.cell(row_orig, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    # Procurar vermelhos
    for col_idx in [8, 9, 10, 11]:
        cell = ws_orig.cell(row_orig, col_idx)

        tem_vermelho = False

        if cell.fill and cell.fill.start_color:
            color_str = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color_str:
                tem_vermelho = True

        if not tem_vermelho:
            continue

        # TEM VERMELHO
        row_count += 1
        integracao_col = MAPA_INTEGRACAO_COL[col_idx]

        # Dados
        ws_new.cell(linha_saida, 1).value = row_count

        # Integracao: tentar extrair dos anuncios, se nao tiver usar a coluna
        integracao = None
        if id_tiny in dados_anuncios and dados_anuncios[id_tiny].get('integracao'):
            integracao = dados_anuncios[id_tiny]['integracao']
        if not integracao:
            integracao = integracao_col

        ws_new.cell(linha_saida, 2).value = integracao

        ws_new.cell(linha_saida, 3).value = id_tiny

        # Titulo: tentar extrair dos anuncios, se nao tiver usar do relatorio
        titulo = None
        if id_tiny in dados_anuncios and dados_anuncios[id_tiny].get('titulo'):
            titulo = dados_anuncios[id_tiny]['titulo']
        if not titulo:
            titulo = titulo_ml

        ws_new.cell(linha_saida, 4).value = titulo

        # SKU: extrair do Tiny
        sku_tiny = produtos_tiny.get(id_tiny, {}).get('sku')
        ws_new.cell(linha_saida, 5).value = sku_tiny

        # Preco de custo (vazio)
        ws_new.cell(linha_saida, 6).value = None

        # Preco de venda (Cadastro Tiny)
        preco_cadastro = ws_orig.cell(row_orig, 7).value
        ws_new.cell(linha_saida, 7).value = preco_cadastro

        # Preco prom (ML + 0.01)
        preco_esperado = round(preco_ml + 0.01, 2)
        ws_new.cell(linha_saida, 8).value = preco_esperado

        # Status
        if id_tiny not in produtos_tiny:
            status = "ERRO 404"
            total_erro += 1

            fill_red = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            font_white = Font(color="FFFFFF", bold=True)

            for col in range(1, 10):
                ws_new.cell(linha_saida, col).fill = fill_red
                ws_new.cell(linha_saida, col).font = font_white

        else:
            preco_atual = produtos_tiny[id_tiny]['preco']

            if abs(float(preco_atual) - preco_esperado) < 0.01:
                status = "OK"
                total_ok += 1

                fill_green = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                font_white = Font(color="FFFFFF", bold=True)

                for col in range(1, 10):
                    ws_new.cell(linha_saida, col).fill = fill_green
                    ws_new.cell(linha_saida, col).font = font_white

            else:
                status = "ATENCAO"
                total_erro += 1

                fill_orange = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                font_white = Font(color="FFFFFF", bold=True)

                for col in range(1, 10):
                    ws_new.cell(linha_saida, col).fill = fill_orange
                    ws_new.cell(linha_saida, col).font = font_white

        ws_new.cell(linha_saida, 9).value = status

        # Formatacao
        for col in [6, 7, 8]:
            cell = ws_new.cell(linha_saida, col)
            if cell.value is not None:
                cell.number_format = 'R$ #,##0.00'

        # Alinhamento
        ws_new.cell(linha_saida, 1).alignment = Alignment(horizontal="center")
        for col in range(2, 10):
            ws_new.cell(linha_saida, col).alignment = Alignment(horizontal="left", vertical="center")

        linha_saida += 1

# Ajustar larguras
ws_new.column_dimensions['A'].width = 8
ws_new.column_dimensions['B'].width = 20
ws_new.column_dimensions['C'].width = 16
ws_new.column_dimensions['D'].width = 50
ws_new.column_dimensions['E'].width = 18
ws_new.column_dimensions['F'].width = 16
ws_new.column_dimensions['G'].width = 18
ws_new.column_dimensions['H'].width = 18
ws_new.column_dimensions['I'].width = 18

# Congelar
ws_new.freeze_panes = "A2"

# Salvar
output_path = Path(r"C:\Users\user\Downloads\PLANILHA_FINAL_IMPORTACAO_TINY.xlsx")
wb_new.save(output_path)

print(f"\n✅ Planilha salva: {output_path.name}")

print(f"\n" + "="*100)
print("RESUMO FINAL")
print("="*100)

print(f"\n  Total de linhas com vermelho: {row_count}")
print(f"  Atualizado com sucesso (OK): {total_ok}")
print(f"  Com atencao/erro: {total_erro}")

if row_count > 0:
    print(f"  Taxa de sucesso: {100*total_ok/row_count:.1f}%")

print(f"\n✅ PLANILHA PRONTA PARA IMPORTAR NO TINY!")
