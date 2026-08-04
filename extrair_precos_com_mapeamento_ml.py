#!/usr/bin/env python3
"""
EXTRAIR PRECOS: Usar DESCRICAO do arquivo de Mapeamento ML (Anuncios)
para buscar nas tabelas do Tiny
"""
import os
import time
import requests
import openpyxl
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from difflib import SequenceMatcher

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n╔════════════════════════════════════════════════════════════════════════════════╗")
print("║  EXTRAIR PRECOS: Mapear pela DESCRICAO do arquivo ML (Anuncios)               ║")
print("╚════════════════════════════════════════════════════════════════════════════════╝")

# IDs das tabelas
TABELAS = {"PDV": 982, "Shopee": 989, "Amazon": 991, "TikTok": 990}

# 1. CARREGAR MAPEAMENTO ML (descricoes dos anuncios)
print("\n[1] Carregando descricoes do Mapeamento ML...")

mapping_desc = {}  # Linha ML → descricao

mapping_files = [
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-48.xls',
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-50.xls',
]

for mfile in mapping_files:
    if Path(mfile).exists():
        try:
            df = pd.read_excel(mfile, sheet_name=0)
            for idx, row in df.iterrows():
                titulo = row.get('Título')
                if titulo:
                    mapping_desc[titulo] = titulo
            print(f"    {Path(mfile).name}: {len(df)} anuncios")
        except Exception as e:
            print(f"    Erro: {e}")

print(f"    Total: {len(mapping_desc)} titulos carregados")

# 2. CARREGAR TODAS AS EXCECOES DAS TABELAS
print("\n[2] Carregando excecoes das tabelas...")

todas_excecoes = {}  # (tabela, codigo) → preco

for nome_tabela, id_tabela in TABELAS.items():
    print(f"    {nome_tabela}...", end=" ", flush=True)

    url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}"

    try:
        resp = requests.get(url, headers=headers, timeout=20)

        if resp.status_code == 200:
            data = resp.json()
            excecoes = data.get("excecoes", [])

            for exc in excecoes:
                codigo = str(exc.get("codigo", "")).upper()
                preco = exc.get("preco")

                todas_excecoes[(nome_tabela, codigo)] = preco

            print(f"{len(excecoes)}")
        else:
            print(f"Erro {resp.status_code}")

    except Exception as e:
        print(f"Erro")

    time.sleep(0.5)

print(f"    Total de excecoes: {len(todas_excecoes)}")

# 3. ABRIR RELATORIO
print("\n[3] Abrindo relatorio...")
file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_FINAL.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"    {file_path.name} ({ws.max_row - 1} linhas)")

# 4. PARA CADA LINHA, BUSCAR DESCRICAO E PRECO
print("\n[4] Buscando precos pela descricao...")

stats = {"processado": 0, "encontrado": 0}

for row_idx in range(2, ws.max_row + 1):
    titulo_ml = ws.cell(row_idx, 2).value  # Coluna B: Título ML
    sku_tiny = ws.cell(row_idx, 4).value   # Coluna D: SKU Tiny

    if not titulo_ml and not sku_tiny:
        continue

    stats["processado"] += 1

    # Extrair palavras-chave da descricao ML
    desc_search = str(titulo_ml).upper() if titulo_ml else str(sku_tiny).upper()
    palavras = [p.strip() for p in desc_search.split() if len(p) > 3]

    # Buscar em cada tabela
    precos_encontrados = {}

    for nome_tabela in TABELAS.keys():
        melhor_match = None
        melhor_score = 0

        # Procurar nos codigos (SKUs) das excecoes
        for (tab, codigo), preco in todas_excecoes.items():
            if tab != nome_tabela:
                continue

            # Contar matches
            matches = sum(1 for palavra in palavras if palavra in codigo)

            if matches > 0 and matches > melhor_score:
                melhor_score = matches
                melhor_match = preco

        if melhor_match:
            precos_encontrados[nome_tabela] = melhor_match
            stats["encontrado"] += 1

    # Preencher precos
    if precos_encontrados:
        ws.cell(row_idx, 8, precos_encontrados.get("PDV"))        # H
        ws.cell(row_idx, 9, precos_encontrados.get("Shopee"))     # I
        ws.cell(row_idx, 10, precos_encontrados.get("Amazon"))    # J
        ws.cell(row_idx, 11, precos_encontrados.get("TikTok"))    # K

        for col in [8, 9, 10, 11]:
            if ws.cell(row_idx, col).value:
                ws.cell(row_idx, col).number_format = 'R$ #,##0.00'

    if stats["processado"] % 20 == 0:
        desc_curta = str(titulo_ml)[:35] if titulo_ml else str(sku_tiny)[:35]
        print(f"    [{stats['processado']:2d}] {desc_curta}: {len(precos_encontrados)} precos")

# 5. SALVAR
print("\n[5] Salvando...")
wb.save(file_path)
wb.close()

print(f"\n{'='*100}")
print(f"SUCESSO!")
print(f"{'='*100}")
print(f"\n[RESULTADO]")
print(f"  Total processado: {stats['processado']}")
print(f"  Com precos encontrados: {stats['encontrado']}")
print(f"\n[ARQUIVO]")
print(f"  {file_path}")
print(f"  Pronto para download!")
