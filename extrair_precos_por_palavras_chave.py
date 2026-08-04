#!/usr/bin/env python3
"""
EXTRAIR PREÇOS: Buscar por PALAVRAS-CHAVE da descrição
Exemplo: "Assento Almofadado Thema" busca "ASSENTO" + "THEMA"
"""
import os
import time
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n╔════════════════════════════════════════════════════════════════════════════════╗")
print("║  EXTRAIR PRECOS: Buscar por PALAVRAS-CHAVE (flexivel)                         ║")
print("╚════════════════════════════════════════════════════════════════════════════════╝")

# IDs das tabelas
TABELAS = {"PDV": 982, "Shopee": 989, "Amazon": 991, "TikTok": 990}

# 1. CARREGAR TODAS AS EXCECOES
print("\n[1] Carregando todas as excecoes...")

todas_excecoes_por_tabela = {}

for nome_tabela, id_tabela in TABELAS.items():
    print(f"    {nome_tabela}...", end=" ", flush=True)

    url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}"

    try:
        resp = requests.get(url, headers=headers, timeout=20)

        if resp.status_code == 200:
            data = resp.json()
            excecoes = data.get("excecoes", [])

            todas_excecoes_por_tabela[nome_tabela] = excecoes

            print(f"{len(excecoes)} excecoes")
        else:
            print(f"Erro {resp.status_code}")

    except Exception as e:
        print(f"Erro: {str(e)[:30]}")

    time.sleep(0.5)

print(f"\n    Total carregado:")
for tabela, excecoes in todas_excecoes_por_tabela.items():
    print(f"      {tabela}: {len(excecoes)}")

# 2. FUNCAO PARA BUSCAR POR PALAVRAS-CHAVE
def buscar_preco_por_palavras_chave(descricao):
    """
    Busca na descricao por palavras-chave (maiuscula, sem acentos)
    Retorna: {PDV: preco, Shopee: preco, Amazon: preco, TikTok: preco}
    """
    if not descricao:
        return {}

    # Extrair palavras-chave (palavras > 3 caracteres)
    desc_upper = str(descricao).upper()
    palavras = [p.strip() for p in desc_upper.split() if len(p) > 3]

    precos = {}

    for tabela_nome, excecoes in todas_excecoes_por_tabela.items():
        melhor_match = None
        melhor_score = 0

        for exc in excecoes:
            codigo = str(exc.get("codigo", "")).upper()

            # Contar quantas palavras-chave estao no codigo
            matches = 0
            for palavra in palavras:
                if palavra in codigo:
                    matches += 1

            # Se tem pelo menos 1 match, eh candidato
            if matches > 0 and matches > melhor_score:
                melhor_score = matches
                melhor_match = exc.get("preco")

        if melhor_match:
            precos[tabela_nome] = melhor_match

    return precos

# 3. ABRIR RELATORIO
print("\n[2] Abrindo relatorio...")
file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COMPLETO_ATUALIZADO.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"    {file_path.name} ({ws.max_row - 1} linhas)")

# 4. PROCESSAR CADA PRODUTO
print("\n[3] Buscando precos por palavras-chave...")

stats = {"processado": 0, "encontrado": 0, "com_preco": 0}

for row_idx in range(2, ws.max_row + 1):
    desc = ws.cell(row_idx, 5).value  # Coluna E: Descricao Tiny

    if not desc:
        continue

    stats["processado"] += 1

    # Buscar precos
    precos_encontrados = buscar_preco_por_palavras_chave(desc)

    if precos_encontrados:
        stats["encontrado"] += 1

        # Preencher nas colunas
        ws.cell(row_idx, 8, precos_encontrados.get("PDV"))        # Col H: PDV
        ws.cell(row_idx, 9, precos_encontrados.get("Shopee"))     # Col I: Shopee
        ws.cell(row_idx, 10, precos_encontrados.get("Amazon"))    # Col J: Amazon
        ws.cell(row_idx, 11, precos_encontrados.get("TikTok"))    # Col K: TikTok

        # Formatar como moeda
        for col in [8, 9, 10, 11]:
            if ws.cell(row_idx, col).value:
                ws.cell(row_idx, col).number_format = 'R$ #,##0.00'
                stats["com_preco"] += 1

    if stats["processado"] % 20 == 0:
        desc_curta = str(desc)[:35]
        print(f"    [{stats['processado']:2d}] {desc_curta}: {len(precos_encontrados)} precos")

# 5. SALVAR
print("\n[4] Salvando...")
wb.save(file_path)
wb.close()

print(f"\n{'='*100}")
print(f"SUCESSO!")
print(f"{'='*100}")
print(f"\n[ESTATISTICAS]")
print(f"  Total processado: {stats['processado']}")
print(f"  Encontrados: {stats['encontrado']}")
print(f"  Precos adicionados: {stats['com_preco']}")
print(f"\n[ARQUIVO]")
print(f"  {file_path}")
print(f"  Pronto para download!")
