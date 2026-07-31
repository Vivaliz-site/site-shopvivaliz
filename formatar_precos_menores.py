#!/usr/bin/env python3
"""
FORMATAR: Pintar em VERMELHO as celulas onde preco < preco_ml
"""
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill

print("\nFormatando precos menores que ML...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_FINAL_6_PRECOS.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Cores
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
red_font = openpyxl.styles.Font(color="FFFFFF", bold=True)

count_formatted = 0

for row_idx in range(2, ws.max_row + 1):
    preco_ml = ws.cell(row_idx, 6).value  # Coluna F: Preco ML

    if not preco_ml:
        continue

    try:
        preco_ml = float(preco_ml)
    except:
        continue

    # Verificar cada coluna de preco (Cadastro, PDV, Shopee, Amazon, TikTok)
    for col_idx in [7, 8, 9, 10, 11]:  # G, H, I, J, K
        preco = ws.cell(row_idx, col_idx).value

        if not preco:
            continue

        try:
            preco = float(preco)
        except:
            continue

        # Se preco < preco_ml, pintar de vermelho
        if preco < preco_ml:
            cell = ws.cell(row_idx, col_idx)
            cell.fill = red_fill
            cell.font = red_font
            count_formatted += 1

print(f"  Formatadas {count_formatted} celulas em vermelho")

# Salvar
print(f"\nSalvando...")
wb.save(file_path)
wb.close()

print(f"Pronto: {file_path.name}")
