#!/usr/bin/env python3
"""
Gerar planilha fazendo MATCH por TITULO
- Relatorio ID Tiny + Titulo -> procurar no arquivo anuncios por titulo similar
- Extrair: Identificador (ID anuncio), Integracao, SKU
"""
import os
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv
import requests
import difflib

try:
    import xlrd
except:
    os.system("pip install xlrd -q")
    import xlrd

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("GERAR PLANILHA COM MATCH POR TITULO")
print("="*100)

# 1. EXTRAIR TODOS OS ANUNCIOS
print("\n[1] Extraindo dados dos anuncios...")

downloads_path = Path(r"C:\Users\user\Downloads")
anuncios_files = sorted(downloads_path.glob("anuncios_2026-07-26-*.xls"))

todos_anuncios = []

for file_path in anuncios_files:
    try:
        workbook = xlrd.open_workbook(str(file_path))
        worksheet = workbook.sheet_by_index(0)

        for row in range(1, worksheet.nrows):
            id_anuncio = worksheet.cell_value(row, 0)
            integracao = worksheet.cell_value(row, 1)
            identificador = worksheet.cell_value(row, 2)
            titulo = worksheet.cell_value(row, 3)
            sku = worksheet.cell_value(row, 4)

            todos_anuncios.append({
                'id_anuncio': id_anuncio,
                'integracao': integracao,
                'identificador': identificador,
                'titulo': str(titulo) if titulo else '',
                'sku': sku
            })

    except Exception as e:
        if "Workbook corruption" not in str(e):
            pass

print(f"    Extraidos {len(todos_anuncios)} anuncios")

# 2. LER RELATORIO
print("\n[2] Lendo relatorio...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb_orig = openpyxl.load_workbook(file_path)
ws_orig = wb_orig.active

# 3. CARREGAR DADOS DO TINY
print("\n[3] Carregando dados do Tiny...")

produtos_tiny = {}
offset = 0

while True:
    url_produtos = f"https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset={offset}"
    resp = requests.get(url_produtos, headers=headers, timeout=30)

    itens = resp.json().get("itens", [])

    if not itens:
        break

    for prod in itens:
        id_prod = prod.get("id")
        produtos_tiny[id_prod] = {
            'sku': prod.get("sku"),
            'descricao': prod.get("descricao"),
            'preco': prod.get("precos", {}).get("preco")
        }

    offset += 100

print(f"    Carregados {len(produtos_tiny)} produtos")

# 4. CRIAR PLANILHA
print("\n[4] Criando planilha...")

wb_new = openpyxl.Workbook()
ws_new = wb_new.active

headers_corretos = [
    "Id",
    "Integração",
    "Identificador",
    "Título",
    "Produto (SKU)",
    "Preço de custo",
    "Preço",
    "Preço promocional"
]

# Escrever headers
style_header = Font(bold=True, color="FFFFFF", size=11)
fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

for col_idx, header in enumerate(headers_corretos, 1):
    cell = ws_new.cell(1, col_idx)
    cell.value = header
    cell.font = style_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Preencher dados
linha_saida = 2
row_count = 0

MAPA_COL = {
    8: "PDV",
    9: "Shopee",
    10: "Amazon",
    11: "TikTok"
}

# Processar relatorio
for row_orig in range(2, ws_orig.max_row + 1):
    titulo_ml = ws_orig.cell(row_orig, 2).value
    id_tiny = ws_orig.cell(row_orig, 3).value
    preco_ml = ws_orig.cell(row_orig, 6).value

    if not id_tiny or not preco_ml or not titulo_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
        titulo_ml_str = str(titulo_ml)
    except:
        continue

    # Identificar colunas com vermelho
    colunas_com_vermelho = []

    for col_idx in [8, 9, 10, 11]:
        cell = ws_orig.cell(row_orig, col_idx)

        if cell.fill and cell.fill.start_color:
            color_str = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color_str:
                colunas_com_vermelho.append(col_idx)

    if not colunas_com_vermelho:
        continue

    preco_esperado = round(preco_ml + 0.01, 2)
    preco_cadastro = ws_orig.cell(row_orig, 7).value
    sku_tiny = produtos_tiny.get(id_tiny, {}).get('sku')

    # Para cada coluna com vermelho
    for col_idx in colunas_com_vermelho:
        integracao_esperada = MAPA_COL[col_idx]

        # Procurar anuncio por TITULO + INTEGRACAO
        melhor_match = None
        melhor_score = 0

        for anuncio in todos_anuncios:
            if anuncio['integracao'] and str(anuncio['integracao']).strip().lower() == integracao_esperada.lower():
                # Comparar titulos
                score = difflib.SequenceMatcher(None, titulo_ml_str.lower(), anuncio['titulo'].lower()).ratio()

                if score > melhor_score:
                    melhor_score = score
                    melhor_match = anuncio

        if melhor_match and melhor_score > 0.5:  # 50% similaridade minima
            row_count += 1

            ws_new.cell(linha_saida, 1).value = row_count
            ws_new.cell(linha_saida, 2).value = integracao_esperada
            ws_new.cell(linha_saida, 3).value = melhor_match['identificador']
            ws_new.cell(linha_saida, 4).value = melhor_match['titulo']
            ws_new.cell(linha_saida, 5).value = melhor_match['sku'] or sku_tiny
            ws_new.cell(linha_saida, 6).value = None
            ws_new.cell(linha_saida, 7).value = preco_cadastro
            ws_new.cell(linha_saida, 8).value = preco_esperado

            # Cor
            if id_tiny not in produtos_tiny:
                fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                font = Font(color="FFFFFF", bold=True)
            elif abs(float(produtos_tiny[id_tiny]['preco']) - preco_esperado) < 0.01:
                fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                font = Font(color="FFFFFF", bold=True)
            else:
                fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                font = Font(color="FFFFFF", bold=True)

            for col in range(1, 9):
                ws_new.cell(linha_saida, col).fill = fill
                ws_new.cell(linha_saida, col).font = font

            # Formato
            for col in [6, 7, 8]:
                ws_new.cell(linha_saida, col).number_format = 'R$ #,##0.00'

            linha_saida += 1

        else:
            # Sem match
            row_count += 1

            ws_new.cell(linha_saida, 1).value = row_count
            ws_new.cell(linha_saida, 2).value = integracao_esperada
            ws_new.cell(linha_saida, 3).value = None
            ws_new.cell(linha_saida, 4).value = titulo_ml_str
            ws_new.cell(linha_saida, 5).value = sku_tiny
            ws_new.cell(linha_saida, 6).value = None
            ws_new.cell(linha_saida, 7).value = preco_cadastro
            ws_new.cell(linha_saida, 8).value = preco_esperado

            fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            font = Font(color="FFFFFF", bold=True)

            for col in range(1, 9):
                ws_new.cell(linha_saida, col).fill = fill
                ws_new.cell(linha_saida, col).font = font

            linha_saida += 1

# Ajustar larguras
ws_new.column_dimensions['A'].width = 8
ws_new.column_dimensions['B'].width = 20
ws_new.column_dimensions['C'].width = 25
ws_new.column_dimensions['D'].width = 45
ws_new.column_dimensions['E'].width = 18
ws_new.column_dimensions['F'].width = 16
ws_new.column_dimensions['G'].width = 16
ws_new.column_dimensions['H'].width = 18

ws_new.freeze_panes = "A2"

# Salvar
output_path = Path(r"C:\Users\user\Downloads\PLANILHA_COM_IDENTIFICADORES.xlsx")
wb_new.save(output_path)

print(f"\n✅ Planilha salva: {output_path.name}")

print(f"\n" + "="*100)
print("RESUMO")
print("="*100)

print(f"\n  Total de linhas: {row_count}")
print(f"  Arquivo: {output_path}")
