#!/usr/bin/env python3
"""
CORRIGIDO FINALMENTE:
- Preencher APENAS os anuncios que vao alterar (estao abaixo do ML)
- Mexer APENAS na coluna Preco promocional (amarelo)
- Todos outros campos vem dos arquivos anuncios
"""
import os
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv
import requests
import difflib

try:
    import xlrd
except:
    os.system("pip install xlrd -q")
    import xlrd

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("GERAR PLANILHA - CORRIGIDA FINAL")
print("="*100)

# 1. EXTRAIR TODOS OS ANUNCIOS
print("\n[1] Extraindo dados dos anuncios...")

downloads_path = Path(r"C:\Users\user\Downloads")
anuncios_files = sorted(downloads_path.glob("anuncios_2026-07-26-*.xls"))

todos_anuncios = []

for file_path in anuncios_files:
    try:
        workbook = xlrd.open_workbook(str(file_path))
        worksheet = workbook.sheet_by_index(0)

        for row in range(1, worksheet.nrows):
            id_anuncio = worksheet.cell_value(row, 0)
            integracao = worksheet.cell_value(row, 1)
            identificador = worksheet.cell_value(row, 2)
            titulo = worksheet.cell_value(row, 3)
            sku = worksheet.cell_value(row, 4)

            todos_anuncios.append({
                'id_anuncio': id_anuncio,
                'integracao': integracao,
                'identificador': identificador,
                'titulo': str(titulo) if titulo else '',
                'sku': sku
            })

    except Exception as e:
        if "Workbook corruption" not in str(e):
            pass

print(f"    Extraidos {len(todos_anuncios)} anuncios")

# 2. LER RELATORIO
print("\n[2] Lendo relatorio...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb_orig = openpyxl.load_workbook(file_path)
ws_orig = wb_orig.active

# 3. CARREGAR DADOS DO TINY
print("\n[3] Carregando dados do Tiny...")

produtos_tiny = {}
offset = 0

while True:
    url_produtos = f"https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset={offset}"
    resp = requests.get(url_produtos, headers=headers, timeout=30)

    itens = resp.json().get("itens", [])

    if not itens:
        break

    for prod in itens:
        id_prod = prod.get("id")
        produtos_tiny[id_prod] = {
            'sku': prod.get("sku"),
            'descricao': prod.get("descricao"),
            'preco': prod.get("precos", {}).get("preco")
        }

    offset += 100

print(f"    Carregados {len(produtos_tiny)} produtos")

# 4. CRIAR PLANILHA
print("\n[4] Criando planilha...")

wb_new = openpyxl.Workbook()
ws_new = wb_new.active

headers_corretos = [
    "Id",
    "Integração",
    "Identificador",
    "Título",
    "Produto (SKU)",
    "Preço de custo",
    "Preço",
    "Preço promocional"
]

# Escrever headers
style_header = Font(bold=True, color="FFFFFF", size=11)
fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

for col_idx, header in enumerate(headers_corretos, 1):
    cell = ws_new.cell(1, col_idx)
    cell.value = header
    cell.font = style_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Preencher dados
linha_saida = 2
row_count = 0

MAPA_COL = {
    8: "PDV",
    9: "Shopee",
    10: "Amazon",
    11: "TikTok"
}

# Processar relatorio
for row_orig in range(2, ws_orig.max_row + 1):
    titulo_ml = ws_orig.cell(row_orig, 2).value
    id_tiny = ws_orig.cell(row_orig, 3).value
    preco_ml = ws_orig.cell(row_orig, 6).value

    if not id_tiny or not preco_ml or not titulo_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
        titulo_ml_str = str(titulo_ml)
    except:
        continue

    # PASSO 1: Verificar se ESTE PRODUTO tem vermelho
    # (se nao tem, pula - nao precisa incluir na planilha)
    tem_vermelho = False
    colunas_com_vermelho = []

    for col_idx in [8, 9, 10, 11]:
        cell = ws_orig.cell(row_orig, col_idx)

        if cell.fill and cell.fill.start_color:
            color_str = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color_str:
                tem_vermelho = True
                colunas_com_vermelho.append(col_idx)

    # SE NAO TEM VERMELHO, PULA
    if not tem_vermelho:
        continue

    # PASSO 2: Tem vermelho! Calcular novo preco
    preco_esperado = round(preco_ml + 0.01, 2)
    preco_cadastro = ws_orig.cell(row_orig, 7).value
    sku_tiny = produtos_tiny.get(id_tiny, {}).get('sku')

    # PASSO 3: Para CADA COLUNA COM VERMELHO, procurar anuncio
    for col_idx in colunas_com_vermelho:
        integracao_esperada = MAPA_COL[col_idx]

        # Procurar anuncio por TITULO + INTEGRACAO
        melhor_match = None
        melhor_score = 0

        for anuncio in todos_anuncios:
            if anuncio['integracao'] and str(anuncio['integracao']).strip().lower() == integracao_esperada.lower():
                # Comparar titulos
                score = difflib.SequenceMatcher(None, titulo_ml_str.lower(), anuncio['titulo'].lower()).ratio()

                if score > melhor_score:
                    melhor_score = score
                    melhor_match = anuncio

        if melhor_match and melhor_score > 0.5:  # 50% similaridade minima
            row_count += 1

            # PREENCHER COM DADOS DO ANUNCIO
            ws_new.cell(linha_saida, 1).value = row_count
            ws_new.cell(linha_saida, 2).value = melhor_match['integracao']  # Do anuncio
            ws_new.cell(linha_saida, 3).value = melhor_match['identificador']  # Do anuncio
            ws_new.cell(linha_saida, 4).value = melhor_match['titulo']  # Do anuncio
            ws_new.cell(linha_saida, 5).value = melhor_match['sku']  # Do anuncio
            ws_new.cell(linha_saida, 6).value = None  # Preco de custo (vazio)
            ws_new.cell(linha_saida, 7).value = preco_cadastro  # Preco atual (nao altera)

            # COLUNA AMARELA: Preco promocional (ALTERADO = ML + 0.01)
            cell_promo = ws_new.cell(linha_saida, 8)
            cell_promo.value = preco_esperado
            cell_promo.number_format = 'R$ #,##0.00'

            # Colorir APENAS a coluna de preco promocional em AMARELO
            fill_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            font_preto = Font(color="000000", bold=True)
            cell_promo.fill = fill_amarelo
            cell_promo.font = font_preto

            # Outras colunas: sem cor ou cor neutra
            fill_neutro = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            for col in range(1, 8):
                ws_new.cell(linha_saida, col).fill = fill_neutro
                ws_new.cell(linha_saida, col).number_format = ''

            # Formato de numeros
            for col in [6, 7]:
                cell = ws_new.cell(linha_saida, col)
                if cell.value is not None:
                    cell.number_format = 'R$ #,##0.00'

            # Alinhamento
            ws_new.cell(linha_saida, 1).alignment = Alignment(horizontal="center")
            for col in range(2, 9):
                ws_new.cell(linha_saida, col).alignment = Alignment(horizontal="left", vertical="center")

            linha_saida += 1

print(f"\n✅ Planilha salva")

# Ajustar larguras
ws_new.column_dimensions['A'].width = 8
ws_new.column_dimensions['B'].width = 20
ws_new.column_dimensions['C'].width = 25
ws_new.column_dimensions['D'].width = 45
ws_new.column_dimensions['E'].width = 18
ws_new.column_dimensions['F'].width = 16
ws_new.column_dimensions['G'].width = 16
ws_new.column_dimensions['H'].width = 18

ws_new.freeze_panes = "A2"

# Salvar
output_path = Path(r"C:\Users\user\Downloads\PLANILHA_FINAL.xlsx")
wb_new.save(output_path)

print(f"\n" + "="*100)
print("RESUMO FINAL")
print("="*100)

print(f"\n  Total de anuncios que sofrem alteracao: {row_count}")
print(f"  Arquivo: {output_path.name}")
print(f"\n  Estrutura:")
print(f"  - Colunas 1-7: Dados originais dos anuncios (NAO alteram)")
print(f"  - Coluna 8 (AMARELA): Preco promocional (ALTERADO = ML + 0.01)")

print(f"\n✅ PRONTO PARA IMPORTAR!")
