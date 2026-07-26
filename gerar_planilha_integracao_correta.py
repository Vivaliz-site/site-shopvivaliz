#!/usr/bin/env python3
"""
Gerar PLANILHA CORRETA:
- Apenas as integrações que NÃO sao Mercado Livre (PDV, Shopee, Amazon, TikTok)
- Para cada linha com vermelho, criar uma entrada
- Usar preço ML como REFERENCIA apenas
"""
import os
import requests
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("GERAR PLANILHA COM INTEGRACAO CORRETA")
print("="*100)

# 1. LER RELATORIO
print("\n[1] Lendo relatorio original...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb_orig = openpyxl.load_workbook(file_path)
ws_orig = wb_orig.active

# 2. CARREGAR DADOS DO TINY
print("\n[2] Carregando dados do Tiny...")

produtos_tiny = {}
offset = 0

while True:
    url_produtos = f"https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset={offset}"
    resp = requests.get(url_produtos, headers=headers, timeout=30)

    itens = resp.json().get("itens", [])

    if not itens:
        break

    for prod in itens:
        id_prod = prod.get("id")
        produtos_tiny[id_prod] = {
            'sku': prod.get("sku"),
            'descricao': prod.get("descricao"),
            'preco': prod.get("precos", {}).get("preco")
        }

    offset += 100

print(f"    Carregados {len(produtos_tiny)} produtos")

# 3. MAPEAR COLUNAS VERMELHAS PARA INTEGRACAO
# Col 8 = PDV, Col 9 = Shopee, Col 10 = Amazon, Col 11 = TikTok
MAPA_INTEGRACAO = {
    8: "PDV",
    9: "Shopee",
    10: "Amazon",
    11: "TikTok"
}

# 4. CRIAR NOVO WORKBOOK
print("\n[3] Criando planilha...")

wb_new = openpyxl.Workbook()
ws_new = wb_new.active

# Headers
headers_corretos = [
    "Id",
    "Integracao",
    "Identificacao",
    "Titulo",
    "Produto",
    "Preco de custo",
    "Preco de venda",
    "Preco prom",
    "Status"
]

# Escrever headers
style_header = Font(bold=True, color="FFFFFF", size=11)
fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
alignment_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, header in enumerate(headers_corretos, 1):
    cell = ws_new.cell(1, col_idx)
    cell.value = header
    cell.font = style_header
    cell.fill = fill_header
    cell.alignment = alignment_header

# Preencher dados
linha_saida = 2
row_count = 0
total_ok = 0
total_erro = 0

# Processar cada linha do relatorio
for row_orig in range(2, ws_orig.max_row + 1):
    titulo_ml = ws_orig.cell(row_orig, 2).value
    id_tiny = ws_orig.cell(row_orig, 3).value
    preco_ml = ws_orig.cell(row_orig, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    # Procurar colunas com VERMELHO (PDV, Shopee, Amazon, TikTok)
    for col_idx in [8, 9, 10, 11]:
        cell = ws_orig.cell(row_orig, col_idx)

        tem_vermelho = False

        if cell.fill and cell.fill.start_color:
            color_str = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color_str:
                tem_vermelho = True

        if not tem_vermelho:
            continue

        # TEM VERMELHO! Criar entrada
        row_count += 1
        integracao = MAPA_INTEGRACAO[col_idx]

        # DADOS
        ws_new.cell(linha_saida, 1).value = row_count
        ws_new.cell(linha_saida, 2).value = integracao
        ws_new.cell(linha_saida, 3).value = id_tiny
        ws_new.cell(linha_saida, 4).value = titulo_ml
        ws_new.cell(linha_saida, 5).value = produtos_tiny.get(id_tiny, {}).get('sku')
        ws_new.cell(linha_saida, 6).value = None  # Preco de custo (vazio)

        # Preco de venda (Cadastro Tiny ATUAL)
        preco_cadastro = ws_orig.cell(row_orig, 7).value
        ws_new.cell(linha_saida, 7).value = preco_cadastro

        # Preco prom (ML + 0.01) - REFERENCIA DO PREÇO A APLICAR
        preco_esperado = round(preco_ml + 0.01, 2)
        ws_new.cell(linha_saida, 8).value = preco_esperado

        # STATUS
        if id_tiny not in produtos_tiny:
            status = "ERRO 404"
            total_erro += 1

            fill_red = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            font_white = Font(color="FFFFFF", bold=True)

            for col in range(1, 10):
                ws_new.cell(linha_saida, col).fill = fill_red
                ws_new.cell(linha_saida, col).font = font_white

        else:
            preco_atual = produtos_tiny[id_tiny]['preco']

            if abs(float(preco_atual) - preco_esperado) < 0.01:
                status = "OK"
                total_ok += 1

                fill_green = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                font_white = Font(color="FFFFFF", bold=True)

                for col in range(1, 10):
                    ws_new.cell(linha_saida, col).fill = fill_green
                    ws_new.cell(linha_saida, col).font = font_white

            else:
                status = f"ATENCAO"
                total_erro += 1

                fill_orange = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                font_white = Font(color="FFFFFF", bold=True)

                for col in range(1, 10):
                    ws_new.cell(linha_saida, col).fill = fill_orange
                    ws_new.cell(linha_saida, col).font = font_white

        ws_new.cell(linha_saida, 9).value = status

        # Formatacao de numeros
        for col in [6, 7, 8]:
            cell = ws_new.cell(linha_saida, col)
            if cell.value is not None:
                cell.number_format = 'R$ #,##0.00'

        # Alinhamento
        ws_new.cell(linha_saida, 1).alignment = Alignment(horizontal="center")
        for col in range(2, 10):
            ws_new.cell(linha_saida, col).alignment = Alignment(horizontal="left", vertical="center")

        linha_saida += 1

# Ajustar larguras
ws_new.column_dimensions['A'].width = 8
ws_new.column_dimensions['B'].width = 20
ws_new.column_dimensions['C'].width = 16
ws_new.column_dimensions['D'].width = 45
ws_new.column_dimensions['E'].width = 18
ws_new.column_dimensions['F'].width = 16
ws_new.column_dimensions['G'].width = 18
ws_new.column_dimensions['H'].width = 18
ws_new.column_dimensions['I'].width = 18

# Congelar primeira linha
ws_new.freeze_panes = "A2"

# Salvar
output_path = Path(r"C:\Users\user\Downloads\PLANILHA_IMPORTACAO_INTEGRACAO_2026-07-26.xlsx")
wb_new.save(output_path)

print(f"\n✅ Planilha salva: {output_path.name}")

print(f"\n" + "="*100)
print("RESUMO FINAL")
print("="*100)

print(f"\n  Total de linhas com vermelho: {row_count}")
print(f"  Atualizado com sucesso (OK): {total_ok}")
print(f"  Com atencao/erro: {total_erro}")

if row_count > 0:
    print(f"  Taxa de sucesso: {100*total_ok/row_count:.1f}%")

print(f"\n✅ PRONTO PARA IMPORTAR NO TINY!")
print(f"   Arquivo: {output_path.name}")
