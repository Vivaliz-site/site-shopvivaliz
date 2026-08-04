#!/usr/bin/env python3
"""
Gerar planilha com MULTIPLOS ANUNCIOS por produto
Se um produto tem 3 anúncios (ML, Shopee, Amazon), criar 3 linhas
"""
import os
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv
import requests

try:
    import xlrd
except:
    os.system("pip install xlrd -q")
    import xlrd

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("GERAR PLANILHA COM MULTIPLOS ANUNCIOS POR PRODUTO")
print("="*100)

# 1. EXTRAIR DADOS DOS ANUNCIOS
print("\n[1] Extraindo dados dos anuncios...")

downloads_path = Path(r"C:\Users\user\Downloads")
anuncios_files = sorted(downloads_path.glob("anuncios_2026-07-26-*.xls"))

# Usar estrutura: {id_tiny: [anuncios_list]}
dados_por_produto = {}

for file_path in anuncios_files:
    try:
        workbook = xlrd.open_workbook(str(file_path))
        worksheet = workbook.sheet_by_index(0)

        headers_arquivo = {}

        for col in range(worksheet.ncols):
            header = str(worksheet.cell_value(0, col))
            if header and header != 'None':
                headers_arquivo[header] = col

        # Processar linhas
        for row in range(1, worksheet.nrows):
            id_tiny = None

            for col_name in ["Id", "ID"]:
                if col_name in headers_arquivo:
                    try:
                        id_tiny = int(worksheet.cell_value(row, headers_arquivo[col_name]))
                        break
                    except:
                        pass

            if not id_tiny:
                continue

            # Extrair dados
            integracao = None
            titulo = None
            id_anuncio = None
            sku_produto = None

            # Integracao
            for col_name in ["Integração", "Integracao", "Canal"]:
                if col_name in headers_arquivo:
                    integracao = worksheet.cell_value(row, headers_arquivo[col_name])
                    if integracao and str(integracao).strip():
                        break

            # Titulo
            for col_name in ["Título", "Titulo", "Title", "TITLE"]:
                if col_name in headers_arquivo:
                    titulo = worksheet.cell_value(row, headers_arquivo[col_name])
                    if titulo and str(titulo).strip():
                        break

            # ID do anúncio
            for col_name in ["Identificador", "Identificacao", "ITEM_ID", "PRODUCT_NUMBER"]:
                if col_name in headers_arquivo:
                    id_anuncio = worksheet.cell_value(row, headers_arquivo[col_name])
                    if id_anuncio and str(id_anuncio).strip():
                        break

            # SKU
            for col_name in ["Produto (SKU)", "Produto", "SKU", "PRODUCT_TINY"]:
                if col_name in headers_arquivo:
                    sku_produto = worksheet.cell_value(row, headers_arquivo[col_name])
                    if sku_produto and str(sku_produto).strip():
                        break

            # Adicionar ao dict (pode ter múltiplos anúncios por produto)
            if id_tiny not in dados_por_produto:
                dados_por_produto[id_tiny] = []

            dados_por_produto[id_tiny].append({
                'integracao': integracao,
                'titulo': titulo,
                'id_anuncio': id_anuncio,
                'sku_produto': sku_produto
            })

    except Exception as e:
        if "Workbook corruption" not in str(e):
            pass

print(f"    Extraidos {len(dados_por_produto)} produtos com anuncios")
print(f"    Exemplo de produto com multiplos anuncios:")

for id_tiny, anuncios in list(dados_por_produto.items())[:3]:
    if len(anuncios) > 1:
        print(f"      ID {id_tiny}: {len(anuncios)} anuncios")
        for i, anuncio in enumerate(anuncios[:3], 1):
            print(f"        {i}. {anuncio['integracao']} - {anuncio['id_anuncio']}")

# 2. LER RELATORIO ORIGINAL
print("\n[2] Lendo relatorio original...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb_orig = openpyxl.load_workbook(file_path)
ws_orig = wb_orig.active

# 3. CARREGAR DADOS DO TINY
print("\n[3] Carregando dados do Tiny...")

produtos_tiny = {}
offset = 0

while True:
    url_produtos = f"https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset={offset}"
    resp = requests.get(url_produtos, headers=headers, timeout=30)

    itens = resp.json().get("itens", [])

    if not itens:
        break

    for prod in itens:
        id_prod = prod.get("id")
        produtos_tiny[id_prod] = {
            'sku': prod.get("sku"),
            'descricao': prod.get("descricao"),
            'preco': prod.get("precos", {}).get("preco")
        }

    offset += 100

print(f"    Carregados {len(produtos_tiny)} produtos")

# 4. CRIAR PLANILHA FINAL
print("\n[4] Criando planilha final...")

wb_new = openpyxl.Workbook()
ws_new = wb_new.active

# Headers
headers_corretos = [
    "Id",
    "Integracao",
    "Anuncio",
    "Titulo",
    "Produto",
    "Preco de custo",
    "Preco de venda",
    "Preco prom",
    "Status"
]

# Escrever headers
style_header = Font(bold=True, color="FFFFFF", size=11)
fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

for col_idx, header in enumerate(headers_corretos, 1):
    cell = ws_new.cell(1, col_idx)
    cell.value = header
    cell.font = style_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Preencher dados
linha_saida = 2
row_count = 0
total_ok = 0
total_erro = 0

MAPA_INTEGRACAO_COL = {
    8: "PDV",
    9: "Shopee",
    10: "Amazon",
    11: "TikTok"
}

# Processar cada linha DO RELATORIO
for row_orig in range(2, ws_orig.max_row + 1):
    titulo_ml = ws_orig.cell(row_orig, 2).value
    id_tiny = ws_orig.cell(row_orig, 3).value
    preco_ml = ws_orig.cell(row_orig, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    # Procurar vermelhos
    for col_idx in [8, 9, 10, 11]:
        cell = ws_orig.cell(row_orig, col_idx)

        tem_vermelho = False

        if cell.fill and cell.fill.start_color:
            color_str = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color_str:
                tem_vermelho = True

        if not tem_vermelho:
            continue

        # TEM VERMELHO - procurar anuncios deste produto
        integracao_esperada = MAPA_INTEGRACAO_COL[col_idx]

        anuncios_produto = dados_por_produto.get(id_tiny, [])

        if not anuncios_produto:
            # Produto sem anuncios registrados, criar uma linha mesmo assim
            row_count += 1

            ws_new.cell(linha_saida, 1).value = row_count
            ws_new.cell(linha_saida, 2).value = integracao_esperada
            ws_new.cell(linha_saida, 3).value = None
            ws_new.cell(linha_saida, 4).value = titulo_ml
            ws_new.cell(linha_saida, 5).value = produtos_tiny.get(id_tiny, {}).get('sku')
            ws_new.cell(linha_saida, 6).value = None
            ws_new.cell(linha_saida, 7).value = ws_orig.cell(row_orig, 7).value
            ws_new.cell(linha_saida, 8).value = round(preco_ml + 0.01, 2)
            ws_new.cell(linha_saida, 9).value = "SEM ANUNCIO"

            # Cor vermelha
            fill_red = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            font_white = Font(color="FFFFFF", bold=True)

            for col in range(1, 10):
                ws_new.cell(linha_saida, col).fill = fill_red
                ws_new.cell(linha_saida, col).font = font_white

            total_erro += 1
            linha_saida += 1

        else:
            # Produto com anuncios - criar UMA LINHA POR ANUNCIO
            for anuncio in anuncios_produto:
                row_count += 1

                ws_new.cell(linha_saida, 1).value = row_count
                ws_new.cell(linha_saida, 2).value = anuncio['integracao'] or integracao_esperada
                ws_new.cell(linha_saida, 3).value = anuncio['id_anuncio']
                ws_new.cell(linha_saida, 4).value = anuncio['titulo'] or titulo_ml
                ws_new.cell(linha_saida, 5).value = anuncio['sku_produto'] or produtos_tiny.get(id_tiny, {}).get('sku')
                ws_new.cell(linha_saida, 6).value = None
                ws_new.cell(linha_saida, 7).value = ws_orig.cell(row_orig, 7).value
                ws_new.cell(linha_saida, 8).value = round(preco_ml + 0.01, 2)

                # Status
                if id_tiny not in produtos_tiny:
                    status = "ERRO 404"
                    total_erro += 1

                    fill_red = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    font_white = Font(color="FFFFFF", bold=True)

                    for col in range(1, 10):
                        ws_new.cell(linha_saida, col).fill = fill_red
                        ws_new.cell(linha_saida, col).font = font_white

                else:
                    preco_atual = produtos_tiny[id_tiny]['preco']
                    preco_esperado = round(preco_ml + 0.01, 2)

                    if abs(float(preco_atual) - preco_esperado) < 0.01:
                        status = "OK"
                        total_ok += 1

                        fill_green = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                        font_white = Font(color="FFFFFF", bold=True)

                        for col in range(1, 10):
                            ws_new.cell(linha_saida, col).fill = fill_green
                            ws_new.cell(linha_saida, col).font = font_white

                    else:
                        status = "ATENCAO"
                        total_erro += 1

                        fill_orange = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                        font_white = Font(color="FFFFFF", bold=True)

                        for col in range(1, 10):
                            ws_new.cell(linha_saida, col).fill = fill_orange
                            ws_new.cell(linha_saida, col).font = font_white

                ws_new.cell(linha_saida, 9).value = status

                # Formatacao
                for col in [6, 7, 8]:
                    cell = ws_new.cell(linha_saida, col)
                    if cell.value is not None:
                        cell.number_format = 'R$ #,##0.00'

                # Alinhamento
                ws_new.cell(linha_saida, 1).alignment = Alignment(horizontal="center")
                for col in range(2, 10):
                    ws_new.cell(linha_saida, col).alignment = Alignment(horizontal="left", vertical="center")

                linha_saida += 1

# Ajustar larguras
ws_new.column_dimensions['A'].width = 8
ws_new.column_dimensions['B'].width = 20
ws_new.column_dimensions['C'].width = 25
ws_new.column_dimensions['D'].width = 50
ws_new.column_dimensions['E'].width = 20
ws_new.column_dimensions['F'].width = 16
ws_new.column_dimensions['G'].width = 18
ws_new.column_dimensions['H'].width = 18
ws_new.column_dimensions['I'].width = 18

ws_new.freeze_panes = "A2"

# Salvar
output_path = Path(r"C:\Users\user\Downloads\PLANILHA_MULTIPLOS_ANUNCIOS.xlsx")
wb_new.save(output_path)

print(f"\n✅ Planilha salva: {output_path.name}")

print(f"\n" + "="*100)
print("RESUMO FINAL")
print("="*100)

print(f"\n  Total de linhas geradas: {row_count}")
print(f"  (pode ter multiplos anuncios por produto)")
print(f"\n  Atualizado com sucesso (OK): {total_ok}")
print(f"  Com atencao/erro: {total_erro}")

if row_count > 0:
    print(f"  Taxa de sucesso: {100*total_ok/row_count:.1f}%")

print(f"\n✅ PLANILHA PRONTA PARA IMPORTAR!")
print(f"   - Cada linha é um anúncio diferente")
print(f"   - Um produto pode ter multiplas linhas (um por integração)")
