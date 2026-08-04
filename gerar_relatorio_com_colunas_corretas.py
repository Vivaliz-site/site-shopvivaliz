#!/usr/bin/env python3
"""
Gerar RELATORIO FINAL com EXATAMENTE estas colunas:
- Id
- Integracao
- Identificacao
- Titulo
- Produto (SKU)
- Preco de custo
- Preco de venda
- Preco prom
- Status
"""
import os
import requests
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("GERAR RELATORIO COM COLUNAS CORRETAS")
print("="*100)

# 1. LER RELATORIO ORIGINAL
print("\n[1] Lendo relatorio original...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb_orig = openpyxl.load_workbook(file_path)
ws_orig = wb_orig.active

# 2. CARREGAR DADOS ATUAIS DO TINY
print("\n[2] Carregando dados do Tiny...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

produtos_tiny = {}

for prod in resp.json().get("itens", []):
    id_prod = prod.get("id")
    produtos_tiny[id_prod] = {
        'sku': prod.get("sku"),
        'descricao': prod.get("descricao"),
        'preco': prod.get("precos", {}).get("preco")
    }

print(f"    Carregados {len(produtos_tiny)} produtos")

# 3. CRIAR NOVO WORKBOOK
print("\n[3] Criando relatorio...")

wb_new = openpyxl.Workbook()
ws_new = wb_new.active

# Headers - CONFORME SOLICITADO
headers_corretos = [
    "Id",                      # 1 - numerar sequencialmente
    "Integracao",              # 2 - "Mercado Livre"
    "Identificacao",           # 3 - VARIATION_ID ou PRODUCT_NUMBER do ML
    "Titulo",                  # 4 - Título do anúncio
    "Produto",                 # 5 - SKU do Tiny
    "Preco de custo",          # 6 - Preço de custo (se temos)
    "Preco de venda",          # 7 - Preço de venda no Tiny (Cadastro)
    "Preco prom",              # 8 - Preço promocional (ML + 0.01)
    "Status"                   # 9 - Status da atualização
]

# Escrever headers
style_header = Font(bold=True, color="FFFFFF", size=11)
fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
alignment_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, header in enumerate(headers_corretos, 1):
    cell = ws_new.cell(1, col_idx)
    cell.value = header
    cell.font = style_header
    cell.fill = fill_header
    cell.alignment = alignment_header

# Preencher dados
linha_saida = 2
row_count = 0
total_ok = 0
total_erro = 0

for row_orig in range(2, ws_orig.max_row + 1):
    titulo_ml = ws_orig.cell(row_orig, 2).value
    id_tiny = ws_orig.cell(row_orig, 3).value
    preco_ml = ws_orig.cell(row_orig, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    row_count += 1

    # DADOS
    ws_new.cell(linha_saida, 1).value = row_count  # Id sequencial
    ws_new.cell(linha_saida, 2).value = "Mercado Livre"  # Integracao
    ws_new.cell(linha_saida, 3).value = id_tiny  # Identificacao (ID no Tiny)
    ws_new.cell(linha_saida, 4).value = titulo_ml  # Titulo
    ws_new.cell(linha_saida, 5).value = produtos_tiny.get(id_tiny, {}).get('sku')  # Produto (SKU)

    # Preco de custo (deixar vazio, nao temos)
    ws_new.cell(linha_saida, 6).value = None

    # Preco de venda (Cadastro Tiny)
    preco_cadastro = ws_orig.cell(row_orig, 7).value
    ws_new.cell(linha_saida, 7).value = preco_cadastro

    # Preco prom (ML + 0.01)
    preco_esperado = round(preco_ml + 0.01, 2)
    ws_new.cell(linha_saida, 8).value = preco_esperado

    # STATUS
    if id_tiny not in produtos_tiny:
        status = "ERRO 404"
        total_erro += 1

        fill_red = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        font_white = Font(color="FFFFFF", bold=True)

        for col in range(1, 10):
            ws_new.cell(linha_saida, col).fill = fill_red
            ws_new.cell(linha_saida, col).font = font_white

    else:
        preco_atual = produtos_tiny[id_tiny]['preco']

        if abs(float(preco_atual) - preco_esperado) < 0.01:
            status = "OK - Atualizado"
            total_ok += 1

            fill_green = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            font_white = Font(color="FFFFFF", bold=True)

            for col in range(1, 10):
                ws_new.cell(linha_saida, col).fill = fill_green
                ws_new.cell(linha_saida, col).font = font_white

        else:
            status = f"ATENCAO - R$ {preco_atual}"
            total_erro += 1

            fill_orange = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            font_white = Font(color="FFFFFF", bold=True)

            for col in range(1, 10):
                ws_new.cell(linha_saida, col).fill = fill_orange
                ws_new.cell(linha_saida, col).font = font_white

    ws_new.cell(linha_saida, 9).value = status

    # Formatacao de numeros (precos)
    for col in [6, 7, 8]:
        cell = ws_new.cell(linha_saida, col)
        if cell.value is not None:
            cell.number_format = 'R$ #,##0.00'

    # Alinhamento
    ws_new.cell(linha_saida, 1).alignment = Alignment(horizontal="center")
    for col in range(2, 10):
        ws_new.cell(linha_saida, col).alignment = Alignment(horizontal="left", vertical="center")

    linha_saida += 1

# Ajustar larguras
ws_new.column_dimensions['A'].width = 8
ws_new.column_dimensions['B'].width = 18
ws_new.column_dimensions['C'].width = 16
ws_new.column_dimensions['D'].width = 40
ws_new.column_dimensions['E'].width = 18
ws_new.column_dimensions['F'].width = 16
ws_new.column_dimensions['G'].width = 16
ws_new.column_dimensions['H'].width = 16
ws_new.column_dimensions['I'].width = 30

# Congelar primeira linha
ws_new.freeze_panes = "A2"

# Salvar
output_path = Path(r"C:\Users\user\Downloads\RELATORIO_PRECOS_ATUALIZADOS_2026-07-26.xlsx")
wb_new.save(output_path)

print(f"\n✅ Relatorio salvo: {output_path.name}")

print(f"\n" + "="*100)
print("RESUMO FINAL")
print("="*100)

print(f"\n  Total de anuncios: {row_count}")
print(f"  Atualizado com sucesso: {total_ok}")
print(f"  Com atencao/erro: {total_erro}")

if row_count > 0:
    print(f"  Taxa de sucesso: {100*total_ok/row_count:.1f}%")

print(f"\n✅ Arquivo pronto em: {output_path.name}")
