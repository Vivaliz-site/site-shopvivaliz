#!/usr/bin/env python3
"""
Gerar RELATORIO FINAL com 37 colunas, seguindo padrão do arquivo Anuncios-ML-com-precos-Tiny.xlsx
"""
import os
import requests
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("GERAR RELATORIO FINAL COM 37 COLUNAS")
print("="*100)

# 1. LER RELATORIO ORIGINAL (com 6 precos)
print("\n[1] Lendo relatorio original...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb_orig = openpyxl.load_workbook(file_path)
ws_orig = wb_orig.active

# 2. LER ARQUIVO EXEMPLO (pra pegar as 25 primeiras colunas de ML)
print("\n[2] Lendo arquivo exemplo...")

file_exemplo = Path(r"C:\Users\user\Downloads\Anuncios-ML-com-precos-Tiny.xlsx")
wb_exemplo = openpyxl.load_workbook(file_exemplo)
ws_exemplo = wb_exemplo.active

# Pegar headers das primeiras 25 colunas
headers_ml = []
for col in range(1, 26):
    headers_ml.append(ws_exemplo.cell(1, col).value)

# 3. CARREGAR DADOS DO TINY
print("\n[3] Carregando dados do Tiny...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

produtos_tiny = {}

for prod in resp.json().get("itens", []):
    id_prod = prod.get("id")
    produtos_tiny[id_prod] = {
        'sku': prod.get("sku"),
        'descricao': prod.get("descricao"),
        'preco': prod.get("precos", {}).get("preco")
    }

print(f"    Carregados {len(produtos_tiny)} produtos")

# 4. CRIAR NOVO WORKBOOK COM 37 COLUNAS
print("\n[4] Criando novo relatorio...")

wb_new = openpyxl.Workbook()
ws_new = wb_new.active

# Headers
headers_completos = headers_ml + [
    "ID produto Tiny",
    "SKU Tiny",
    "Descrição Tiny",
    "Preço Mercado Livre original",
    "Preço cadastro Tiny",
    "Preço tabela PDV",
    "Preço tabela Shopee",
    "Preço tabela Amazon",
    "Preço tabela TikTok",
    "Status da localização",
    "Critério utilizado",
    "Observação"
]

# Escrever headers
for col_idx, header in enumerate(headers_completos, 1):
    cell = ws_new.cell(1, col_idx)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Preencher dados
linha_saida = 2

for row_orig in range(2, ws_orig.max_row + 1):
    titulo_ml = ws_orig.cell(row_orig, 2).value
    id_tiny = ws_orig.cell(row_orig, 3).value
    preco_ml = ws_orig.cell(row_orig, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    # Procurar a mesma linha no arquivo exemplo
    row_exemplo = None

    for r in range(2, ws_exemplo.max_row + 1):
        titulo_exemplo = ws_exemplo.cell(r, 5).value  # Coluna 5 = TITLE

        if titulo_exemplo and titulo_ml and str(titulo_exemplo).strip() == str(titulo_ml).strip():
            row_exemplo = r
            break

    # Copiar colunas 1-25 do arquivo exemplo (ou do original se não encontrou)
    if row_exemplo:
        for col in range(1, 26):
            cell_src = ws_exemplo.cell(row_exemplo, col)
            cell_dst = ws_new.cell(linha_saida, col)
            cell_dst.value = cell_src.value
    else:
        # Se não achou no exemplo, deixar em branco as 25 colunas ML
        pass

    # Preencher colunas 26-37
    ws_new.cell(linha_saida, 26).value = id_tiny
    ws_new.cell(linha_saida, 27).value = produtos_tiny.get(id_tiny, {}).get('sku')
    ws_new.cell(linha_saida, 28).value = produtos_tiny.get(id_tiny, {}).get('descricao')
    ws_new.cell(linha_saida, 29).value = preco_ml

    # Preços das tabelas (copiar do relatorio original)
    ws_new.cell(linha_saida, 30).value = ws_orig.cell(row_orig, 7).value  # Cadastro
    ws_new.cell(linha_saida, 31).value = ws_orig.cell(row_orig, 8).value  # PDV
    ws_new.cell(linha_saida, 32).value = ws_orig.cell(row_orig, 9).value  # Shopee
    ws_new.cell(linha_saida, 33).value = ws_orig.cell(row_orig, 10).value # Amazon
    ws_new.cell(linha_saida, 34).value = ws_orig.cell(row_orig, 11).value # TikTok

    # Status (verificar se foi atualizado)
    if id_tiny not in produtos_tiny:
        status = "ERRO 404"
        obs = "Produto não encontrado no Tiny"
    else:
        preco_atual = produtos_tiny[id_tiny]['preco']
        preco_esperado = round(preco_ml + 0.01, 2)

        if abs(float(preco_atual) - preco_esperado) < 0.01:
            status = "OK - Atualizado"
            obs = f"Preco de {ws_orig.cell(row_orig, 6).value:.2f} para {preco_atual}"

            # Colorir linha de verde
            fill_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            for col in range(1, 38):
                ws_new.cell(linha_saida, col).fill = fill_green

        elif any(ws_orig.cell(row_orig, c).fill and 'FF0000' in str(ws_orig.cell(row_orig, c).fill.start_color.rgb) for c in [7,8,9,10,11]):
            status = "NAO ATUALIZADO"
            obs = f"Preço esperado {preco_esperado}, obteve {preco_atual}"

            # Colorir linha de vermelho
            fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            font_white = Font(color="FFFFFF")

            for col in range(1, 38):
                ws_new.cell(linha_saida, col).fill = fill_red
                ws_new.cell(linha_saida, col).font = font_white

        else:
            status = "OK (sem ajuste necessario)"
            obs = ""

    ws_new.cell(linha_saida, 35).value = status
    ws_new.cell(linha_saida, 36).value = "Comparacao ML vs Tiny"
    ws_new.cell(linha_saida, 37).value = obs

    linha_saida += 1

# Ajustar larguras de coluna
for col in range(1, 38):
    ws_new.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

# Salvar
output_path = Path(r"C:\Users\user\Downloads\RELATORIO_FINAL_37_COLUNAS_2026-07-26.xlsx")
wb_new.save(output_path)

print(f"\n✅ Relatorio salvo: {output_path.name}")
print(f"   Total de linhas: {linha_saida - 2}")
