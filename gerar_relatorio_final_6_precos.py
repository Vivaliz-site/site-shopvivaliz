#!/usr/bin/env python3
"""
RELATÓRIO FINAL: Merge do arquivo Tiny (com 6 preços) + Planilha Base ML
Resultado: 6 preços por produto (ML + Cadastro + PDV + Shopee + Amazon + TikTok)
"""
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

print("\n╔════════════════════════════════════════════════════════════════════════════════╗")
print("║  RELATÓRIO FINAL: 6 PREÇOS (ML + CADASTRO + PDV + SHOPEE + AMAZON + TIKTOK)   ║")
print("╚════════════════════════════════════════════════════════════════════════════════╝")

# 1. CARREGAR ARQUIVO DO TINY (com 6 preços)
print("\n📥 Carregando arquivo do Tiny...")
tiny_file = Path(r"C:\Users\user\Downloads\Anuncios-ML-com-precos-Tiny.xlsx")
wb_tiny = openpyxl.load_workbook(tiny_file, data_only=True)
ws_tiny = wb_tiny.active

print(f"✅ Carregado: {tiny_file.name}")
print(f"   Linhas: {ws_tiny.max_row}, Colunas: {ws_tiny.max_column}")

# Encontrar os índices das colunas que precisamos
header_tiny = {}
for col_idx in range(1, ws_tiny.max_column + 1):
    col_name = ws_tiny.cell(1, col_idx).value
    if col_name:
        header_tiny[str(col_name).strip()] = col_idx

print(f"\n✅ Colunas encontradas no arquivo Tiny:")
cols_needed = [
    "TITLE", "PRICE", "ID produto Tiny", "SKU Tiny", "Descrição Tiny",
    "Preço Mercado Livre ", "Preço cadastro Tiny",
    "Preço tabela PDV", "Preço tabela Shopee", "Preço tabela Amazon", "Preço tabela TikTok",
    "Status da localizaçã"
]

for col in cols_needed:
    if col in header_tiny:
        print(f"   ✅ {col:30s} → Coluna {header_tiny[col]}")
    else:
        # Tentar variações
        for key in header_tiny.keys():
            if col.lower().replace(" ", "").replace("ã", "a") in key.lower().replace(" ", "").replace("ã", "a"):
                header_tiny[col] = header_tiny[key]
                print(f"   ✅ {col:30s} → Coluna {header_tiny[col]} (encontrada como '{key}')")
                break
        else:
            print(f"   ❌ {col:30s} → NÃO ENCONTRADA")

# 2. CRIAR NOVO ARQUIVO DE SAÍDA
print("\n📝 Criando relatório final...")

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Preços Comparativos"

# CABEÇALHO
headers_out = [
    "Linha",
    "Título ML",
    "ID Tiny",
    "SKU Tiny",
    "Descrição Tiny",
    "Preço ML",
    "Preço Cadastro",
    "Preço PDV",
    "Preço Shopee",
    "Preço Amazon",
    "Preço TikTok",
    "Status"
]

for col_idx, header_text in enumerate(headers_out, 1):
    cell = ws_out.cell(1, col_idx, header_text)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 3. PREENCHER DADOS
print("\n🔄 Processando...")

stats = {"total": 0, "found": 0, "errors": 0}
out_row = 2

for in_row in range(2, ws_tiny.max_row + 1):
    # Ler dados do Tiny
    title = ws_tiny.cell(in_row, header_tiny.get("TITLE", 5)).value
    price_ml = ws_tiny.cell(in_row, header_tiny.get("Preço Mercado Livre ", 29)).value
    id_tiny = ws_tiny.cell(in_row, header_tiny.get("ID produto Tiny", 26)).value
    sku_tiny = ws_tiny.cell(in_row, header_tiny.get("SKU Tiny", 27)).value
    desc_tiny = ws_tiny.cell(in_row, header_tiny.get("Descrição Tiny", 28)).value
    price_cadastro = ws_tiny.cell(in_row, header_tiny.get("Preço cadastro Tiny", 30)).value
    price_pdv = ws_tiny.cell(in_row, header_tiny.get("Preço tabela PDV", 31)).value
    price_shopee = ws_tiny.cell(in_row, header_tiny.get("Preço tabela Shopee", 32)).value
    price_amazon = ws_tiny.cell(in_row, header_tiny.get("Preço tabela Amazon", 33)).value
    price_tiktok = ws_tiny.cell(in_row, header_tiny.get("Preço tabela TikTok", 34)).value
    status = ws_tiny.cell(in_row, header_tiny.get("Status da localizaçã", 35)).value

    # Validações
    if not title or not price_ml:
        continue

    stats["total"] += 1

    # Determinar status
    if id_tiny and str(id_tiny).strip() != "":
        stats["found"] += 1
        final_status = "✅ OK"
        # Cor verde para encontrado
        color = "E2EFDA"
    else:
        stats["errors"] += 1
        final_status = "❌ Não encontrado"
        # Cor vermelha para erro
        color = "FCE4D6"

    # Preencher linha
    ws_out.cell(out_row, 1, in_row - 1)  # Número da linha original
    ws_out.cell(out_row, 2, str(title)[:50] if title else "")
    ws_out.cell(out_row, 3, id_tiny)
    ws_out.cell(out_row, 4, sku_tiny)
    ws_out.cell(out_row, 5, str(desc_tiny)[:50] if desc_tiny else "")
    ws_out.cell(out_row, 6, price_ml)
    ws_out.cell(out_row, 7, price_cadastro)
    ws_out.cell(out_row, 8, price_pdv)
    ws_out.cell(out_row, 9, price_shopee)
    ws_out.cell(out_row, 10, price_amazon)
    ws_out.cell(out_row, 11, price_tiktok)
    ws_out.cell(out_row, 12, final_status)

    # Aplicar cor
    for col_idx in range(1, 13):
        ws_out.cell(out_row, col_idx).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws_out.cell(out_row, col_idx).alignment = Alignment(horizontal="center", vertical="center")

    # Formatar colunas de preço como moeda
    for col in [6, 7, 8, 9, 10, 11]:
        ws_out.cell(out_row, col).number_format = 'R$ #,##0.00'

    if stats["total"] % 20 == 0:
        print(f"  [{in_row}] {str(title)[:40]}")

    out_row += 1

# Ajustar largura das colunas
widths = [10, 35, 12, 12, 35, 14, 14, 14, 14, 14, 14, 15]
for col_idx, width in enumerate(widths, 1):
    ws_out.column_dimensions[get_column_letter(col_idx)].width = width

# Congelar cabeçalho
ws_out.freeze_panes = "A2"

# SALVAR
output_file = Path(r"C:\Users\user\Downloads\RELATORIO_FINAL_6_PRECOS.xlsx")
print(f"\n💾 Salvando: {output_file.name}")
wb_out.save(output_file)

wb_tiny.close()

# RELATÓRIO
print(f"\n{'='*80}")
print(f"✅ RELATÓRIO GERADO COM SUCESSO")
print(f"{'='*80}")
print(f"\n📊 ESTATÍSTICAS:")
print(f"  Total processado: {stats['total']}")
print(f"  ✅ Encontrados: {stats['found']}")
print(f"  ❌ Erros: {stats['errors']}")
print(f"  Taxa de sucesso: {100*stats['found']//max(1,stats['total'])}%")
print(f"\n📁 Arquivo: {output_file}")
print(f"\n✅ Pronto para baixar!")
