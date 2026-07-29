#!/usr/bin/env python3
"""
Gerar relatório final consolidado com:
1. IDs processados
2. Preços antes/depois
3. Status da atualização
4. Validação (dados salvos no Tiny)
"""
import os
import time
import requests
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("GERAR RELATORIO FINAL CONSOLIDADO")
print("="*100)

# 1. LER RELATORIO ORIGINAL
print("\n[1] Lendo relatorio original...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb_orig = openpyxl.load_workbook(file_path)
ws_orig = wb_orig.active

# 2. CARREGAR DADOS ATUAIS DO TINY
print("\n[2] Carregando dados atuais do Tiny...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

produtos_map = {}

for prod in resp.json().get("itens", []):
    id_prod = prod.get("id")
    produtos_map[id_prod] = {
        'preco': prod.get("precos", {}).get("preco")
    }

print(f"    Carregados {len(produtos_map)} produtos")

# 3. CRIAR NOVO RELATORIO COM VALIDACAO
print("\n[3] Criando novo relatorio...")

# Copiar estrutura do relatorio original
wb_new = openpyxl.Workbook()
ws_new = wb_new.active

# Copiar headers
for col in range(1, 13):
    cell_src = ws_orig.cell(1, col)
    cell_dst = ws_new.cell(1, col)
    cell_dst.value = cell_src.value
    cell_dst.font = Font(bold=True)
    cell_dst.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

# Adicionar colunas de validacao
ws_new.cell(1, 13).value = "Status"
ws_new.cell(1, 14).value = "Preco Tiny Atual"
ws_new.cell(1, 15).value = "Validacao"

ws_new.cell(1, 13).font = Font(bold=True)
ws_new.cell(1, 14).font = Font(bold=True)
ws_new.cell(1, 15).font = Font(bold=True)

# Preencher dados
total_processados = 0
total_validados_ok = 0

for row_idx in range(2, ws_orig.max_row + 1):
    id_tiny = ws_orig.cell(row_idx, 3).value
    preco_ml = ws_orig.cell(row_idx, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    # Copiar dados originais
    for col in range(1, 12):
        cell_src = ws_orig.cell(row_idx, col)
        cell_dst = ws_new.cell(row_idx, col)
        cell_dst.value = cell_src.value

    # Verificar se tinha vermelho
    teve_vermelho = False

    for col_idx in [7, 8, 9, 10, 11]:
        cell = ws_orig.cell(row_idx, col_idx)

        if cell.fill and cell.fill.start_color:
            color_str = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color_str:
                teve_vermelho = True
                break

    if not teve_vermelho:
        ws_new.cell(row_idx, 13).value = "OK (sem vermelho)"
        continue

    total_processados += 1

    # Validar contra Tiny
    if id_tiny not in produtos_map:
        ws_new.cell(row_idx, 13).value = "ERRO (404)"
        ws_new.cell(row_idx, 14).value = "-"
        ws_new.cell(row_idx, 15).value = "FALHA"
        continue

    preco_atual = produtos_map[id_tiny]['preco']
    preco_esperado = round(preco_ml + 0.01, 2)

    ws_new.cell(row_idx, 13).value = "Atualizado"
    ws_new.cell(row_idx, 14).value = preco_atual

    if abs(float(preco_atual) - preco_esperado) < 0.01:
        ws_new.cell(row_idx, 15).value = "OK"
        total_validados_ok += 1

        # Colorir com verde
        fill_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        for col in range(13, 16):
            ws_new.cell(row_idx, col).fill = fill_green

    else:
        ws_new.cell(row_idx, 15).value = f"FALHA (esperado {preco_esperado})"

        # Colorir com vermelho
        fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        font_white = Font(color="FFFFFF")

        for col in range(13, 16):
            ws_new.cell(row_idx, col).fill = fill_red
            ws_new.cell(row_idx, col).font = font_white

    time.sleep(0.05)

# 4. SALVAR RELATORIO
print(f"\n[4] Salvando relatorio...")

output_path = Path(r"C:\Users\user\Downloads\RELATORIO_FINAL_VALIDADO_2026-07-26.xlsx")
wb_new.save(output_path)

print(f"    Salvo em: {output_path}")

# 5. RELATORIO FINAL
print(f"\n" + "="*100)
print("RELATORIO FINAL")
print("="*100)

print(f"\n[RESUMO]")
print(f"  Total de linhas: {ws_orig.max_row - 1}")
print(f"  Com vermelho (processadas): {total_processados}")
print(f"  Sem vermelho (OK): {(ws_orig.max_row - 1) - total_processados}")

print(f"\n[VALIDACAO]")
print(f"  Validadas com sucesso: {total_validados_ok}/{total_processados}")
print(f"  Taxa de sucesso: {100*total_validados_ok/total_processados:.1f}%")

print(f"\n✅ Relatorio salvo: {output_path.name}")
