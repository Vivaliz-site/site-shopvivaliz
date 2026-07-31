#!/usr/bin/env python3
"""
Gerar RELATORIO FINAL com estrutura corrigida:
- Apenas dados REAIS do relatorio original + Tiny
- Sem copiar dados do arquivo exemplo
- 37 colunas no total (12 do relatorio original + 25 vazias de ML)
"""
import os
import requests
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("GERAR RELATORIO FINAL - DADOS REAIS")
print("="*100)

# 1. LER RELATORIO ORIGINAL
print("\n[1] Lendo relatorio original...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb_orig = openpyxl.load_workbook(file_path)
ws_orig = wb_orig.active

# 2. CARREGAR DADOS ATUAIS DO TINY
print("\n[2] Carregando dados do Tiny...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

produtos_tiny = {}

for prod in resp.json().get("itens", []):
    id_prod = prod.get("id")
    produtos_tiny[id_prod] = {
        'sku': prod.get("sku"),
        'descricao': prod.get("descricao"),
        'preco': prod.get("precos", {}).get("preco")
    }

print(f"    Carregados {len(produtos_tiny)} produtos")

# 3. CRIAR NOVO WORKBOOK
print("\n[3] Criando novo relatorio...")

wb_new = openpyxl.Workbook()
ws_new = wb_new.active

# Headers - 37 colunas
# 25 vazias de ML + 12 do nosso relatorio
headers_completos = [
    "FAMILY_ID",            # 1
    "ITEM_ID",              # 2
    "PRODUCT_NUMBER",       # 3
    "VARIATION_ID",         # 4
    "TITLE",                # 5
    "VARIATIONS",           # 6
    "STORE_STOCK_QUANTITY", # 7
    "STOCK_FULL",           # 8
    "PRICE",                # 9
    "CURRENCY_ID",          # 10
    "TIERED_PRICING_PRICE_1",    # 11
    "TIERED_PRICING_QUANTITY_1", # 12
    "TIERED_PRICING_PRICE_2",    # 13
    "TIERED_PRICING_QUANTITY_2", # 14
    "TIERED_PRICING_PRICE_3",    # 15
    "TIERED_PRICING_QUANTITY_3", # 16
    "TIERED_PRICING_PRICE_4",    # 17
    "TIERED_PRICING_QUANTITY_4", # 18
    "TIERED_PRICING_PRICE_5",    # 19
    "TIERED_PRICING_QUANTITY_5", # 20
    "CONDITION",                 # 21
    "SHIPPING_METHOD",           # 22
    "LISTING_TYPE",              # 23
    "FEE_PER_SALE",              # 24
    "STATUS",                    # 25
    "ID produto Tiny",           # 26
    "SKU Tiny",                  # 27
    "Descrição Tiny",            # 28
    "Preço Mercado Livre",       # 29
    "Preço Cadastro Tiny",       # 30
    "Preço PDV",                 # 31
    "Preço Shopee",              # 32
    "Preço Amazon",              # 33
    "Preço TikTok",              # 34
    "Status Atualização",        # 35
    "Critério",                  # 36
    "Observação"                 # 37
]

# Escrever headers
for col_idx, header in enumerate(headers_completos, 1):
    cell = ws_new.cell(1, col_idx)
    cell.value = header
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Preencher dados
linha_saida = 2
total_processados = 0
total_ok = 0
total_erro = 0

for row_orig in range(2, ws_orig.max_row + 1):
    titulo_ml = ws_orig.cell(row_orig, 2).value
    id_tiny = ws_orig.cell(row_orig, 3).value
    preco_ml = ws_orig.cell(row_orig, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    total_processados += 1

    # Colunas 1-25: vazias (dados ML não preenchemos)
    # Deixar em branco propositalmente

    # Colunas 26-34: dados do relatorio
    ws_new.cell(linha_saida, 26).value = id_tiny
    ws_new.cell(linha_saida, 27).value = produtos_tiny.get(id_tiny, {}).get('sku')
    ws_new.cell(linha_saida, 28).value = produtos_tiny.get(id_tiny, {}).get('descricao')
    ws_new.cell(linha_saida, 29).value = preco_ml

    # Preços das tabelas (do relatorio com formatacao)
    preco_cadastro = ws_orig.cell(row_orig, 7).value
    preco_pdv = ws_orig.cell(row_orig, 8).value
    preco_shopee = ws_orig.cell(row_orig, 9).value
    preco_amazon = ws_orig.cell(row_orig, 10).value
    preco_tiktok = ws_orig.cell(row_orig, 11).value

    ws_new.cell(linha_saida, 30).value = preco_cadastro
    ws_new.cell(linha_saida, 31).value = preco_pdv
    ws_new.cell(linha_saida, 32).value = preco_shopee
    ws_new.cell(linha_saida, 33).value = preco_amazon
    ws_new.cell(linha_saida, 34).value = preco_tiktok

    # Verificar se tinha vermelho na linha original
    tem_vermelho = False

    for col_idx in [7, 8, 9, 10, 11]:
        cell = ws_orig.cell(row_orig, col_idx)

        if cell.fill and cell.fill.start_color:
            color_str = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color_str:
                tem_vermelho = True
                break

    if not tem_vermelho:
        ws_new.cell(linha_saida, 35).value = "OK (sem ajuste)"
        ws_new.cell(linha_saida, 36).value = ""
        ws_new.cell(linha_saida, 37).value = ""
        total_ok += 1

    else:
        # Validar contra Tiny
        if id_tiny not in produtos_tiny:
            ws_new.cell(linha_saida, 35).value = "ERRO 404"
            ws_new.cell(linha_saida, 36).value = "Não encontrado"
            ws_new.cell(linha_saida, 37).value = "ID não existe no Tiny"
            total_erro += 1

            # Colorir com vermelho
            fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            font_white = Font(color="FFFFFF")

            for col in range(26, 38):
                ws_new.cell(linha_saida, col).fill = fill_red
                ws_new.cell(linha_saida, col).font = font_white

        else:
            preco_atual = produtos_tiny[id_tiny]['preco']
            preco_esperado = round(preco_ml + 0.01, 2)

            if abs(float(preco_atual) - preco_esperado) < 0.01:
                ws_new.cell(linha_saida, 35).value = "OK - Atualizado"
                ws_new.cell(linha_saida, 36).value = "Preco ML + 0.01"
                ws_new.cell(linha_saida, 37).value = f"Preco atual: R$ {preco_atual}"
                total_ok += 1

                # Colorir com verde
                fill_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

                for col in range(26, 38):
                    ws_new.cell(linha_saida, col).fill = fill_green

            else:
                ws_new.cell(linha_saida, 35).value = "NAO ATUALIZADO"
                ws_new.cell(linha_saida, 36).value = "Preco diferente"
                ws_new.cell(linha_saida, 37).value = f"Esperado R$ {preco_esperado}, obteve R$ {preco_atual}"
                total_erro += 1

                # Colorir com laranja
                fill_orange = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

                for col in range(26, 38):
                    ws_new.cell(linha_saida, col).fill = fill_orange

    # Formatacao de numeros
    for col in [29, 30, 31, 32, 33, 34]:
        ws_new.cell(linha_saida, col).number_format = '0.00'

    linha_saida += 1

# Ajustar larguras de coluna
for col in range(1, 38):
    ws_new.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

# Congelar primeira linha
ws_new.freeze_panes = "A2"

# Salvar
output_path = Path(r"C:\Users\user\Downloads\RELATORIO_FINAL_37_COLUNAS_2026-07-26.xlsx")
wb_new.save(output_path)

print(f"\n✅ Relatorio salvo: {output_path.name}")

print(f"\n" + "="*100)
print("RESUMO")
print("="*100)

print(f"\n  Total de linhas: {total_processados}")
print(f"  OK (sem ajuste ou atualizado): {total_ok}")
print(f"  Erro (nao atualizado ou 404): {total_erro}")
print(f"\n  Taxa de sucesso: {100*total_ok/total_processados:.1f}%")
