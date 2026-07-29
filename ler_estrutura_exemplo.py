#!/usr/bin/env python3
import openpyxl
from pathlib import Path

file = Path(r"C:\Users\user\Downloads\Anuncios-ML-com-precos-Tiny.xlsx")
wb = openpyxl.load_workbook(file)
ws = wb.active

print("HEADERS DO ARQUIVO EXEMPLO:")
print("="*100)

for col in range(1, ws.max_column + 1):
    header = ws.cell(1, col).value
    if header:
        print(f"  Col {col:2d}: {header}")

print(f"\nTotal de colunas: {ws.max_column}")
print(f"Total de linhas: {ws.max_row}")

print("\n\nPRIMEIRAS 2 LINHAS DE DADOS:")
print("="*100)

for row in range(2, 4):
    print(f"\nLinha {row}:")
    for col in range(1, min(ws.max_column + 1, 15)):
        header = ws.cell(1, col).value
        value = ws.cell(row, col).value
        print(f"  {header:20s}: {value}")
