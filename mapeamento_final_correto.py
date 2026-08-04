#!/usr/bin/env python3
"""
MAPEAMENTO FINAL CORRETO:
Usar SKU do arquivo de Mapeamento para encontrar o PRODUTO TINY correto
SKU → ID do Produto Tiny
"""
import os
import time
import requests
import openpyxl
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("MAPEAMENTO FINAL CORRETO: Usar SKU para mapear ML → Produto Tiny → Precos Tabelas")
print("="*100)

# 1. CARREGAR MAPEAMENTO (com SKU)
print("\n[1] Carregando mapeamento (Id, Identificacao, SKU)...")

mapping_sku = {}  # SKU → {titulo, id_integracao}

mapping_files = [
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-48.xls',
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-50.xls',
]

for mfile in mapping_files:
    if Path(mfile).exists():
        try:
            df = pd.read_excel(mfile, sheet_name=0)
            for idx, row in df.iterrows():
                sku = str(row.get('Produto (SKU)')).strip().upper()
                titulo = row.get('Título')
                id_integracao = row.get('Id')

                if sku and sku != 'NAN':
                    mapping_sku[sku] = {
                        'titulo': titulo,
                        'id_integracao': id_integracao
                    }

            print(f"    {Path(mfile).name}: {len(df)} items")
        except Exception as e:
            print(f"    Erro: {e}")

print(f"    Total com SKU: {len(mapping_sku)}")
print(f"    Exemplos: {list(mapping_sku.keys())[:3]}")

# 2. CARREGAR PRODUTOS TINY (com SKU como chave)
print("\n[2] Carregando produtos Tiny (mapeando por SKU)...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

if resp.status_code != 200:
    print(f"Erro: {resp.status_code}")
    exit(1)

produtos_tiny_por_sku = {}  # SKU → {id, descricao, preco_cadastro}

for prod in resp.json().get("itens", []):
    sku = str(prod.get("sku", "")).strip().upper()
    id_tiny = prod.get("id")
    desc = prod.get("descricao")
    preco_cad = prod.get("precos", {}).get("preco")

    if sku:
        produtos_tiny_por_sku[sku] = {
            'id': id_tiny,
            'descricao': desc,
            'preco_cadastro': preco_cad
        }

print(f"    Carregados {len(produtos_tiny_por_sku)} produtos (com SKU)")

# 3. CARREGAR TABELAS DE PRECO
print("\n[3] Carregando excecoes das tabelas...")

TABELAS = {"PDV": 982, "Shopee": 989, "Amazon": 991, "TikTok": 990}
tabelas_data = {}  # id_produto → {tabela → preco}

for nome_tabela, id_tabela in TABELAS.items():
    print(f"    {nome_tabela}...", end=" ", flush=True)

    url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}"

    try:
        resp = requests.get(url, headers=headers, timeout=20)

        if resp.status_code == 200:
            data = resp.json()
            excecoes = data.get("excecoes", [])

            for exc in excecoes:
                id_prod = exc.get("idProduto")
                preco = exc.get("preco")

                if id_prod not in tabelas_data:
                    tabelas_data[id_prod] = {}

                tabelas_data[id_prod][nome_tabela] = preco

            print(f"{len(excecoes)}")
        else:
            print(f"Erro {resp.status_code}")

    except Exception as e:
        print(f"Erro")

    time.sleep(0.5)

print(f"    Total de produtos com excecoes: {len(tabelas_data)}")

# 4. CARREGAR PLANILHA BASE ML
print("\n[4] Carregando planilha base ML...")

ml_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
wb_ml = openpyxl.load_workbook(ml_file, data_only=True)
ws_ml = wb_ml.active

header = {}
for col_idx in range(1, ws_ml.max_column + 1):
    val = ws_ml.cell(1, col_idx).value
    if val:
        header[str(val).strip()] = col_idx

title_col = header.get("TITLE", 5)
price_col = header.get("PRICE", 9)

print(f"    {ml_file.name}: {ws_ml.max_row} linhas")

# 5. CRIAR RELATORIO
print("\n[5] Criando relatorio final...")

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Precos Comparativos"

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

headers_out = [
    "Linha ML",
    "Titulo ML",
    "ID Tiny",
    "SKU Tiny",
    "Descricao Tiny",
    "Preco ML",
    "Preco Cadastro",
    "Preco PDV",
    "Preco Shopee",
    "Preco Amazon",
    "Preco TikTok",
    "Status"
]

for col_idx, header_text in enumerate(headers_out, 1):
    cell = ws_out.cell(1, col_idx, header_text)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 6. PROCESSAR
print("\n[6] Processando 89 produtos...\n")

stats = {"total": 0, "encontrados": 0, "com_preco": 0}
out_row = 2

for in_row in range(6, min(6 + 89, ws_ml.max_row + 1)):
    title_ml = ws_ml.cell(in_row, title_col).value
    price_ml = ws_ml.cell(in_row, price_col).value

    if not title_ml or not price_ml:
        continue

    title_ml = str(title_ml).strip()
    price_ml = float(price_ml)

    if price_ml <= 0:
        continue

    stats["total"] += 1

    # Procurar nos SKUs do mapeamento por similaridade com titulo
    id_tiny = None
    sku_match = None
    best_score = 0

    title_lower = title_ml.lower()

    for sku, mapping_data in mapping_sku.items():
        titulo_map = mapping_data['titulo'].lower()

        # Match simples: se o titulo ML esta contido no titulo do mapeamento
        if titulo_map.startswith(title_lower[:20]):
            # Procurar o SKU no Tiny
            if sku in produtos_tiny_por_sku:
                id_tiny = produtos_tiny_por_sku[sku]['id']
                sku_match = sku
                best_score = 1.0
                break

    # Se nao encontrou por match simples, tentar por palavra-chave
    if not id_tiny:
        palavras = [p for p in title_lower.split() if len(p) > 3]

        for sku, mapping_data in mapping_sku.items():
            titulo_map = mapping_data['titulo'].lower()

            matches = sum(1 for palavra in palavras if palavra in titulo_map)

            if matches > best_score:
                best_score = matches

                # Procurar o SKU no Tiny
                if sku in produtos_tiny_por_sku:
                    id_tiny = produtos_tiny_por_sku[sku]['id']
                    sku_match = sku

    # Obter dados
    if id_tiny and sku_match and sku_match in produtos_tiny_por_sku:
        stats["encontrados"] += 1

        prod_tiny = produtos_tiny_por_sku[sku_match]
        sku_tiny = sku_match
        desc_tiny = prod_tiny['descricao']
        price_cadastro = prod_tiny['preco_cadastro']

        # Precos das tabelas
        precos_tabelas = tabelas_data.get(id_tiny, {})

        if precos_tabelas:
            stats["com_preco"] += len(precos_tabelas)

        status = "OK"
        color = "E2EFDA"

    else:
        sku_tiny = ""
        desc_tiny = ""
        price_cadastro = None
        precos_tabelas = {}
        status = "NAO ENCONTRADO"
        color = "FCE4D6"

    # Preencher
    ws_out.cell(out_row, 1, in_row)
    ws_out.cell(out_row, 2, title_ml[:50])
    ws_out.cell(out_row, 3, id_tiny)
    ws_out.cell(out_row, 4, sku_tiny)
    ws_out.cell(out_row, 5, desc_tiny[:50] if desc_tiny else "")
    ws_out.cell(out_row, 6, price_ml)
    ws_out.cell(out_row, 7, price_cadastro)
    ws_out.cell(out_row, 8, precos_tabelas.get("PDV"))
    ws_out.cell(out_row, 9, precos_tabelas.get("Shopee"))
    ws_out.cell(out_row, 10, precos_tabelas.get("Amazon"))
    ws_out.cell(out_row, 11, precos_tabelas.get("TikTok"))
    ws_out.cell(out_row, 12, status)

    # Formatar
    for col_idx in range(1, 13):
        cell = ws_out.cell(out_row, col_idx)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in [6, 7, 8, 9, 10, 11]:
            cell.number_format = 'R$ #,##0.00'

    if stats["total"] % 15 == 0:
        print(f"  [{stats['total']:2d}] {title_ml[:45]}")

    out_row += 1

# Ajustar
widths = [10, 35, 12, 12, 35, 14, 14, 14, 14, 14, 14, 15]
for col_idx, width in enumerate(widths, 1):
    ws_out.column_dimensions[get_column_letter(col_idx)].width = width

ws_out.freeze_panes = "A2"

# Salvar
output_file = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_CORRETO.xlsx")
print(f"\n[7] Salvando...")
wb_out.save(output_file)

wb_ml.close()

# Relatorio
print(f"\n" + "="*100)
print(f"SUCESSO!")
print("="*100)
print(f"\n[RESULTADO]")
print(f"  Total processado: {stats['total']}")
print(f"  Encontrados no Tiny: {stats['encontrados']}")
print(f"  Precos de tabelas encontrados: {stats['com_preco']}")
print(f"\n[ARQUIVO]")
print(f"  {output_file}")
