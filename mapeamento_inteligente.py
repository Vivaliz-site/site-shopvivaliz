#!/usr/bin/env python3
"""
MAPEAMENTO INTELIGENTE:
1. Carregar arquivo ML com TITULO e PRECO
2. Para cada item ML, encontrar o PRODUTO EXATO no Tiny (por ID do mapeamento)
3. Buscar os 4 preços de tabelas PARA AQUELE PRODUTO ESPECIFICO
"""
import os
import time
import requests
import openpyxl
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from difflib import SequenceMatcher

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n╔════════════════════════════════════════════════════════════════════════════════╗")
print("║  MAPEAMENTO INTELIGENTE: ML → Tiny → Tabelas de Preço                        ║")
print("╚════════════════════════════════════════════════════════════════════════════════╝")

# 1. CARREGAR ARQUIVO DE MAPEAMENTO (ML → Tiny ID)
print("\n[1] Carregando mapeamento ML → Tiny ID...")

mapping_ml_tiny = {}  # Titulo ML → {id_tiny, sku, descricao}

mapping_files = [
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-48.xls',
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-50.xls',
]

for mfile in mapping_files:
    if Path(mfile).exists():
        try:
            df = pd.read_excel(mfile, sheet_name=0)
            for idx, row in df.iterrows():
                id_tiny = int(row.get('Id'))
                titulo_ml = row.get('Título')
                sku = row.get('Produto (SKU)')

                mapping_ml_tiny[titulo_ml] = {
                    'id': id_tiny,
                    'sku': sku
                }

            print(f"    {Path(mfile).name}: {len(df)} items")
        except Exception as e:
            print(f"    Erro: {e}")

print(f"    Total mapeado: {len(mapping_ml_tiny)}")

# 2. CARREGAR PLANILHA BASE ML (89 produtos)
print("\n[2] Carregando planilha base ML...")

ml_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
wb_ml = openpyxl.load_workbook(ml_file, data_only=True)
ws_ml = wb_ml.active

# Encontrar colunas
header = {}
for col_idx in range(1, ws_ml.max_column + 1):
    val = ws_ml.cell(1, col_idx).value
    if val:
        header[str(val).strip()] = col_idx

title_col = header.get("TITLE", 5)
price_col = header.get("PRICE", 9)

print(f"    {ml_file.name}: {ws_ml.max_row} linhas")

# 3. CARREGAR PRODUTOS TINY (com preços do cadastro)
print("\n[3] Carregando produtos Tiny...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

if resp.status_code != 200:
    print(f"❌ Erro: {resp.status_code}")
    exit(1)

produtos_tiny = {}  # ID → {sku, descricao, preco_cadastro}

for prod in resp.json().get("itens", []):
    id_tiny = prod.get("id")
    sku = prod.get("sku")
    desc = prod.get("descricao")
    preco_cad = prod.get("precos", {}).get("preco")

    produtos_tiny[id_tiny] = {
        'sku': sku,
        'descricao': desc,
        'preco_cadastro': preco_cad
    }

print(f"    Carregados {len(produtos_tiny)} produtos")

# 4. CARREGAR TODAS AS EXCECOES DAS TABELAS
print("\n[4] Carregando excecoes das tabelas...")

TABELAS = {"PDV": 982, "Shopee": 989, "Amazon": 991, "TikTok": 990}
tabelas_data = {}  # id_tiny → {tabela → preco}

for nome_tabela, id_tabela in TABELAS.items():
    print(f"    {nome_tabela}...", end=" ", flush=True)

    url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}"

    try:
        resp = requests.get(url, headers=headers, timeout=20)

        if resp.status_code == 200:
            data = resp.json()
            excecoes = data.get("excecoes", [])

            for exc in excecoes:
                id_prod = exc.get("idProduto")
                preco = exc.get("preco")

                if id_prod not in tabelas_data:
                    tabelas_data[id_prod] = {}

                tabelas_data[id_prod][nome_tabela] = preco

            print(f"{len(excecoes)}")
        else:
            print(f"Erro {resp.status_code}")

    except Exception as e:
        print(f"Erro")

    time.sleep(0.5)

print(f"    Total de produtos com excecoes: {len(tabelas_data)}")

# 5. CRIAR RELATORIO FINAL
print("\n[5] Criando relatorio final...")

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Precos Comparativos"

# CABECALHO
headers_out = [
    "Linha ML",
    "Titulo ML",
    "ID Tiny",
    "SKU Tiny",
    "Descricao Tiny",
    "Preco ML",
    "Preco Cadastro",
    "Preco PDV",
    "Preco Shopee",
    "Preco Amazon",
    "Preco TikTok",
    "Status"
]

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

for col_idx, header_text in enumerate(headers_out, 1):
    cell = ws_out.cell(1, col_idx, header_text)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# DADOS
print("\n[6] Processando 89 produtos...")

stats = {"total": 0, "found": 0, "with_prices": 0}
out_row = 2

for in_row in range(6, min(6 + 89, ws_ml.max_row + 1)):
    title_ml = ws_ml.cell(in_row, title_col).value
    price_ml = ws_ml.cell(in_row, price_col).value

    if not title_ml or not price_ml:
        continue

    title_ml = str(title_ml).strip()
    price_ml = float(price_ml)

    if price_ml <= 0:
        continue

    stats["total"] += 1

    # ESTRATEGIA 1: Buscar EXATO no mapeamento
    id_tiny = None

    if title_ml in mapping_ml_tiny:
        id_tiny = mapping_ml_tiny[title_ml]['id']
    else:
        # ESTRATEGIA 2: Buscar por similaridade
        title_lower = title_ml.lower()
        best_match = -1

        for titulo_map, dados in mapping_ml_tiny.items():
            titulo_map_lower = titulo_map.lower()
            score = SequenceMatcher(None, title_lower, titulo_map_lower).ratio()

            if score > 0.7 and len(titulo_map_lower) > best_match:
                best_match = len(titulo_map_lower)
                id_tiny = dados['id']

    # Obter dados do Tiny
    if id_tiny and id_tiny in produtos_tiny:
        stats["found"] += 1

        prod_tiny = produtos_tiny[id_tiny]
        sku_tiny = prod_tiny['sku']
        desc_tiny = prod_tiny['descricao']
        price_cadastro = prod_tiny['preco_cadastro']

        # Buscar precos nas tabelas
        precos_tabelas = tabelas_data.get(id_tiny, {})

        if precos_tabelas:
            stats["with_prices"] += len(precos_tabelas)

        status = "OK"
        color = "E2EFDA"

    else:
        sku_tiny = ""
        desc_tiny = ""
        price_cadastro = None
        precos_tabelas = {}
        status = "NAO ENCONTRADO"
        color = "FCE4D6"

    # Preencher linha
    ws_out.cell(out_row, 1, in_row)
    ws_out.cell(out_row, 2, title_ml[:50])
    ws_out.cell(out_row, 3, id_tiny)
    ws_out.cell(out_row, 4, sku_tiny)
    ws_out.cell(out_row, 5, desc_tiny[:50] if desc_tiny else "")
    ws_out.cell(out_row, 6, price_ml)
    ws_out.cell(out_row, 7, price_cadastro)
    ws_out.cell(out_row, 8, precos_tabelas.get("PDV"))
    ws_out.cell(out_row, 9, precos_tabelas.get("Shopee"))
    ws_out.cell(out_row, 10, precos_tabelas.get("Amazon"))
    ws_out.cell(out_row, 11, precos_tabelas.get("TikTok"))
    ws_out.cell(out_row, 12, status)

    # Formatar
    for col_idx in range(1, 13):
        cell = ws_out.cell(out_row, col_idx)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in [6, 7, 8, 9, 10, 11]:
            cell.number_format = 'R$ #,##0.00'

    if stats["total"] % 20 == 0:
        print(f"    [{stats['total']:2d}] {title_ml[:40]}")

    out_row += 1

# Ajustar colunas
widths = [10, 35, 12, 12, 35, 14, 14, 14, 14, 14, 14, 15]
for col_idx, width in enumerate(widths, 1):
    ws_out.column_dimensions[get_column_letter(col_idx)].width = width

ws_out.freeze_panes = "A2"

# SALVAR
output_file = Path(r"C:\Users\user\Downloads\RELATORIO_MAPEAMENTO_INTELIGENTE.xlsx")
print(f"\n[7] Salvando...")
wb_out.save(output_file)

wb_ml.close()

# RELATORIO
print(f"\n{'='*100}")
print(f"SUCESSO!")
print(f"{'='*100}")
print(f"\n[RESULTADO]")
print(f"  Total processado: {stats['total']}")
print(f"  Encontrados no Tiny: {stats['found']}")
print(f"  Precos de tabelas encontrados: {stats['with_prices']}")
print(f"\n[ARQUIVO]")
print(f"  {output_file}")
print(f"  Pronto para download!")
