#!/usr/bin/env python3
"""
ESTRATEGIA CORRIGIDA: Usar o SKU do Mapeamento DIRETAMENTE para buscar no Tiny
Nao fazer busca por palavras-chave, apenas SKU para SKU
"""
import os
import time
import requests
import openpyxl
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("MAPEAMENTO POR SKU DIRETO: SKU Mapeamento → Produto Tiny → Precos")
print("="*100)

# 1. CARREGAR MAPEAMENTO (SKU → Titulo ML)
print("\n[1] Carregando mapeamento (SKU)...")

sku_to_titulo = {}  # SKU → titulo_ml

mapping_files = [
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-48.xls',
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-50.xls',
]

for mfile in mapping_files:
    if Path(mfile).exists():
        df = pd.read_excel(mfile, sheet_name=0)
        for idx, row in df.iterrows():
            sku = str(row.get('Produto (SKU)')).strip()
            titulo = row.get('Título')

            if sku and sku.upper() != 'NAN':
                sku_to_titulo[sku.upper()] = titulo

        print(f"    {Path(mfile).name}: {len(df)} items")

print(f"    Total SKUs: {len(sku_to_titulo)}")

# 2. CARREGAR PRODUTOS TINY (SKU → ID + Dados)
print("\n[2] Carregando produtos Tiny...")

url = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url, headers=headers, timeout=30)

sku_to_product = {}  # SKU → {id, descricao, preco_cadastro}

for prod in resp.json().get("itens", []):
    sku = str(prod.get("sku", "")).strip().upper()

    if sku:
        sku_to_product[sku] = {
            'id': prod.get("id"),
            'descricao': prod.get("descricao"),
            'preco_cadastro': prod.get("precos", {}).get("preco")
        }

print(f"    Total: {len(sku_to_product)} produtos")

# Encontrados vs Nao encontrados
encontrados = set(sku_to_titulo.keys()) & set(sku_to_product.keys())
nao_encontrados = set(sku_to_titulo.keys()) - set(sku_to_product.keys())

print(f"    SKUs encontrados no Tiny: {len(encontrados)}")
print(f"    SKUs NAO encontrados: {len(nao_encontrados)}")
print(f"    Exemplos nao encontrados: {list(nao_encontrados)[:5]}")

# 3. CARREGAR TABELAS DE PRECO
print("\n[3] Carregando excecoes das tabelas...")

TABELAS = {"PDV": 982, "Shopee": 989, "Amazon": 991, "TikTok": 990}
tabelas_data = {}  # id_produto → {tabela → preco}

for nome_tabela, id_tabela in TABELAS.items():
    print(f"    {nome_tabela}...", end=" ", flush=True)

    url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}"

    try:
        resp = requests.get(url, headers=headers, timeout=20)

        if resp.status_code == 200:
            excecoes = resp.json().get("excecoes", [])

            for exc in excecoes:
                id_prod = exc.get("idProduto")
                preco = exc.get("preco")

                if id_prod not in tabelas_data:
                    tabelas_data[id_prod] = {}

                tabelas_data[id_prod][nome_tabela] = preco

            print(f"{len(excecoes)}")
        else:
            print(f"Erro {resp.status_code}")

    except Exception as e:
        print(f"Erro")

    time.sleep(0.5)

print(f"    Total de produtos com excecoes: {len(tabelas_data)}")

# 4. CARREGAR PLANILHA BASE ML
print("\n[4] Carregando planilha base ML...")

ml_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
wb_ml = openpyxl.load_workbook(ml_file, data_only=True)
ws_ml = wb_ml.active

header = {}
for col_idx in range(1, ws_ml.max_column + 1):
    val = ws_ml.cell(1, col_idx).value
    if val:
        header[str(val).strip()] = col_idx

title_col = header.get("TITLE", 5)
price_col = header.get("PRICE", 9)

print(f"    {ml_file.name}")

# 5. CRIAR RELATORIO
print("\n[5] Processando 89 produtos...")

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Precos Comparativos"

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

headers_out = [
    "Linha ML",
    "Titulo ML",
    "ID Tiny",
    "SKU Tiny",
    "Descricao Tiny",
    "Preco ML",
    "Preco Cadastro",
    "Preco PDV",
    "Preco Shopee",
    "Preco Amazon",
    "Preco TikTok",
    "Status"
]

for col_idx, header_text in enumerate(headers_out, 1):
    cell = ws_out.cell(1, col_idx, header_text)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# PROCESSAR
stats = {"total": 0, "encontrados": 0, "com_preco": 0}
out_row = 2

for in_row in range(6, min(6 + 89, ws_ml.max_row + 1)):
    title_ml = ws_ml.cell(in_row, title_col).value
    price_ml = ws_ml.cell(in_row, price_col).value

    if not title_ml or not price_ml:
        continue

    title_ml = str(title_ml).strip()
    price_ml = float(price_ml)

    if price_ml <= 0:
        continue

    stats["total"] += 1

    # Procurar SKU correspondente no mapeamento
    # Estrategia: procurar o SKU que tem titulo similar ao titulo ML

    title_lower = title_ml.lower()
    best_sku = None
    best_score = 0

    for sku, titulo_map in sku_to_titulo.items():
        titulo_map_lower = titulo_map.lower()

        # Match simples: verificar se tem palavras em comum
        score = 0

        # Se titulo_map contem a maior parte do title_ml
        if title_lower in titulo_map_lower:
            score = len(title_lower) / len(titulo_map_lower)
        elif titulo_map_lower in title_lower:
            score = len(titulo_map_lower) / len(title_lower)
        else:
            # Contar palavras em comum
            palavras_ml = [p for p in title_lower.split() if len(p) > 3]
            palavras_map = titulo_map_lower.split()

            matches = sum(1 for p_ml in palavras_ml if any(p_ml in p_map for p_map in palavras_map))
            score = matches / len(palavras_ml) if palavras_ml else 0

        if score > best_score:
            best_score = score
            best_sku = sku

    # Se encontrou um SKU com score > 0.5, procurar no Tiny
    id_tiny = None
    sku_tiny = None

    if best_sku and best_score > 0.5 and best_sku in sku_to_product:
        id_tiny = sku_to_product[best_sku]['id']
        sku_tiny = best_sku
        desc_tiny = sku_to_product[best_sku]['descricao']
        price_cadastro = sku_to_product[best_sku]['preco_cadastro']

        precos_tabelas = tabelas_data.get(id_tiny, {})

        stats["encontrados"] += 1
        if precos_tabelas:
            stats["com_preco"] += len(precos_tabelas)

        status = "OK"
        color = "E2EFDA"

    else:
        desc_tiny = ""
        price_cadastro = None
        precos_tabelas = {}
        status = "NAO ENCONTRADO"
        color = "FCE4D6"

    # Preencher
    ws_out.cell(out_row, 1, in_row)
    ws_out.cell(out_row, 2, title_ml[:50])
    ws_out.cell(out_row, 3, id_tiny)
    ws_out.cell(out_row, 4, sku_tiny)
    ws_out.cell(out_row, 5, desc_tiny[:50] if desc_tiny else "")
    ws_out.cell(out_row, 6, price_ml)
    ws_out.cell(out_row, 7, price_cadastro)
    ws_out.cell(out_row, 8, precos_tabelas.get("PDV"))
    ws_out.cell(out_row, 9, precos_tabelas.get("Shopee"))
    ws_out.cell(out_row, 10, precos_tabelas.get("Amazon"))
    ws_out.cell(out_row, 11, precos_tabelas.get("TikTok"))
    ws_out.cell(out_row, 12, status)

    # Formatar
    for col_idx in range(1, 13):
        cell = ws_out.cell(out_row, col_idx)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in [6, 7, 8, 9, 10, 11]:
            cell.number_format = 'R$ #,##0.00'

    if stats["total"] % 15 == 0:
        print(f"  [{stats['total']:2d}] {title_ml[:45]} - {status}")

    out_row += 1

# Salvar
widths = [10, 35, 12, 12, 35, 14, 14, 14, 14, 14, 14, 15]
for col_idx, width in enumerate(widths, 1):
    ws_out.column_dimensions[get_column_letter(col_idx)].width = width

ws_out.freeze_panes = "A2"

output_file = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_CORRETO.xlsx")
print(f"\n[6] Salvando {output_file.name}...")
wb_out.save(output_file)

wb_ml.close()

print(f"\n" + "="*100)
print("SUCESSO!")
print("="*100)
print(f"\n[RESULTADO]")
print(f"  Total processado: {stats['total']}")
print(f"  Encontrados no Tiny: {stats['encontrados']}")
print(f"  Precos de tabelas: {stats['com_preco']}")
print(f"\nArquivo: {output_file}")
