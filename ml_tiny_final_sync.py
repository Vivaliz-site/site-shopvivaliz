#!/usr/bin/env python3
"""
VERSÃO FINAL: Sincronização ML → Tiny (Preços)
Com DEBUG e cache funcionando
"""
import os
import json
import sys
import time
from pathlib import Path
from typing import Dict, Optional
from datetime import datetime
from difflib import SequenceMatcher

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import openpyxl
from openpyxl.styles import Font, PatternFill
from dotenv import load_dotenv

# CARREGAR ENV
env_files = [
    Path(r"C:\site-shopvivaliz\.env.local"),
    Path(r"C:\site-shopvivaliz\.env"),
]

for f in env_files:
    if f.exists():
        load_dotenv(f)
        break

# CLIENT
class TinyClient:
    def __init__(self):
        self.token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")
        if not self.token:
            print("❌ Token não encontrado")
            sys.exit(1)

        self.base_url = "https://api.tiny.com.br/public-api/v3"
        self.session = requests.Session()
        self.session.mount("https://", HTTPAdapter(max_retries=Retry(total=2, backoff_factor=1)))
        self.req_count = 0
        self.products = None

        print("✅ Client inicializado")

    def get_products(self, limit=500, offset=0):
        """Busca produtos com tratamento de erro"""
        self.req_count += 1
        headers = {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json"
        }

        url = f"{self.base_url}/produtos"
        params = {"limit": limit, "offset": offset}

        try:
            resp = self.session.get(url, headers=headers, params=params, timeout=20)
            print(f"  Requisição {self.req_count}: GET /produtos limit={limit} offset={offset}")
            print(f"    Status: {resp.status_code}")

            if resp.status_code == 200:
                data = resp.json()
                items = data.get("itens", [])
                print(f"    Retornou: {len(items)} itens")
                return data
            elif resp.status_code == 429:
                print(f"    ⚠️  Rate limit, aguardando...")
                time.sleep(60)
                return self.get_products(limit, offset)
            else:
                print(f"    ❌ Erro: {resp.status_code}")
                return None
        except Exception as e:
            print(f"    ❌ Exceção: {e}")
            return None

    def load_all_products(self):
        """Carrega TODOS os produtos"""
        print("\n🔄 Carregando todos os produtos...")

        all_products = []
        offset = 0
        limit = 500

        while True:
            result = self.get_products(limit, offset)
            if not result or "itens" not in result:
                print("  Nenhum resultado, parando")
                break

            items = result.get("itens", [])
            all_products.extend(items)
            print(f"  Total acumulado: {len(all_products)}")

            pagination = result.get("paginacao", {})
            total = pagination.get("total", 0)

            print(f"  Paginação: total={total}, offset={offset}, limit={limit}")

            if offset + limit >= total or len(items) == 0:
                print(f"  Fim alcançado")
                break

            offset += limit
            time.sleep(1)

        print(f"\n✅ Carregado: {len(all_products)} produtos")
        self.products = all_products
        return all_products

    def find_product_by_title(self, title, threshold=0.4):
        """Busca produto por título"""
        if self.products is None:
            return None

        title_lower = title.lower().strip()
        best = None
        best_score = threshold

        for prod in self.products:
            desc = prod.get("descricao", "").lower()
            score = SequenceMatcher(None, title_lower, desc).ratio()

            if score > best_score:
                best = prod
                best_score = score

        return best

    def update_price(self, product_id, new_price):
        """Atualiza preço do produto"""
        self.req_count += 1
        headers = {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json"
        }

        url = f"{self.base_url}/produtos/{product_id}"
        payload = {"precos": {"preco": float(new_price)}}

        try:
            resp = self.session.put(url, headers=headers, json=payload, timeout=20)
            print(f"    PUT /produtos/{product_id} → Status {resp.status_code}")
            return resp.status_code in [200, 204]
        except Exception as e:
            print(f"    ❌ Erro: {e}")
            return False

# MAIN
def main():
    print("\n╔════════════════════════════════════════════════════════════════╗")
    print("║  SINCRONIZAÇÃO FINAL: ML → Tiny                              ║")
    print("║  " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "                              ║")
    print("╚════════════════════════════════════════════════════════════════╝")

    # Abrir planilha
    input_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
    print(f"\n📂 Abrindo: {input_file.name}")

    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb.active

    # Encontrar colunas
    title_col = None
    price_col = None

    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col_idx).value
        if cell_value == "TITLE":
            title_col = col_idx
        elif cell_value == "PRICE":
            price_col = col_idx

    print(f"✅ Título: coluna {title_col}")
    print(f"✅ Preço: coluna {price_col}")

    # Cliente
    client = TinyClient()

    # Carregar produtos
    client.load_all_products()

    if not client.products:
        print("❌ Nenhum produto carregado!")
        return

    # Criar saída
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active

    cols = ["Linha", "Título ML", "Preço ML", "ID Tiny", "Descrição Tiny", "Preço Tiny", "Diferença", "Novo Preço", "Status"]
    for i, col in enumerate(cols, 1):
        ws_out.cell(1, i, col)

    # Estilo cabeçalho
    for col in range(1, len(cols) + 1):
        cell = ws_out.cell(1, col)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Processar
    print(f"\n🔍 Processando anúncios...")

    stats = {
        "total": 0,
        "found": 0,
        "cheaper": 0,
        "updated": 0,
    }

    out_row = 2

    for in_row in range(6, min(6 + 89, ws.max_row + 1)):
        title_ml = ws.cell(in_row, title_col).value
        price_ml = ws.cell(in_row, price_col).value

        if not title_ml or not price_ml:
            continue

        title_ml = str(title_ml).strip()
        price_ml = float(price_ml)

        if price_ml <= 0:
            continue

        stats["total"] += 1

        # Buscar
        prod = client.find_product_by_title(title_ml, threshold=0.35)

        if not prod:
            if stats["total"] % 20 == 0:
                print(f"  [{in_row}] Não encontrado")
            continue

        stats["found"] += 1

        pid = prod.get("id")
        desc = prod.get("descricao", "")[:50]
        price_tiny = prod.get("precos", {}).get("preco", 0)

        # Verificar mais barato
        if price_tiny > 0 and price_tiny < price_ml:
            stats["cheaper"] += 1

            new_price = price_ml + 0.01
            diff = price_ml - price_tiny

            # Atualizar
            ok = client.update_price(pid, new_price)

            if ok:
                stats["updated"] += 1
                status = "✅ OK"
            else:
                status = "❌ ERRO"

            # Registrar
            ws_out.cell(out_row, 1, in_row)
            ws_out.cell(out_row, 2, title_ml[:40])
            ws_out.cell(out_row, 3, price_ml)
            ws_out.cell(out_row, 4, pid)
            ws_out.cell(out_row, 5, desc)
            ws_out.cell(out_row, 6, price_tiny)
            ws_out.cell(out_row, 7, diff)
            ws_out.cell(out_row, 8, new_price)
            ws_out.cell(out_row, 9, status)

            print(f"  ⚠️  [{in_row}] {title_ml[:40]}")
            print(f"      ML R$ {price_ml:.2f} > Tiny R$ {price_tiny:.2f} → Corrigido: R$ {new_price:.2f} {status}")

            out_row += 1
        else:
            if stats["total"] % 20 == 0:
                print(f"  [{in_row}] OK")

        time.sleep(0.3)

    # Salvar
    output = Path(r"C:\Users\user\Downloads\Relatorio_Precos_Inconsistentes.xlsx")
    print(f"\n💾 Salvando: {output.name}")
    wb_out.save(output)
    wb.close()

    # Resultado
    print(f"\n{'='*80}")
    print(f"✅ CONCLUÍDO")
    print(f"{'='*80}")
    print(f"\nANÚNCIOS PROCESSADOS: {stats['total']}")
    print(f"ENCONTRADOS NO TINY: {stats['found']}")
    print(f"MAIS BARATOS NO TINY: {stats['cheaper']} ⚠️")
    print(f"ATUALIZADOS: {stats['updated']}")
    print(f"\nArquivo: {output}")
    print(f"Requisições API: {client.req_count}")

if __name__ == "__main__":
    main()
