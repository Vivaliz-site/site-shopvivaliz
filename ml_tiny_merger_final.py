#!/usr/bin/env python3
"""
Merger completo: Mercado Livre + Tiny/Olist ERP
Busca preços do Tiny para cada anúncio do ML na mesma linha
"""
import os
import json
import sys
from pathlib import Path
from typing import Dict, Optional, List, Tuple
from datetime import datetime

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# ═══════════════════════════════════════════════════════════════════════════════
# CARREGAMENTO DE AMBIENTE
# ═══════════════════════════════════════════════════════════════════════════════

env_files = [
    Path(r"C:\site-shopvivaliz\.env.local"),
    Path(r"C:\site-shopvivaliz\.env"),
    Path("/home/ubuntu/site-shopvivaliz/.env"),
]

for f in env_files:
    if f.exists():
        load_dotenv(f)
        print(f"✅ Variáveis carregadas de {f}")
        break

# ═══════════════════════════════════════════════════════════════════════════════
# CLIENT API TINY
# ═══════════════════════════════════════════════════════════════════════════════

class TinyAPIClient:
    def __init__(self):
        self.access_token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")
        if not self.access_token:
            print("❌ ERRO: Token não configurado")
            sys.exit(1)

        self.base_url = "https://api.tiny.com.br/public-api/v3"
        self.session = requests.Session()
        retry = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
        adapter = HTTPAdapter(max_retries=retry)
        self.session.mount("https://", adapter)

        self.cache = {}
        self.request_count = 0

        print(f"✅ Cliente API inicializado")
        print(f"   Base URL: {self.base_url}")

    def _make_request(self, method: str, endpoint: str, params: Dict = None) -> Optional[Dict]:
        """Faz requisição com retry"""
        self.request_count += 1

        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": "ShopVivaliz/ML-Merger"
        }

        url = f"{self.base_url}{endpoint}"

        try:
            if method == "GET":
                response = self.session.get(url, headers=headers, params=params, timeout=15)
            else:
                response = self.session.request(method, url, headers=headers, timeout=15)

            if response.status_code == 200:
                return response.json()
            else:
                print(f"   ⚠️  HTTP {response.status_code} em {endpoint}")
                return None

        except Exception as e:
            print(f"   ❌ Erro na requisição: {e}")
            return None

    def search_products(self, limit: int = 100, offset: int = 0) -> Optional[Dict]:
        """Lista produtos com paginação"""
        return self._make_request("GET", "/produtos", params={"limit": limit, "offset": offset})

    def find_product_by_sku(self, sku: str) -> Optional[Dict]:
        """Busca produto por SKU"""
        if sku in self.cache:
            return self.cache[sku]

        # Buscar produtos
        result = self.search_products(limit=100)
        if not result or "itens" not in result:
            return None

        for product in result.get("itens", []):
            if product.get("sku", "").strip().lower() == sku.strip().lower():
                self.cache[sku] = product
                return product

        # Se não encontrou na primeira página, tenta paginação
        pagination = result.get("paginacao", {})
        total = pagination.get("total", 0)
        limit = pagination.get("limit", 100)

        for offset in range(limit, total, limit):
            result = self.search_products(limit=limit, offset=offset)
            if not result or "itens" not in result:
                continue

            for product in result.get("itens", []):
                if product.get("sku", "").strip().lower() == sku.strip().lower():
                    self.cache[sku] = product
                    return product

        return None

    def get_product_by_id(self, product_id: int) -> Optional[Dict]:
        """Busca produto pelo ID do Tiny"""
        return self._make_request("GET", f"/produtos/{product_id}")

# ═══════════════════════════════════════════════════════════════════════════════
# PROCESSAMENTO DE PLANILHA
# ═══════════════════════════════════════════════════════════════════════════════

def process_ml_spreadsheet():
    """Processa planilha do Mercado Livre e enriquece com dados do Tiny"""

    input_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
    if not input_file.exists():
        print(f"❌ Arquivo não encontrado: {input_file}")
        sys.exit(1)

    print(f"\n📂 Abrindo planilha: {input_file.name}")
    wb_in = openpyxl.load_workbook(input_file, data_only=True)
    ws_in = wb_in.active

    # Análise de colunas - linha 1 tem os nomes técnicos
    header = []
    for col_idx in range(1, ws_in.max_column + 1):
        header.append(ws_in.cell(1, col_idx).value)

    # Dados começam na linha 6 (linhas 2-5 são metadados do Mercado Livre)
    data_start_row = 6
    data_rows = ws_in.max_row - data_start_row + 1

    print(f"✅ Planilha carregada: {data_rows} anúncios, {len(header)} colunas")
    print(f"\n📋 Colunas originais:")
    for idx, col in enumerate(header, 1):
        print(f"  [{idx:2d}] {col}")

    # Identificar coluna de preço do ML e coluna de SKU/código
    price_col_idx = None
    sku_col_idx = None
    product_number_col_idx = None

    for idx, col_name in enumerate(header, 1):
        col_lower = str(col_name).lower() if col_name else ""
        if "price" in col_lower and price_col_idx is None:
            price_col_idx = idx
        if "product_number" in col_lower.replace("_", ""):
            product_number_col_idx = idx

    print(f"\n🔍 Colunas identificadas:")
    print(f"  Coluna de Preço ML: {price_col_idx} ({header[price_col_idx-1] if price_col_idx else 'NÃO ENCONTRADA'})")
    print(f"  Coluna de PRODUCT_NUMBER: {product_number_col_idx} ({header[product_number_col_idx-1] if product_number_col_idx else 'NÃO ENCONTRADA'})")

    # Inicializar cliente API
    print(f"\n🔗 Conectando à API Tiny...")
    client = TinyAPIClient()

    # Criar nova planilha
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Anúncios com Preços"

    # Cabeçalho original + novas colunas
    new_columns = [
        "ID produto Tiny",
        "SKU Tiny",
        "Descrição Tiny",
        "Preço Mercado Livre original",
        "Preço cadastro Tiny",
        "Preço tabela PDV",
        "Preço tabela Shopee",
        "Preço tabela Amazon",
        "Preço tabela TikTok",
        "Status da localização",
        "Critério utilizado",
        "Observação"
    ]

    # Escrever cabeçalho
    for col_idx, col_name in enumerate(header, 1):
        ws_out.cell(1, col_idx, col_name)

    for col_idx, col_name in enumerate(new_columns, len(header) + 1):
        ws_out.cell(1, col_idx, col_name)

    print(f"✅ Cabeçalho criado com {len(header) + len(new_columns)} colunas")

    # Estilo para cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(header) + len(new_columns) + 1):
        cell = ws_out.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Processamento de linhas
    print(f"\n🔍 Processando {data_rows} anúncios...")

    stats = {
        "total": 0,
        "localized": 0,
        "not_localized": 0,
        "by_product_number": 0,
        "by_sku": 0,
        "prices_found": 0,
    }

    for out_row_idx, in_row_idx in enumerate(range(data_start_row, ws_in.max_row + 1), 2):
        # Copiar linhas originais
        for col_idx in range(1, len(header) + 1):
            value = ws_in.cell(in_row_idx, col_idx).value
            ws_out.cell(out_row_idx, col_idx, value)

        # Extrair identificadores
        product_number = ws_in.cell(in_row_idx, product_number_col_idx).value if product_number_col_idx else None
        price_ml = ws_in.cell(in_row_idx, price_col_idx).value if price_col_idx else None

        stats["total"] += 1

        # Buscar produto no Tiny
        tiny_product = None
        search_criterion = None

        if product_number:
            tiny_product = client.find_product_by_sku(str(product_number).strip())
            if tiny_product:
                search_criterion = "PRODUCT_NUMBER exato"
                stats["by_product_number"] += 1

        # Se não encontrou, tentar por SKU (se houver coluna diferente)
        if not tiny_product and "sku" in str(header).lower():
            for idx, col_name in enumerate(header, 1):
                if "sku" in str(col_name).lower() and idx != product_number_col_idx:
                    sku = ws_in.cell(in_row_idx, idx).value
                    if sku:
                        tiny_product = client.find_product_by_sku(str(sku).strip())
                        if tiny_product:
                            search_criterion = f"SKU exato (coluna {col_name})"
                            stats["by_sku"] += 1
                            break

        # Preencher colunas de resultado
        out_col = len(header) + 1

        if tiny_product:
            stats["localized"] += 1

            # ID produto Tiny
            ws_out.cell(out_row_idx, out_col, tiny_product.get("id"))
            out_col += 1

            # SKU Tiny
            ws_out.cell(out_row_idx, out_col, tiny_product.get("sku"))
            out_col += 1

            # Descrição Tiny
            ws_out.cell(out_row_idx, out_col, tiny_product.get("descricao"))
            out_col += 1

            # Preço ML original
            ws_out.cell(out_row_idx, out_col, price_ml)
            out_col += 1

            # Preço cadastro Tiny
            precos = tiny_product.get("precos", {})
            preco_cadastro = precos.get("preco")
            ws_out.cell(out_row_idx, out_col, preco_cadastro)
            if preco_cadastro:
                stats["prices_found"] += 1
            out_col += 1

            # Tabelas de preços (vazias - endpoint não localizado)
            for _ in range(4):
                ws_out.cell(out_row_idx, out_col, None)
                out_col += 1

            # Status
            ws_out.cell(out_row_idx, out_col, "LOCALIZADO")
            out_col += 1

            # Critério
            ws_out.cell(out_row_idx, out_col, search_criterion)
            out_col += 1

            # Observação
            obs = ""
            if preco_cadastro is None:
                obs += "Sem preço cadastrado no Tiny. "
            ws_out.cell(out_row_idx, out_col, obs.strip())

            if out_row_idx % 10 == 0:
                print(f"  ✓ Linha {out_row_idx}: {tiny_product.get('descricao', 'Sem descrição')[:40]}...")

        else:
            stats["not_localized"] += 1

            # Preencher apenas com preço ML e status
            out_col += 3  # Pular ID, SKU, Descrição

            # Preço ML original
            ws_out.cell(out_row_idx, out_col, price_ml)
            out_col += 1

            # Preço cadastro e tabelas (vazias)
            for _ in range(5):
                ws_out.cell(out_row_idx, out_col, None)
                out_col += 1

            # Status
            ws_out.cell(out_row_idx, out_col, "NÃO LOCALIZADO")
            out_col += 1

            # Critério
            ws_out.cell(out_row_idx, out_col, "Produto não encontrado no Tiny")
            out_col += 1

            # Observação
            id_value = product_number or f"Linha {out_row_idx}"
            ws_out.cell(out_row_idx, out_col, f"Verificar identificação: {id_value}")

    # Formatação final
    print(f"\n✨ Formatando planilha...")

    # Congelar cabeçalho
    ws_out.freeze_panes = "A2"

    # Filtros
    ws_out.auto_filter.ref = f"A1:{get_column_letter(len(header) + len(new_columns))}{data_rows + 1}"

    # Ajustar largura das colunas
    for col in range(1, len(header) + len(new_columns) + 1):
        ws_out.column_dimensions[get_column_letter(col)].width = 15

    # Salvar
    output_file = Path(r"C:\Users\user\Downloads\Anuncios-ML-com-precos-Tiny.xlsx")
    print(f"\n💾 Salvando arquivo: {output_file.name}")
    wb_out.save(output_file)

    wb_in.close()
    wb_out.close()

    # Relatório final
    print(f"\n" + "=" * 80)
    print(f"✅ PROCESSAMENTO CONCLUÍDO")
    print(f"=" * 80)
    print(f"\n📊 RELATÓRIO:")
    print(f"  Total de anúncios processados: {stats['total']}")
    print(f"  Anúncios localizados no Tiny: {stats['localized']} ({stats['localized']/stats['total']*100:.1f}%)")
    print(f"  Anúncios NÃO localizados: {stats['not_localized']} ({stats['not_localized']/stats['total']*100:.1f}%)")
    print(f"  Localizados por PRODUCT_NUMBER: {stats['by_product_number']}")
    print(f"  Localizados por SKU: {stats['by_sku']}")
    print(f"  Produtos com preço cadastrado: {stats['prices_found']}")
    print(f"\n📁 Arquivo gerado:")
    print(f"  {output_file}")
    print(f"  Tamanho: {output_file.stat().st_size / 1024:.1f} KB")
    print(f"\n📈 Requisições API: {client.request_count}")
    print(f"\n⚠️  NOTA: Colunas de tabelas de preços (PDV, Shopee, Amazon, TikTok) estão vazias.")
    print(f"  O endpoint correspondente na API Tiny v3 não foi localizado.")
    print(f"  Você pode preencher manualmente ou consultar outro endpoint.")
    print(f"\n✅ Preço do Mercado Livre original foi preservado na coluna específica.")


if __name__ == "__main__":
    print("╔════════════════════════════════════════════════════════════════╗")
    print("║  Merger: Mercado Livre + Preços do Tiny ERP                   ║")
    print("║  Data: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "                              ║")
    print("╚════════════════════════════════════════════════════════════════╝")

    process_ml_spreadsheet()
