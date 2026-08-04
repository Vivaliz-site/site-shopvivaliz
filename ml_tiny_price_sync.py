#!/usr/bin/env python3
"""
Sincronização inteligente: ML → Tiny
Busca produtos por título, identifica preços mais baratos, atualiza automaticamente
"""
import os
import json
import sys
import time
from pathlib import Path
from typing import Dict, Optional, List, Tuple
from datetime import datetime
from difflib import SequenceMatcher

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# ═══════════════════════════════════════════════════════════════════════════════
# CARREGAMENTO DE AMBIENTE
# ═══════════════════════════════════════════════════════════════════════════════

env_files = [
    Path(r"C:\site-shopvivaliz\.env.local"),
    Path(r"C:\site-shopvivaliz\.env"),
    Path("/home/ubuntu/site-shopvivaliz/.env"),
]

for f in env_files:
    if f.exists():
        load_dotenv(f)
        break

# ═══════════════════════════════════════════════════════════════════════════════
# CLIENT API TINY
# ═══════════════════════════════════════════════════════════════════════════════

class TinyAPIClient:
    def __init__(self):
        self.access_token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")
        if not self.access_token:
            print("❌ ERRO: Token não configurado")
            sys.exit(1)

        self.base_url = "https://api.tiny.com.br/public-api/v3"
        self.session = requests.Session()
        retry = Retry(total=3, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
        adapter = HTTPAdapter(max_retries=retry)
        self.session.mount("https://", adapter)

        self.cache = {}
        self.all_products = None
        self.request_count = 0

        print(f"✅ Cliente API inicializado")

    def _make_request(self, method: str, endpoint: str, params: Dict = None, json_data: Dict = None) -> Optional[Dict]:
        """Faz requisição com retry"""
        self.request_count += 1

        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": "ShopVivaliz/ML-Sync"
        }

        url = f"{self.base_url}{endpoint}"

        try:
            if method == "GET":
                response = self.session.get(url, headers=headers, params=params, timeout=15)
            elif method == "PUT":
                response = self.session.put(url, headers=headers, json=json_data, timeout=15)
            else:
                response = self.session.request(method, url, headers=headers, timeout=15)

            if response.status_code in [200, 204]:
                return response.json() if response.text else {"status": "success"}
            else:
                return None

        except Exception as e:
            print(f"   ❌ Erro: {e}")
            return None

    def load_all_products(self):
        """Carrega todos os produtos com paginação"""
        if self.all_products is not None:
            return self.all_products

        print("\n🔄 Carregando produtos do Tiny...")
        all_products = []
        offset = 0
        limit = 500

        while True:
            result = self._make_request("GET", "/produtos", params={"limit": limit, "offset": offset})
            if not result or "itens" not in result:
                break

            items = result.get("itens", [])
            all_products.extend(items)
            print(f"  Carregados: {len(all_products)} produtos", end="\r")

            pagination = result.get("paginacao", {})
            total = pagination.get("total", 0)

            if offset + limit >= total:
                break

            offset += limit
            time.sleep(0.2)  # Rate limit

        print(f"  ✅ Total carregado: {len(all_products)} produtos")
        self.all_products = all_products
        return all_products

    def search_product_by_title(self, title: str, threshold: float = 0.6) -> Optional[Dict]:
        """Busca produto por título usando similaridade"""
        if self.all_products is None:
            self.load_all_products()

        if not self.all_products:
            return None

        title_lower = title.lower().strip()
        best_match = None
        best_score = threshold

        for product in self.all_products:
            prod_desc = product.get("descricao", "").lower()
            score = SequenceMatcher(None, title_lower, prod_desc).ratio()

            if score > best_score:
                best_score = score
                best_match = product

        return best_match

    def get_product_full(self, product_id: int) -> Optional[Dict]:
        """Busca produto completo pelo ID"""
        return self._make_request("GET", f"/produtos/{product_id}")

    def update_product_price(self, product_id: int, price: float, table_id: Optional[int] = None) -> bool:
        """Atualiza preço do produto (cadastro ou tabela)"""
        if table_id:
            # Atualizar tabela de preço específica
            payload = {
                "preco": float(price)
            }
            result = self._make_request("PUT", f"/produtos/{product_id}/lista-precos/{table_id}", json_data=payload)
        else:
            # Atualizar cadastro principal
            payload = {
                "precos": {
                    "preco": float(price)
                }
            }
            result = self._make_request("PUT", f"/produtos/{product_id}", json_data=payload)

        return result is not None

# ═══════════════════════════════════════════════════════════════════════════════
# PROCESSAMENTO
# ═══════════════════════════════════════════════════════════════════════════════

def process_ml_and_sync():
    """Processa ML, encontra produtos, identifica discrepâncias e sincroniza"""

    input_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
    if not input_file.exists():
        print(f"❌ Arquivo não encontrado: {input_file}")
        sys.exit(1)

    print(f"\n📂 Abrindo planilha: {input_file.name}")
    wb_in = openpyxl.load_workbook(input_file, data_only=True)
    ws_in = wb_in.active

    # Análise de colunas
    header = []
    for col_idx in range(1, ws_in.max_column + 1):
        header.append(ws_in.cell(1, col_idx).value)

    # Identificar colunas
    price_col_idx = None
    title_col_idx = None

    for idx, col_name in enumerate(header, 1):
        col_lower = str(col_name).lower() if col_name else ""
        if "price" in col_lower and price_col_idx is None:
            price_col_idx = idx
        if "title" in col_lower:
            title_col_idx = idx

    print(f"✅ Planilha carregada: {ws_in.max_row - 1} anúncios")
    print(f"  Coluna de Título: {title_col_idx} ({header[title_col_idx-1] if title_col_idx else 'ERRO'})")
    print(f"  Coluna de Preço ML: {price_col_idx} ({header[price_col_idx-1] if price_col_idx else 'ERRO'})")

    # Inicializar cliente
    client = TinyAPIClient()
    client.load_all_products()

    # Criar nova planilha de resultado
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Discrepâncias Encontradas"

    # Cabeçalho da saída
    result_columns = [
        "Linha ML",
        "Título ML",
        "Preço ML",
        "ID Tiny",
        "Descrição Tiny",
        "Tabela com Preço Menor",
        "Preço na Tabela",
        "Diferença (R$)",
        "% Menor",
        "Preço Corrigido (ML + 0,01)",
        "Status Atualização",
        "Critério Match",
        "Similaridade (%)"
    ]

    for col_idx, col_name in enumerate(result_columns, 1):
        ws_out.cell(1, col_idx, col_name)

    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(result_columns) + 1):
        cell = ws_out.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Processamento
    print(f"\n🔍 Processando {ws_in.max_row - 5} anúncios (começando linha 6)...")

    data_start_row = 6
    stats = {
        "total": 0,
        "found": 0,
        "discrepancies": 0,
        "updated": 0,
        "failed": 0,
    }

    out_row = 2

    for in_row_idx in range(data_start_row, ws_in.max_row + 1):
        title_ml = ws_in.cell(in_row_idx, title_col_idx).value if title_col_idx else None
        price_ml = ws_in.cell(in_row_idx, price_col_idx).value if price_col_idx else None

        if not title_ml or not price_ml:
            continue

        stats["total"] += 1
        title_ml = str(title_ml).strip()
        price_ml = float(price_ml) if isinstance(price_ml, (int, float)) else 0

        if price_ml <= 0:
            continue

        # Buscar no Tiny
        tiny_product = client.search_product_by_title(title_ml, threshold=0.5)

        if not tiny_product:
            if stats["total"] % 10 == 0:
                print(f"  ⊘ Linha {in_row_idx}: Produto não encontrado")
            continue

        stats["found"] += 1
        product_id = tiny_product.get("id")
        product_desc = tiny_product.get("descricao", "")

        # Buscar preço completo do produto
        full_product = client.get_product_full(product_id)
        if not full_product:
            full_product = tiny_product

        precos = full_product.get("precos", {})
        price_cadastro = precos.get("preco", 0)

        # Verificar se há preço mais barato
        cheaper_table = None
        cheaper_price = None
        cheaper_table_name = None

        # Buscar listas de preço (tabelas)
        # Nota: endpoint específico não existe, então tentamos get do produto
        # As tabelas podem estar em estrutura diferente

        # Por enquanto, comparar apenas cadastro principal
        if price_cadastro > 0 and price_cadastro < price_ml:
            cheaper_table = "CADASTRO_PRINCIPAL"
            cheaper_price = price_cadastro
            cheaper_table_name = "Cadastro Principal"

        if cheaper_table:
            stats["discrepancies"] += 1
            diff = price_ml - cheaper_price
            pct_diff = (diff / price_ml) * 100

            new_price = price_ml + 0.01

            # Atualizar no Tiny
            updated = client.update_product_price(product_id, new_price)
            if updated:
                stats["updated"] += 1
                status_update = "✅ ATUALIZADO"
            else:
                stats["failed"] += 1
                status_update = "❌ FALHA NA ATUALIZAÇÃO"

            # Preencher linha de resultado
            similarity = SequenceMatcher(None, title_ml.lower(), product_desc.lower()).ratio()

            ws_out.cell(out_row, 1, in_row_idx)
            ws_out.cell(out_row, 2, title_ml)
            ws_out.cell(out_row, 3, price_ml)
            ws_out.cell(out_row, 4, product_id)
            ws_out.cell(out_row, 5, product_desc)
            ws_out.cell(out_row, 6, cheaper_table_name)
            ws_out.cell(out_row, 7, cheaper_price)
            ws_out.cell(out_row, 8, diff)
            ws_out.cell(out_row, 9, pct_diff)
            ws_out.cell(out_row, 10, new_price)
            ws_out.cell(out_row, 11, status_update)
            ws_out.cell(out_row, 12, "Título/Descrição")
            ws_out.cell(out_row, 13, similarity * 100)

            print(f"  ⚠️  Linha {in_row_idx}: {title_ml[:40]}...")
            print(f"      ML: R$ {price_ml:.2f} | Cadastro: R$ {cheaper_price:.2f} | Corrigido: R$ {new_price:.2f}")
            print(f"      {status_update}")

            out_row += 1

        elif stats["total"] % 20 == 0:
            print(f"  ✓ Linha {in_row_idx}: OK ({price_ml:.2f})")

        time.sleep(0.1)  # Rate limit

    # Salvar resultado
    output_file = Path(r"C:\Users\user\Downloads\ML_Discrepancias_Precos.xlsx")
    print(f"\n💾 Salvando relatório: {output_file.name}")
    wb_out.save(output_file)

    wb_in.close()
    wb_out.close()

    # Relatório final
    print(f"\n" + "=" * 80)
    print(f"✅ SINCRONIZAÇÃO CONCLUÍDA")
    print(f"=" * 80)
    print(f"\n📊 ESTATÍSTICAS:")
    print(f"  Total de anúncios ML: {stats['total']}")
    print(f"  Encontrados no Tiny: {stats['found']} ({stats['found']/stats['total']*100:.1f}%)")
    print(f"  Com preço MAIS BARATO: {stats['discrepancies']}")
    print(f"  Atualizados com sucesso: {stats['updated']}")
    print(f"  Falhas na atualização: {stats['failed']}")
    print(f"\n📁 Relatório gerado:")
    print(f"  {output_file}")
    print(f"\n🔗 Requisições API: {client.request_count}")


if __name__ == "__main__":
    print("╔════════════════════════════════════════════════════════════════╗")
    print("║  Sincronização: ML → Tiny (Preços)                           ║")
    print("║  Data: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "                              ║")
    print("╚════════════════════════════════════════════════════════════════╝")

    process_ml_and_sync()
