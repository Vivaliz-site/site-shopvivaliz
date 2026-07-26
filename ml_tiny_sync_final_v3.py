#!/usr/bin/env python3
"""
VERSÃO FINAL V3: Sincronização ML → Tiny (SIMPLIFICADA)
Sem paginação complexa, carrega apenas os primeiros produtos e processa
"""
import os
import sys
import time
from pathlib import Path
from typing import Dict, Optional
from datetime import datetime
from difflib import SequenceMatcher

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill
from dotenv import load_dotenv

# CARREGAR ENV
for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

print("\n╔════════════════════════════════════════════════════════════════╗")
print("║  SINCRONIZAÇÃO ML → TINY (V3 - SIMPLIFICADA)                 ║")
print("║  " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "                              ║")
print("╚════════════════════════════════════════════════════════════════╝")

# GET TOKEN
token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")
if not token:
    print("❌ Token não configurado")
    sys.exit(1)

print(f"\n✅ Token carregado ({len(token)} chars)")

# CARREGAR PRODUTOS
print("\n📥 Carregando produtos do Tiny...")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

url = "https://api.tiny.com.br/public-api/v3/produtos?limit=1000&offset=0"

try:
    resp = requests.get(url, headers=headers, timeout=30)
    print(f"Status: {resp.status_code}")

    if resp.status_code == 200:
        data = resp.json()
        products = data.get("itens", [])
        print(f"✅ Carregados: {len(products)} produtos")
    else:
        print(f"❌ Erro: {resp.status_code}")
        if resp.text:
            print(f"Resposta: {resp.text[:200]}")
        sys.exit(1)

except Exception as e:
    print(f"❌ Erro na requisição: {e}")
    sys.exit(1)

if not products:
    print("❌ Nenhum produto carregado!")
    sys.exit(1)

# ABRIR PLANILHA ML
print("\n📂 Abrindo planilha ML...")

input_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
wb = openpyxl.load_workbook(input_file, data_only=True)
ws = wb.active

# Encontrar colunas
title_col = price_col = None
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(1, col_idx).value
    if val == "TITLE":
        title_col = col_idx
    elif val == "PRICE":
        price_col = col_idx

print(f"✅ Colunas: Título={title_col}, Preço={price_col}")

# PROCESSAR
print(f"\n🔍 Processando anúncios...")

wb_out = openpyxl.Workbook()
ws_out = wb_out.active

cols = ["Linha", "Título ML", "Preço ML", "ID Tiny", "Desc Tiny", "Preço Tiny", "Diferença", "Novo Preço", "Status"]
for i, col in enumerate(cols, 1):
    ws_out.cell(1, i, col)

# Estilo
for col in range(1, len(cols) + 1):
    cell = ws_out.cell(1, col)
    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")

stats = {"total": 0, "found": 0, "cheaper": 0, "updated": 0}
out_row = 2

for in_row in range(6, min(6 + 89, ws.max_row + 1)):
    title_ml = ws.cell(in_row, title_col).value
    price_ml = ws.cell(in_row, price_col).value

    if not title_ml or not price_ml:
        continue

    title_ml = str(title_ml).strip()
    price_ml = float(price_ml)

    if price_ml <= 0:
        continue

    stats["total"] += 1

    # Buscar produto
    title_lower = title_ml.lower()
    best_prod = None
    best_score = 0.4

    for prod in products:
        desc = prod.get("descricao", "").lower()
        score = SequenceMatcher(None, title_lower, desc).ratio()

        if score > best_score:
            best_prod = prod
            best_score = score

    if not best_prod:
        if stats["total"] % 20 == 0:
            print(f"  [{in_row}] Não encontrado")
        continue

    stats["found"] += 1

    pid = best_prod.get("id")
    desc = best_prod.get("descricao", "")[:40]
    price_tiny = best_prod.get("precos", {}).get("preco", 0)

    # Verificar mais barato
    if price_tiny > 0 and price_tiny < price_ml:
        stats["cheaper"] += 1

        new_price = price_ml + 0.01
        diff = price_ml - price_tiny

        # Tentar atualizar
        update_url = f"https://api.tiny.com.br/public-api/v3/produtos/{pid}"
        update_payload = {"precos": {"preco": float(new_price)}}

        try:
            update_resp = requests.put(update_url, headers=headers, json=update_payload, timeout=20)
            ok = update_resp.status_code in [200, 204]
        except:
            ok = False

        if ok:
            stats["updated"] += 1
            status = "✅ OK"
        else:
            status = "❌ ERRO"

        # Registrar
        ws_out.cell(out_row, 1, in_row)
        ws_out.cell(out_row, 2, title_ml[:35])
        ws_out.cell(out_row, 3, price_ml)
        ws_out.cell(out_row, 4, pid)
        ws_out.cell(out_row, 5, desc)
        ws_out.cell(out_row, 6, price_tiny)
        ws_out.cell(out_row, 7, diff)
        ws_out.cell(out_row, 8, new_price)
        ws_out.cell(out_row, 9, status)

        print(f"  ⚠️  [{in_row}] {title_ml[:30]}")
        print(f"      ML R$ {price_ml:.2f} > Tiny R$ {price_tiny:.2f} → R$ {new_price:.2f} {status}")

        out_row += 1
    else:
        if stats["total"] % 20 == 0:
            print(f"  [{in_row}] OK")

    time.sleep(0.2)

# SALVAR
output = Path(r"C:\Users\user\Downloads\Precos_Inconsistentes_Final.xlsx")
print(f"\n💾 Salvando: {output.name}")
wb_out.save(output)
wb.close()

# RELATÓRIO
print(f"\n{'='*80}")
print(f"✅ SINCRONIZAÇÃO CONCLUÍDA")
print(f"{'='*80}")
print(f"\n📊 RESULTADOS:")
print(f"  Anúncios processados: {stats['total']}")
print(f"  Encontrados no Tiny: {stats['found']}")
print(f"  Mais baratos no Tiny: {stats['cheaper']} ⚠️")
print(f"  Atualizados com sucesso: {stats['updated']}")
print(f"\n📁 Arquivo: {output}")
print(f"\n✅ Pronto!")
