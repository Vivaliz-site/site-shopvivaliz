#!/usr/bin/env python3
"""
PREENCHER: Apenas os 4 preços de tabelas no RELATORIO_6_PRECOS_COMPLETO.xlsx
"""
import os
import time
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n╔════════════════════════════════════════════════════════════════════════════════╗")
print("║  PREENCHER: Preços das Tabelas (PDV, Shopee, Amazon, TikTok)                  ║")
print("╚════════════════════════════════════════════════════════════════════════════════╝")

# IDs das tabelas
TABELAS = {"PDV": 982, "Shopee": 989, "Amazon": 991, "TikTok": 990}

# Abrir arquivo
print("\n📂 Abrindo arquivo...")
file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COMPLETO.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"✅ {file_path.name} aberto")
print(f"   Linhas: {ws.max_row - 1} produtos")

# Processar cada linha
print("\n🔄 Buscando preços...")

stats = {"processado": 0, "com_preco": 0}

for row_idx in range(2, ws.max_row + 1):
    id_tiny_cell = ws.cell(row_idx, 3)
    id_tiny = id_tiny_cell.value

    if not id_tiny:
        continue

    try:
        id_tiny = int(id_tiny)
    except:
        continue

    stats["processado"] += 1

    # Buscar preços das 4 tabelas
    precos = {}

    for nome_tabela, id_tabela in TABELAS.items():
        try:
            url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}?produto_id={id_tiny}"
            resp = requests.get(url, headers=headers, timeout=10)

            if resp.status_code == 200:
                data = resp.json()
                excecoes = data.get("excecoes", [])

                for exc in excecoes:
                    if exc.get("idProduto") == id_tiny:
                        preco = exc.get("preco")
                        if preco:
                            precos[nome_tabela] = preco
                            stats["com_preco"] += 1
                        break

        except:
            pass

        time.sleep(0.2)

    # Preencher colunas
    ws.cell(row_idx, 8, precos.get("PDV"))        # Col H: PDV
    ws.cell(row_idx, 9, precos.get("Shopee"))     # Col I: Shopee
    ws.cell(row_idx, 10, precos.get("Amazon"))    # Col J: Amazon
    ws.cell(row_idx, 11, precos.get("TikTok"))    # Col K: TikTok

    # Formatar como moeda
    for col in [8, 9, 10, 11]:
        ws.cell(row_idx, col).number_format = 'R$ #,##0.00'

    if stats["processado"] % 20 == 0:
        print(f"  [{stats['processado']:2d}] ID {id_tiny}: PDV={precos.get('PDV')}, Shopee={precos.get('Shopee')}")

# Salvar
print(f"\n💾 Salvando...")
try:
    wb.save(file_path)
    wb.close()
    print(f"✅ Arquivo salvo com sucesso!")
except Exception as e:
    print(f"❌ Erro ao salvar: {e}")
    print(f"   Feche o arquivo no Excel e tente novamente")
    wb.close()
    exit(1)

# Relatório
print(f"\n{'='*100}")
print(f"✅ PREENCHIMENTO CONCLUÍDO!")
print(f"{'='*100}")
print(f"\n📊 ESTATÍSTICAS:")
print(f"  Total processado: {stats['processado']}")
print(f"  💰 Preços encontrados: {stats['com_preco']}")
print(f"\n📁 Arquivo: {file_path}")
print(f"✅ Pronto para download!")
