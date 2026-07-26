#!/usr/bin/env python3
"""
PREENCHER: Preços das tabelas no novo arquivo (sem conflito)
"""
import os
import time
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n╔════════════════════════════════════════════════════════════════════════════════╗")
print("║  PREENCHER PREÇOS: RELATORIO_6_PRECOS_COMPLETO_ATUALIZADO.xlsx                ║")
print("╚════════════════════════════════════════════════════════════════════════════════╝")

# IDs das tabelas
TABELAS = {"PDV": 982, "Shopee": 989, "Amazon": 991, "TikTok": 990}

# Abrir arquivo (novo, sem conflito)
print("\n📂 Abrindo arquivo...")
file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COMPLETO_ATUALIZADO.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"✅ {file_path.name}")
print(f"   Linhas: {ws.max_row - 1}")

# Processar
print("\n🔄 Buscando preços das tabelas...")

total = 0
com_preco = 0

for row_idx in range(2, ws.max_row + 1):
    id_tiny = ws.cell(row_idx, 3).value

    if not id_tiny:
        continue

    try:
        id_tiny = int(id_tiny)
    except:
        continue

    total += 1

    # Buscar preços
    precos = {}

    for nome_tabela, id_tabela in TABELAS.items():
        try:
            url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}?produto_id={id_tiny}"
            resp = requests.get(url, headers=headers, timeout=10)

            if resp.status_code == 200:
                data = resp.json()
                excecoes = data.get("excecoes", [])

                for exc in excecoes:
                    if exc.get("idProduto") == id_tiny:
                        preco = exc.get("preco")
                        if preco:
                            precos[nome_tabela] = preco
                            com_preco += 1
                        break

        except:
            pass

        time.sleep(0.2)

    # Preencher
    ws.cell(row_idx, 8, precos.get("PDV"))
    ws.cell(row_idx, 9, precos.get("Shopee"))
    ws.cell(row_idx, 10, precos.get("Amazon"))
    ws.cell(row_idx, 11, precos.get("TikTok"))

    for col in [8, 9, 10, 11]:
        ws.cell(row_idx, col).number_format = 'R$ #,##0.00'

    if total % 20 == 0:
        print(f"  [{total:2d}] ID {id_tiny}")

# Salvar
print(f"\n💾 Salvando...")
wb.save(file_path)
wb.close()

print(f"\n{'='*100}")
print(f"✅ SUCESSO!")
print(f"{'='*100}")
print(f"\n📊 ESTATÍSTICAS:")
print(f"  Total: {total}")
print(f"  Preços encontrados: {com_preco}")
print(f"\n📁 {file_path}")
print(f"✅ Pronto para download!")
