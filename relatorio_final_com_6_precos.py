#!/usr/bin/env python3
"""
RELATÓRIO FINAL: 6 Preços com mapeamento correto
1. Ler planilha base ML (89 produtos)
2. Usar arquivos de mapeamento Tiny (500 + 177 produtos)
3. Buscar preços das 4 tabelas via API
4. Gerar relatório completo com 6 preços
"""
import os
import json
import time
import requests
import openpyxl
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n╔════════════════════════════════════════════════════════════════════════════════╗")
print("║  RELATÓRIO FINAL: 6 PREÇOS COM MAPEAMENTO CORRETO                             ║")
print("╚════════════════════════════════════════════════════════════════════════════════╝")

# IDs das tabelas
TABELAS = {
    "PDV": 982,
    "Shopee": 989,
    "Amazon": 991,
    "TikTok": 990
}

# 1. CARREGAR ARQUIVO DE MAPEAMENTO (Tiny → ML)
print("\n📥 Carregando arquivo de mapeamento do Tiny...")

# Tentar os dois arquivos
mapping_files = [
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-48.xls',
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-50.xls',
]

mapping_dict = {}  # SKU → ID Tiny
mapping_id_dict = {}  # ID Tiny → dados

for mapping_file in mapping_files:
    if not Path(mapping_file).exists():
        continue

    try:
        df_mapping = pd.read_excel(mapping_file, sheet_name=0)

        for idx, row in df_mapping.iterrows():
            id_tiny = row.get('Id')
            sku = row.get('Produto (SKU)')
            titulo = row.get('Título')

            if id_tiny and sku:
                sku = str(sku).strip().upper()
                mapping_dict[sku] = id_tiny
                mapping_id_dict[id_tiny] = {
                    'sku': sku,
                    'titulo': titulo,
                    'id': id_tiny
                }

        print(f"✅ Carregado: {Path(mapping_file).name} ({len(df_mapping)} itens)")

    except Exception as e:
        print(f"⚠️  Erro ao carregar {mapping_file}: {e}")

print(f"   Total de mapeamentos: {len(mapping_dict)}")

# 2. CARREGAR PLANILHA BASE ML
print("\n📂 Carregando planilha base do Mercado Livre...")

ml_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
wb_ml = openpyxl.load_workbook(ml_file, data_only=True)
ws_ml = wb_ml.active

# Encontrar colunas
header_ml = {}
for col_idx in range(1, ws_ml.max_column + 1):
    col_name = ws_ml.cell(1, col_idx).value
    if col_name:
        header_ml[str(col_name).strip()] = col_idx

title_col = header_ml.get("TITLE", 5)
price_col = header_ml.get("PRICE", 9)

print(f"✅ Planilha carregada: {ml_file.name}")
print(f"   Linhas: {ws_ml.max_row}")

# 3. CARREGAR PRODUTOS DO TINY (para dados de cadastro)
print("\n📥 Carregando produtos do Tiny...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

if resp.status_code != 200:
    print(f"❌ Erro: {resp.status_code}")
    exit(1)

products_tiny = {p.get('id'): p for p in resp.json().get("itens", [])}
print(f"✅ Carregados {len(products_tiny)} produtos do Tiny")

# 4. CRIAR ARQUIVO DE SAÍDA
print("\n📝 Criando relatório final...")

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Preços Comparativos"

# CABEÇALHO
headers_out = [
    "Linha ML",
    "Título ML",
    "ID Tiny",
    "SKU Tiny",
    "Descrição Tiny",
    "Preço ML",
    "Preço Cadastro",
    "Preço PDV",
    "Preço Shopee",
    "Preço Amazon",
    "Preço TikTok",
    "Status"
]

for col_idx, header_text in enumerate(headers_out, 1):
    cell = ws_out.cell(1, col_idx, header_text)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 5. PROCESSAR DADOS
print("\n🔄 Processando...")

stats = {"total": 0, "found": 0, "with_prices": 0, "errors": 0}
out_row = 2

for in_row in range(6, min(6 + 89, ws_ml.max_row + 1)):
    title_ml = ws_ml.cell(in_row, title_col).value
    price_ml = ws_ml.cell(in_row, price_col).value

    if not title_ml or not price_ml:
        continue

    title_ml = str(title_ml).strip()
    price_ml = float(price_ml)

    if price_ml <= 0:
        continue

    stats["total"] += 1

    # Procurar produto Tiny por SKU ou similaridade de título
    id_tiny = None
    prod_tiny = None

    # Estratégia 1: Procurar nos mapeamentos existentes (por SKU aproximado)
    title_lower = title_ml.lower()

    for sku, tiny_id in mapping_dict.items():
        score = SequenceMatcher(None, title_lower, sku.lower()).ratio()
        if score > 0.5:
            id_tiny = tiny_id
            break

    # Se não encontrou, buscar nos produtos do Tiny por similaridade de título
    if not id_tiny:
        best_score = 0.4
        for t_id, prod in products_tiny.items():
            desc = prod.get("descricao", "").lower()
            score = SequenceMatcher(None, title_lower, desc).ratio()

            if score > best_score:
                id_tiny = t_id
                best_score = score

    # Obter dados do produto Tiny
    if id_tiny and id_tiny in products_tiny:
        stats["found"] += 1
        prod_tiny = products_tiny[id_tiny]
        sku_tiny = prod_tiny.get("sku", "")
        desc_tiny = prod_tiny.get("descricao", "")
        price_cadastro = prod_tiny.get("precos", {}).get("preco")

        # Buscar preços das tabelas
        precos_tabelas = {}

        for nome_tabela, id_tabela in TABELAS.items():
            try:
                url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}?produto_id={id_tiny}"
                resp = requests.get(url, headers=headers, timeout=10)

                if resp.status_code == 200:
                    data = resp.json()
                    preco = None
                    excecoes = data.get("excecoes", [])

                    for exc in excecoes:
                        if exc.get("idProduto") == id_tiny:
                            preco = exc.get("preco")
                            break

                    if preco:
                        precos_tabelas[nome_tabela] = preco
                        stats["with_prices"] += 1

            except:
                pass

            time.sleep(0.2)

        final_status = "✅ OK"
        color = "E2EFDA"

    else:
        stats["errors"] += 1
        sku_tiny = ""
        desc_tiny = ""
        price_cadastro = None
        precos_tabelas = {}
        final_status = "❌ Não encontrado"
        color = "FCE4D6"

    # Preencher linha
    ws_out.cell(out_row, 1, in_row)
    ws_out.cell(out_row, 2, title_ml[:50])
    ws_out.cell(out_row, 3, id_tiny)
    ws_out.cell(out_row, 4, sku_tiny)
    ws_out.cell(out_row, 5, desc_tiny[:50] if desc_tiny else "")
    ws_out.cell(out_row, 6, price_ml)
    ws_out.cell(out_row, 7, price_cadastro)
    ws_out.cell(out_row, 8, precos_tabelas.get("PDV"))
    ws_out.cell(out_row, 9, precos_tabelas.get("Shopee"))
    ws_out.cell(out_row, 10, precos_tabelas.get("Amazon"))
    ws_out.cell(out_row, 11, precos_tabelas.get("TikTok"))
    ws_out.cell(out_row, 12, final_status)

    # Aplicar cor
    for col_idx in range(1, 13):
        cell = ws_out.cell(out_row, col_idx)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Formatar colunas de preço
        if col_idx in [6, 7, 8, 9, 10, 11]:
            cell.number_format = 'R$ #,##0.00'

    if stats["total"] % 20 == 0:
        print(f"  [{in_row}] {title_ml[:40]}")

    out_row += 1

# Ajustar largura
widths = [10, 35, 12, 12, 35, 14, 14, 14, 14, 14, 14, 15]
for col_idx, width in enumerate(widths, 1):
    ws_out.column_dimensions[get_column_letter(col_idx)].width = width

ws_out.freeze_panes = "A2"

# SALVAR
output_file = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_FINAL.xlsx")
print(f"\n💾 Salvando: {output_file.name}")
wb_out.save(output_file)

wb_ml.close()

# RELATÓRIO
print(f"\n{'='*100}")
print(f"✅ RELATÓRIO GERADO COM SUCESSO!")
print(f"{'='*100}")
print(f"\n📊 ESTATÍSTICAS:")
print(f"  Total de anúncios: {stats['total']}")
print(f"  ✅ Encontrados no Tiny: {stats['found']}")
print(f"  💰 Com preços de tabela: {stats['with_prices']}")
print(f"  ❌ Erros: {stats['errors']}")
print(f"\n📁 Arquivo: {output_file}")
print(f"\n✅ Pronto para download!")
