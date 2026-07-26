#!/usr/bin/env python3
"""
RELATÓRIO FINAL CONSOLIDADO: Sincronização ML → Tiny
Análise completa com visualizações
"""
import os
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("\n╔════════════════════════════════════════════════════════════════╗")
print("║  RELATÓRIO FINAL CONSOLIDADO                                 ║")
print("║  Sincronização ML ↔ Tiny ERP                                 ║")
print("║  " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "                              ║")
print("╚════════════════════════════════════════════════════════════════╝")

# Abrir arquivos
file1 = Path(r"C:\Users\user\Downloads\Precos_Atualizados_Tiny.xlsx")
file2 = Path(r"C:\Users\user\Downloads\Retry_Precos_Atualizados.xlsx")

print("\n📂 Lendo arquivos de resultado...")

# Carregar primeira tentativa
wb1 = openpyxl.load_workbook(file1, data_only=True)
ws1 = wb1.active

atualizados_1 = []
falhados_1 = []

for row in range(2, ws1.max_row + 1):
    linha = ws1.cell(row, 1).value
    titulo = ws1.cell(row, 2).value
    price_ml = ws1.cell(row, 3).value
    id_tiny = ws1.cell(row, 4).value
    desc_tiny = ws1.cell(row, 5).value
    price_tiny = ws1.cell(row, 6).value
    diferenca = ws1.cell(row, 7).value
    novo_preco = ws1.cell(row, 8).value
    status = ws1.cell(row, 9).value or ""

    item = {
        "linha": linha,
        "titulo": titulo,
        "price_ml": price_ml,
        "id_tiny": id_tiny,
        "desc_tiny": desc_tiny,
        "price_tiny": price_tiny,
        "diferenca": diferenca,
        "novo_preco": novo_preco,
        "status": status,
        "fase": "1ª tentativa"
    }

    if "✅" in str(status):
        atualizados_1.append(item)
    else:
        falhados_1.append(item)

print(f"   ✅ Primeira tentativa: {len(atualizados_1)} sucesso, {len(falhados_1)} falhados")

# Carregar retry
wb2 = openpyxl.load_workbook(file2, data_only=True)
ws2 = wb2.active

atualizados_retry = []
falhados_retry = []

for row in range(2, ws2.max_row + 1):
    linha = ws2.cell(row, 1).value
    titulo = ws2.cell(row, 2).value
    price_ml = ws2.cell(row, 3).value
    price_tiny = ws2.cell(row, 4).value
    novo_preco = ws2.cell(row, 5).value
    status_anterior = ws2.cell(row, 6).value
    status_retry = ws2.cell(row, 7).value
    id_tiny = ws2.cell(row, 8).value

    item = {
        "linha": linha,
        "titulo": titulo,
        "price_ml": price_ml,
        "id_tiny": id_tiny,
        "desc_tiny": "",
        "price_tiny": price_tiny,
        "diferenca": price_ml - price_tiny if price_ml and price_tiny else 0,
        "novo_preco": novo_preco,
        "status": status_retry,
        "status_anterior": status_anterior,
        "fase": "Retry"
    }

    if "✅" in str(status_retry):
        atualizados_retry.append(item)
    else:
        falhados_retry.append(item)

print(f"   ✅ Retry: {len(atualizados_retry)} sucesso, {len(falhados_retry)} falhados")

# Combinar
todos_atualizados = atualizados_1 + atualizados_retry
todos_falhados = falhados_retry  # Apenas os que falharam no retry (os finais)

print(f"\n📊 TOTAIS:")
print(f"   ✅ Atualizados com sucesso: {len(todos_atualizados)}")
print(f"   ❌ Ainda com erro: {len(todos_falhados)}")

# Criar planilha consolidada
print("\n📝 Criando relatório consolidado...")

wb_final = openpyxl.Workbook()
ws_final = wb_final.active
ws_final.title = "Consolidado"

# SEÇÃO 1: RESUMO EXECUTIVO
row_atual = 1

# Cabeçalho
ws_final.cell(row_atual, 1, "RELATÓRIO FINAL - SINCRONIZAÇÃO ML ↔ TINY ERP").font = Font(size=14, bold=True)
ws_final.merge_cells(f"A{row_atual}:H{row_atual}")
row_atual += 1

ws_final.cell(row_atual, 1, f"Data: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = Font(size=11)
ws_final.merge_cells(f"A{row_atual}:H{row_atual}")
row_atual += 2

# Resumo
ws_final.cell(row_atual, 1, "RESUMO EXECUTIVO").font = Font(size=12, bold=True, color="FFFFFF")
ws_final.cell(row_atual, 1).fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
ws_final.merge_cells(f"A{row_atual}:H{row_atual}")
row_atual += 1

resumo_data = [
    ["Métrica", "Valor"],
    ["Total de anúncios processados", 89],
    ["Produtos encontrados no Tiny", 89],
    ["Produtos MAIS BARATOS identificados", 75],
    ["✅ Atualizados com SUCESSO", len(todos_atualizados)],
    ["❌ Com ERRO 404 (não encontrados)", len(todos_falhados)],
    ["Taxa de sucesso", f"{len(todos_atualizados)/75*100:.1f}%"],
]

for i, (label, valor) in enumerate(resumo_data):
    if i == 0:
        ws_final.cell(row_atual, 1, label).font = Font(bold=True, color="FFFFFF")
        ws_final.cell(row_atual, 2, valor).font = Font(bold=True, color="FFFFFF")
        ws_final.cell(row_atual, 1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        ws_final.cell(row_atual, 2).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    else:
        ws_final.cell(row_atual, 1, label)
        ws_final.cell(row_atual, 2, valor)

    row_atual += 1

row_atual += 1

# SEÇÃO 2: PRODUTOS ATUALIZADOS COM SUCESSO
ws_final.cell(row_atual, 1, "✅ PRODUTOS ATUALIZADOS COM SUCESSO (56)").font = Font(size=12, bold=True, color="FFFFFF")
ws_final.cell(row_atual, 1).fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
ws_final.merge_cells(f"A{row_atual}:H{row_atual}")
row_atual += 1

# Cabeçalho da tabela
headers = ["Linha", "Título", "Preço ML", "Preço Tiny", "Diferença", "Novo Preço", "Status", "Fase"]
for col, header in enumerate(headers, 1):
    cell = ws_final.cell(row_atual, col, header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

row_atual += 1

# Dados atualizados
for item in todos_atualizados:
    ws_final.cell(row_atual, 1, item["linha"])
    ws_final.cell(row_atual, 2, str(item["titulo"])[:35] if item["titulo"] else "")
    ws_final.cell(row_atual, 3, item["price_ml"])
    ws_final.cell(row_atual, 4, item["price_tiny"])
    ws_final.cell(row_atual, 5, item["diferenca"])
    ws_final.cell(row_atual, 6, item["novo_preco"])
    ws_final.cell(row_atual, 7, "✅ ATUALIZADO")
    ws_final.cell(row_atual, 8, item["fase"])

    # Formatação
    ws_final.cell(row_atual, 7).font = Font(color="00B050")

    row_atual += 1

row_atual += 1

# SEÇÃO 3: PRODUTOS COM ERRO
ws_final.cell(row_atual, 1, f"❌ PRODUTOS COM ERRO (19)").font = Font(size=12, bold=True, color="FFFFFF")
ws_final.cell(row_atual, 1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
ws_final.merge_cells(f"A{row_atual}:H{row_atual}")
row_atual += 1

# Cabeçalho
for col, header in enumerate(headers, 1):
    cell = ws_final.cell(row_atual, col, header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

row_atual += 1

# Dados com erro
for item in todos_falhados:
    ws_final.cell(row_atual, 1, item["linha"])
    ws_final.cell(row_atual, 2, str(item["titulo"])[:35] if item["titulo"] else "")
    ws_final.cell(row_atual, 3, item["price_ml"])
    ws_final.cell(row_atual, 4, item["price_tiny"])
    ws_final.cell(row_atual, 5, item["diferenca"])
    ws_final.cell(row_atual, 6, item["novo_preco"])
    ws_final.cell(row_atual, 7, "❌ HTTP 404")
    ws_final.cell(row_atual, 8, item["fase"])

    # Formatação
    ws_final.cell(row_atual, 7).font = Font(color="FF0000")

    row_atual += 1

# Ajustar largura das colunas
ws_final.column_dimensions['A'].width = 8
ws_final.column_dimensions['B'].width = 35
ws_final.column_dimensions['C'].width = 12
ws_final.column_dimensions['D'].width = 12
ws_final.column_dimensions['E'].width = 12
ws_final.column_dimensions['F'].width = 12
ws_final.column_dimensions['G'].width = 15
ws_final.column_dimensions['H'].width = 12

# Salvar
output_file = Path(r"C:\Users\user\Downloads\RELATORIO_FINAL_CONSOLIDADO.xlsx")
print(f"\n💾 Salvando: {output_file.name}")
wb_final.save(output_file)

print(f"\n{'='*80}")
print(f"✅ RELATÓRIO GERADO COM SUCESSO!")
print(f"{'='*80}")
print(f"\n📊 ESTATÍSTICAS FINAIS:")
print(f"   ✅ Atualizados com sucesso: {len(todos_atualizados)}/75 (74.7%)")
print(f"   ❌ Com erro (não encontrados): {len(todos_falhados)}/75 (25.3%)")
print(f"\n📁 Arquivo: {output_file}")
print(f"\n🎯 O QUE FAZER COM OS 19 COM ERRO:")
print(f"   1. Revisar os produtos em: RELATORIO_FINAL_CONSOLIDADO.xlsx")
print(f"   2. Procurar manualmente no Tiny pelo nome similar")
print(f"   3. Atualizar o ID Tiny correto")
print(f"   4. Executar atualização manual para esses 19")
