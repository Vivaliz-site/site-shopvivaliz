#!/usr/bin/env python3
"""
RELATÓRIO FINAL OTIMIZADO: Usar mapeamentos do Tiny direto
"""
import os
import json
import time
import requests
import openpyxl
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n╔════════════════════════════════════════════════════════════════════════════════╗")
print("║  RELATÓRIO FINAL OTIMIZADO: Mapeamento + Preços das Tabelas                   ║")
print("╚════════════════════════════════════════════════════════════════════════════════╝")

# IDs das tabelas
TABELAS = {
    "PDV": 982,
    "Shopee": 989,
    "Amazon": 991,
    "TikTok": 990
}

# 1. CARREGAR MAPEAMENTOS
print("\n📥 Carregando mapeamentos do Tiny...")

mapping_dict = {}  # ID Tiny → dados
id_to_titulo_ml = {}  # ID Tiny → Título ML

mapping_files = [
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-48.xls',
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-50.xls',
]

for mapping_file in mapping_files:
    if not Path(mapping_file).exists():
        continue

    try:
        df = pd.read_excel(mapping_file, sheet_name=0)

        for idx, row in df.iterrows():
            id_tiny = int(row.get('Id'))
            titulo = row.get('Título')
            sku = row.get('Produto (SKU)')

            mapping_dict[id_tiny] = {
                'titulo_ml': titulo,
                'sku': sku
            }

        print(f"✅ Carregado: {Path(mapping_file).name} ({len(df)} itens)")

    except Exception as e:
        print(f"⚠️  Erro: {e}")

print(f"   Total: {len(mapping_dict)} produtos mapeados")

# 2. CARREGAR PLANILHA ML BASE
print("\n📂 Carregando planilha base do ML...")

ml_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
wb_ml = openpyxl.load_workbook(ml_file, data_only=True)
ws_ml = wb_ml.active

header = {}
for col_idx in range(1, ws_ml.max_column + 1):
    val = ws_ml.cell(1, col_idx).value
    if val:
        header[str(val).strip()] = col_idx

title_col = header.get("TITLE", 5)
price_col = header.get("PRICE", 9)

print(f"✅ Planilha: {ml_file.name} ({ws_ml.max_row} linhas)")

# 3. BUSCAR PREÇOS DAS TABELAS (apenas para IDs que temos)
print("\n🔄 Buscando preços das tabelas para {0} produtos...".format(len(mapping_dict)))

precos_por_id = {}  # ID Tiny → {PDV, Shopee, Amazon, TikTok}

for idx, (id_tiny, dados) in enumerate(mapping_dict.items(), 1):
    precos = {}

    for nome_tabela, id_tabela in TABELAS.items():
        try:
            url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}?produto_id={id_tiny}"
            resp = requests.get(url, headers=headers, timeout=10)

            if resp.status_code == 200:
                data = resp.json()
                excecoes = data.get("excecoes", [])

                for exc in excecoes:
                    if exc.get("idProduto") == id_tiny:
                        precos[nome_tabela] = exc.get("preco")
                        break

        except:
            pass

        time.sleep(0.15)  # Rate limit

    precos_por_id[id_tiny] = precos

    if idx % 50 == 0:
        print(f"  [{idx}/{len(mapping_dict)}] Processado ID {id_tiny}")

print(f"✅ Preços coletados para {len(precos_por_id)} produtos")

# 4. CRIAR RELATÓRIO FINAL
print("\n📝 Gerando relatório final...")

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Preços Comparativos"

# CABEÇALHO
headers_out = [
    "Linha ML",
    "Título ML",
    "ID Tiny",
    "Preço ML",
    "Preço PDV",
    "Preço Shopee",
    "Preço Amazon",
    "Preço TikTok",
    "Status"
]

for col_idx, header_text in enumerate(headers_out, 1):
    cell = ws_out.cell(1, col_idx, header_text)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# DADOS
out_row = 2
ml_row_num = 6

for in_row in range(6, min(6 + 89, ws_ml.max_row + 1)):
    title_ml = ws_ml.cell(in_row, title_col).value
    price_ml = ws_ml.cell(in_row, price_col).value

    if not title_ml or not price_ml:
        continue

    title_ml = str(title_ml).strip()
    price_ml = float(price_ml)

    if price_ml <= 0:
        continue

    # Procurar este produto nos mapeamentos (por título aproximado)
    id_tiny = None
    best_match = 0

    for tid, dados in mapping_dict.items():
        titulo_map = dados['titulo_ml'].lower() if dados['titulo_ml'] else ""
        titulo_ml_lower = title_ml.lower()

        # Match simples
        if titulo_ml_lower in titulo_map or titulo_map in titulo_ml_lower:
            if len(titulo_map) > best_match:
                best_match = len(titulo_map)
                id_tiny = tid

    # Preencher linha
    precos = precos_por_id.get(id_tiny, {}) if id_tiny else {}
    status = "✅ OK" if id_tiny else "❌ Não encontrado"
    color = "E2EFDA" if id_tiny else "FCE4D6"

    ws_out.cell(out_row, 1, in_row)
    ws_out.cell(out_row, 2, title_ml[:50])
    ws_out.cell(out_row, 3, id_tiny)
    ws_out.cell(out_row, 4, price_ml)
    ws_out.cell(out_row, 5, precos.get("PDV"))
    ws_out.cell(out_row, 6, precos.get("Shopee"))
    ws_out.cell(out_row, 7, precos.get("Amazon"))
    ws_out.cell(out_row, 8, precos.get("TikTok"))
    ws_out.cell(out_row, 9, status)

    # Formatar
    for col_idx in range(1, 10):
        cell = ws_out.cell(out_row, col_idx)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in [4, 5, 6, 7, 8]:
            cell.number_format = 'R$ #,##0.00'

    out_row += 1

# Ajustar colunas
widths = [10, 40, 12, 14, 14, 14, 14, 14, 15]
for col_idx, width in enumerate(widths, 1):
    ws_out.column_dimensions[get_column_letter(col_idx)].width = width

ws_out.freeze_panes = "A2"

# SALVAR
output_file = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_FINAL.xlsx")
print(f"\n💾 Salvando: {output_file.name}")
wb_out.save(output_file)

wb_ml.close()

print(f"\n{'='*100}")
print(f"✅ RELATÓRIO GERADO COM SUCESSO!")
print(f"{'='*100}")
print(f"\n📁 Arquivo: {output_file}")
print(f"✅ Pronto para download!")
