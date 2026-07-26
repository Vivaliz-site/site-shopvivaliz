#!/usr/bin/env python3
"""
Reprocessar apenas os 3 IDs que falharam:
- 337703360: Erro 404
- 337705059: Subiu pra 399.01, esperado 374.00
- 337749479: Subiu pra 202.00, esperado 186.00

O problema: atualizar com base no preco ESPERADO (ML + 0.01) pode estar errado
Se há múltiplas linhas com o mesmo ID, pega-se o MENOR preco esperado
"""
import os
import time
import requests
import openpyxl
from pathlib import Path
from collections import defaultdict
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("REPROCESSAR IDS FALHADOS")
print("="*100)

# 1. LER RELATORIO
print("\n[1] Lendo relatorio...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

MAPA_COLUNAS = {
    7: ("Cadastro", 1.00),
    8: ("PDV", 1.45),
    9: ("Shopee", 1.10),
    10: ("Amazon", 1.00),
    11: ("TikTok", 1.10),
}

# IDs que falharam
IDS_FALHADOS = [337703360, 337705059, 337749479]

atualizacoes = {}  # {id_tiny: preco_novo_cadastro}

for row_idx in range(2, ws.max_row + 1):
    id_tiny = ws.cell(row_idx, 3).value
    preco_ml = ws.cell(row_idx, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    if id_tiny not in IDS_FALHADOS:
        continue

    novo_preco = round(preco_ml + 0.01, 2)

    # Verificar cada coluna
    tem_vermelho = False

    for col_idx in [7, 8, 9, 10, 11]:
        cell = ws.cell(row_idx, col_idx)

        if cell.fill and cell.fill.start_color:
            color_str = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color_str:
                local_nome, acrescimo = MAPA_COLUNAS[col_idx]
                preco_cadastro_necessario = round(novo_preco / acrescimo, 2)

                if id_tiny not in atualizacoes or preco_cadastro_necessario < atualizacoes[id_tiny]:
                    # Pegar o MENOR preco_cadastro (pra nao ficar caro demais)
                    atualizacoes[id_tiny] = preco_cadastro_necessario

                tem_vermelho = True
                break

wb.close()

print(f"    Identificados pra reprocessar: {len(atualizacoes)}")

for id_tiny, preco in atualizacoes.items():
    print(f"      ID {id_tiny}: R$ {preco:.2f}")

# 2. CARREGAR DADOS DOS PRODUTOS
print("\n[2] Carregando dados...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

produtos_map = {}

for prod in resp.json().get("itens", []):
    id_prod = prod.get("id")
    produtos_map[id_prod] = {
        'sku': prod.get("sku"),
        'descricao': prod.get("descricao"),
        'preco_atual': prod.get("precos", {}).get("preco")
    }

print(f"    Carregados {len(produtos_map)} produtos")

# 3. ATUALIZAR
print("\n[3] Atualizando...")

for id_tiny, preco_novo in atualizacoes.items():

    if id_tiny not in produtos_map:
        print(f"    ID {id_tiny}: nao carregado (404)")
        continue

    prod_data = produtos_map[id_tiny]

    try:
        payload = {
            "descricao": prod_data['descricao'],
            "sku": prod_data['sku'],
            "precos": {"preco": preco_novo}
        }

        url = f"https://api.tiny.com.br/public-api/v3/produtos/{id_tiny}"
        resp = requests.put(url, headers=headers, json=payload, timeout=20)

        if resp.status_code in [200, 201, 204]:
            print(f"    ID {id_tiny}: atualizado para R$ {preco_novo:.2f}")
        else:
            print(f"    ID {id_tiny}: ERRO {resp.status_code}")
            print(f"      Response: {resp.text[:100]}")

        time.sleep(0.5)

    except Exception as e:
        print(f"    ID {id_tiny}: EXCECAO {str(e)[:50]}")

# 4. VALIDAR
print(f"\n[4] Validando...")

time.sleep(1)

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

produtos_novo = {}

for prod in resp.json().get("itens", []):
    id_prod = prod.get("id")
    produtos_novo[id_prod] = prod.get("precos", {}).get("preco")

print("\n    Validacao final:")

for id_tiny, preco_esperado in atualizacoes.items():
    preco_atual = produtos_novo.get(id_tiny)

    if preco_atual is None:
        print(f"    ID {id_tiny}: ERRO 404")
    elif abs(float(preco_atual) - preco_esperado) < 0.01:
        print(f"    ID {id_tiny}: OK (R$ {preco_atual})")
    else:
        print(f"    ID {id_tiny}: FALHA (esperado R$ {preco_esperado:.2f}, obteve R$ {preco_atual})")
