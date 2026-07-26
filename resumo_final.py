import openpyxl

file_path = r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COMPLETO_ATUALIZADO.xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

print("\n" + "="*100)
print("RELATORIO FINAL: 6 PRECOS - RESUMO EXECUTIVO")
print("="*100)

print(f"\n[ARQUIVO]")
print(f"  RELATORIO_6_PRECOS_COMPLETO_ATUALIZADO.xlsx")

print(f"\n[TOTAL]")
print(f"  Produtos: {ws.max_row - 1}")

print(f"\n[PRECOS PREENCHIDOS]")
print("-"*100)

cols = {
    6: "Preco ML",
    7: "Preco Cadastro",
    8: "Preco PDV",
    9: "Preco Shopee",
    10: "Preco Amazon",
    11: "Preco TikTok",
}

total_precos = 0
for col_idx, col_name in cols.items():
    filled = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, col_idx).value)
    total = ws.max_row - 1
    pct = (filled * 100) // total if total > 0 else 0

    print(f"  {col_name:20s}: {filled:3d}/{total} ({pct:3d}%)")

    total_precos += filled

print(f"\n  TOTAL DE PRECOS PREENCHIDOS: {total_precos}")

print(f"\n[EXEMPLOS COM DADOS COMPLETOS]")
print("-"*100)

count = 0
for row_idx in range(2, ws.max_row + 1):
    pdv = ws.cell(row_idx, 8).value
    shopee = ws.cell(row_idx, 9).value
    amazon = ws.cell(row_idx, 10).value
    tiktok = ws.cell(row_idx, 11).value

    total_tabelas = sum(1 for x in [pdv, shopee, amazon, tiktok] if x)

    if total_tabelas >= 3:
        count += 1
        titulo = ws.cell(row_idx, 2).value
        preco_ml = ws.cell(row_idx, 6).value
        preco_cad = ws.cell(row_idx, 7).value

        print(f"\n{count}. {str(titulo)[:50]}")
        print(f"   ML: R$ {preco_ml} | Cadastro: R$ {preco_cad}")
        print(f"   PDV: {pdv} | Shopee: {shopee} | Amazon: {amazon} | TikTok: {tiktok}")

        if count >= 5:
            break

print("\n" + "="*100)
print("RELATORIO PRONTO PARA DOWNLOAD")
print("="*100)

wb.close()
