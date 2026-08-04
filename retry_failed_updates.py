#!/usr/bin/env python3
"""
RETRY: Tentar atualizar novamente os 28 produtos que falharam
"""
import os
import json
import time
from pathlib import Path
from datetime import datetime
from difflib import SequenceMatcher

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill
from dotenv import load_dotenv

# CARREGAR ENV
for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

print("\n╔════════════════════════════════════════════════════════════════╗")
print("║  RETRY: Atualizar produtos que falharam                      ║")
print("║  " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "                              ║")
print("╚════════════════════════════════════════════════════════════════╝")

# GET TOKEN
token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

# CARREGAR PRODUTOS
print("\n📥 Carregando produtos do Tiny...")

url = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url, headers=headers, timeout=30)

if resp.status_code != 200:
    print(f"❌ Erro: {resp.status_code}")
    exit(1)

products = resp.json().get("itens", [])
print(f"✅ Carregados: {len(products)} produtos")

# ABRIR ARQUIVO ANTERIOR
print("\n📂 Lendo relatório anterior...")

input_file = Path(r"C:\Users\user\Downloads\Precos_Atualizados_Tiny.xlsx")
wb = openpyxl.load_workbook(input_file, data_only=True)
ws = wb.active

# ENCONTRAR FALHAS
print("\n🔍 Procurando por status de FALHA (❌)...")

failed_items = []

for row in range(2, ws.max_row + 1):
    status = ws.cell(row, 9).value or ""

    if "❌" in str(status) or "HTTP 404" in str(status) or "HTTP 429" in str(status):
        linha = ws.cell(row, 1).value
        titulo = ws.cell(row, 2).value
        price_ml = ws.cell(row, 3).value
        price_tiny = ws.cell(row, 6).value
        novo_preco = ws.cell(row, 8).value

        failed_items.append({
            "linha": linha,
            "titulo": titulo,
            "price_ml": price_ml,
            "price_tiny": price_tiny,
            "novo_preco": novo_preco,
            "status_anterior": status
        })

print(f"✅ Encontradas {len(failed_items)} falhas")

if not failed_items:
    print("✅ Nenhuma falha para repetir!")
    exit(0)

# CRIAR SAÍDA COM RETRY
wb_out = openpyxl.Workbook()
ws_out = wb_out.active

cols = ["Linha", "Título", "Preço ML", "Preço Tiny", "Novo Preço", "Status Anterior", "Status RETRY", "ID Produto"]
for i, col in enumerate(cols, 1):
    ws_out.cell(1, i, col)

for col in range(1, len(cols) + 1):
    cell = ws_out.cell(1, col)
    cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")

# PROCESSAR RETRY
print(f"\n🔄 Tentando novamente...")

retry_stats = {"total": 0, "success": 0, "failed": 0}
out_row = 2

for item in failed_items:
    retry_stats["total"] += 1

    titulo = item["titulo"]
    titulo_lower = titulo.lower() if titulo else ""

    # Encontrar produto
    best_prod = None
    best_score = 0.4

    for prod in products:
        desc = prod.get("descricao", "").lower()
        score = SequenceMatcher(None, titulo_lower, desc).ratio()

        if score > best_score:
            best_prod = prod
            best_score = score

    if not best_prod:
        # Ainda não achou
        ws_out.cell(out_row, 1, item["linha"])
        ws_out.cell(out_row, 2, item["titulo"][:40])
        ws_out.cell(out_row, 3, item["price_ml"])
        ws_out.cell(out_row, 4, item["price_tiny"])
        ws_out.cell(out_row, 5, item["novo_preco"])
        ws_out.cell(out_row, 6, item["status_anterior"])
        ws_out.cell(out_row, 7, "❌ FALHA: Produto não encontrado")
        ws_out.cell(out_row, 8, "")

        retry_stats["failed"] += 1
        print(f"  ❌ [{item['linha']}] Não encontrado")
        out_row += 1
        continue

    # Tentar atualizar
    pid = best_prod.get("id")
    desc = best_prod.get("descricao", "")
    sku = best_prod.get("sku", "")

    update_url = f"https://api.tiny.com.br/public-api/v3/produtos/{pid}"
    update_payload = {
        "descricao": desc,
        "sku": sku,
        "precos": {
            "preco": float(item["novo_preco"])
        }
    }

    try:
        update_resp = requests.put(
            update_url,
            headers=headers,
            json=update_payload,
            timeout=20
        )

        if update_resp.status_code in [200, 204]:
            retry_stats["success"] += 1
            status_retry = f"✅ ATUALIZADO (Retry #{retry_stats['success']})"
            print(f"  ✅ [{item['linha']}] {item['titulo'][:30]} → R$ {item['novo_preco']}")
        else:
            retry_stats["failed"] += 1
            status_retry = f"❌ HTTP {update_resp.status_code}"
            print(f"  ❌ [{item['linha']}] Erro {update_resp.status_code}")

    except Exception as e:
        retry_stats["failed"] += 1
        status_retry = f"❌ ERRO: {str(e)[:30]}"
        print(f"  ❌ [{item['linha']}] Exceção")

    # Registrar
    ws_out.cell(out_row, 1, item["linha"])
    ws_out.cell(out_row, 2, item["titulo"][:40])
    ws_out.cell(out_row, 3, item["price_ml"])
    ws_out.cell(out_row, 4, item["price_tiny"])
    ws_out.cell(out_row, 5, item["novo_preco"])
    ws_out.cell(out_row, 6, item["status_anterior"])
    ws_out.cell(out_row, 7, status_retry)
    ws_out.cell(out_row, 8, pid if best_prod else "")

    out_row += 1
    time.sleep(0.5)  # Rate limit

# SALVAR
output = Path(r"C:\Users\user\Downloads\Retry_Precos_Atualizados.xlsx")
print(f"\n💾 Salvando: {output.name}")
wb_out.save(output)

# RELATÓRIO
print(f"\n{'='*80}")
print(f"✅ RETRY CONCLUÍDO")
print(f"{'='*80}")
print(f"\n📊 RESULTADOS DO RETRY:")
print(f"  Total de falhas iniciais: {retry_stats['total']}")
print(f"  ✅ Agora atualizados: {retry_stats['success']}")
print(f"  ❌ Ainda falharam: {retry_stats['failed']}")
print(f"\n📁 Arquivo: {output}")
print(f"✅ Pronto!")
