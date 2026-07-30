#!/usr/bin/env python3
"""
ATUALIZAR E VALIDAR:
1. Ler precos em vermelho do relatorio
2. Atualizar NO TINY para ML + 0.01
3. Extrair dados do Tiny para validar
"""
import os
import time
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("ATUALIZAR PRECOS EM VERMELHO NO TINY E VALIDAR")
print("="*100)

# 1. ABRIR RELATORIO E IDENTIFICAR PRECOS EM VERMELHO
print("\n[1] Lendo precos em vermelho do relatorio...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Cor vermelha
red_color = "FFFF0000"

atualizacoes = []  # Lista de atualizacoes a fazer

for row_idx in range(2, ws.max_row + 1):
    id_tiny = ws.cell(row_idx, 3).value
    titulo_ml = ws.cell(row_idx, 2).value
    preco_ml = ws.cell(row_idx, 6).value

    if not id_tiny or not preco_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    # Verificar se alguma celula desta linha esta vermelha
    tem_vermelho = False

    for col_idx in [7, 8, 9, 10, 11]:
        cell = ws.cell(row_idx, col_idx)

        if cell.fill and cell.fill.start_color:
            color = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color or color == red_color:
                tem_vermelho = True
                break

    if tem_vermelho:
        novo_preco = round(preco_ml + 0.01, 2)

        atualizacoes.append({
            'id_tiny': id_tiny,
            'titulo': titulo_ml,
            'preco_ml': preco_ml,
            'novo_preco': novo_preco
        })

print(f"    Encontrados {len(atualizacoes)} produtos com precos em vermelho")

if atualizacoes:
    print(f"    Exemplos:")
    for atua in atualizacoes[:3]:
        print(f"      ID {atua['id_tiny']}: {atua['titulo'][:40]}")
        print(f"        ML: R$ {atua['preco_ml']} → Novo: R$ {atua['novo_preco']}")

# 2. CARREGAR DADOS DO TINY PARA ATUALIZAR
print("\n[2] Carregando dados dos produtos para atualizacao...")

produtos_tiny = {}

for atua in atualizacoes:
    id_tiny = atua['id_tiny']

    url = f"https://api.tiny.com.br/public-api/v3/produtos/{id_tiny}"

    try:
        resp = requests.get(url, headers=headers, timeout=10)

        if resp.status_code == 200:
            prod = resp.json()
            produtos_tiny[id_tiny] = prod
        else:
            print(f"    Erro ao carregar ID {id_tiny}: {resp.status_code}")

        time.sleep(0.2)

    except Exception as e:
        print(f"    Erro ID {id_tiny}: {str(e)[:50]}")

print(f"    Carregados {len(produtos_tiny)} produtos")

# 3. ATUALIZAR PRECOS NO TINY
print("\n[3] Atualizando precos no Tiny...")

atualizados = 0
erros = 0

for atua in atualizacoes:
    id_tiny = atua['id_tiny']

    if id_tiny not in produtos_tiny:
        print(f"    Pulando {id_tiny} (nao carregado)")
        continue

    prod = produtos_tiny[id_tiny]

    # Preparar payload
    payload = {
        "descricao": prod.get("descricao"),
        "sku": prod.get("sku"),
        "precos": {
            "preco": atua['novo_preco']
        }
    }

    # PUT /produtos/{id}
    url = f"https://api.tiny.com.br/public-api/v3/produtos/{id_tiny}"

    try:
        resp = requests.put(url, headers=headers, json=payload, timeout=20)

        if resp.status_code in [200, 201]:
            atualizados += 1

            if atualizados % 10 == 0:
                print(f"    [{atualizados}] ID {id_tiny}: R$ {atua['novo_preco']} ✓")
        else:
            erros += 1
            print(f"    Erro ID {id_tiny}: {resp.status_code} - {resp.text[:100]}")

        time.sleep(0.5)

    except Exception as e:
        erros += 1
        print(f"    Erro ID {id_tiny}: {str(e)[:50]}")

# 4. VALIDAR - EXTRAIR DADOS DO TINY
print("\n[4] Validando - extraindo dados do Tiny...")

validacao = []

for atua in atualizacoes[:5]:  # Validar primeiros 5
    id_tiny = atua['id_tiny']

    url = f"https://api.tiny.com.br/public-api/v3/produtos/{id_tiny}"

    try:
        resp = requests.get(url, headers=headers, timeout=10)

        if resp.status_code == 200:
            prod = resp.json()
            preco_atual = prod.get("precos", {}).get("preco")

            validacao.append({
                'id': id_tiny,
                'titulo': atua['titulo'],
                'esperado': atua['novo_preco'],
                'atual': preco_atual,
                'ok': abs(preco_atual - atua['novo_preco']) < 0.01 if preco_atual else False
            })

            print(f"    ID {id_tiny}: esperado R$ {atua['novo_preco']}, atual R$ {preco_atual}")

        time.sleep(0.2)

    except Exception as e:
        print(f"    Erro validacao ID {id_tiny}: {str(e)[:50]}")

# RELATORIO
wb.close()

print(f"\n" + "="*100)
print("CONCLUSAO")
print("="*100)
print(f"\n[ATUALIZACOES]")
print(f"  Total identificado: {len(atualizacoes)}")
print(f"  Atualizados com sucesso: {atualizados}")
print(f"  Erros: {erros}")

print(f"\n[VALIDACAO (primeiros 5)]")
for v in validacao:
    status = "✓ OK" if v['ok'] else "✗ ERRO"
    print(f"  ID {v['id']:10d}: {status} - {v['titulo'][:40]}")

print(f"\n✅ Processo concluido!")
