#!/usr/bin/env python3
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font

print("Formatando arquivo novo...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Cor vermelha com texto branco
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
red_font = Font(color="FFFFFF", bold=True)

count = 0

for row_idx in range(2, ws.max_row + 1):
    preco_ml_cell = ws.cell(row_idx, 6)
    preco_ml = preco_ml_cell.value

    if not preco_ml:
        continue

    try:
        preco_ml = float(preco_ml)
    except:
        continue

    # Verificar colunas 7-11 (Cadastro, PDV, Shopee, Amazon, TikTok)
    for col_idx in [7, 8, 9, 10, 11]:
        preco_cell = ws.cell(row_idx, col_idx)
        preco = preco_cell.value

        if not preco:
            continue

        try:
            preco = float(preco)
        except:
            continue

        # Se MENOR que ML, pintar vermelho
        if preco < preco_ml:
            preco_cell.fill = red_fill
            preco_cell.font = red_font
            count += 1

print(f"Formatadas {count} celulas em vermelho")

wb.save(file_path)
wb.close()

print(f"Salvo: {file_path.name}")
print("Pronto!")
