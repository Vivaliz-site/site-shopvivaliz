#!/usr/bin/env python3
"""
VERSÃO FINAL: Relatório com 6 preços (sem conflito de arquivo aberto)
"""
import os
import time
import requests
import openpyxl
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n╔════════════════════════════════════════════════════════════════════════════════╗")
print("║  RELATÓRIO FINAL v2: 6 Preços (ML + Cadastro + PDV + Shopee + Amazon + TikTok)║")
print("╚════════════════════════════════════════════════════════════════════════════════╝")

# IDs das tabelas
TABELAS = {
    "PDV": 982,
    "Shopee": 989,
    "Amazon": 991,
    "TikTok": 990
}

# 1. CARREGAR MAPEAMENTOS
print("\n📥 Carregando mapeamentos do Tiny...")

mapping_data = {}  # ID Tiny → {titulo_ml, sku}

mapping_files = [
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-48.xls',
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-50.xls',
]

for mfile in mapping_files:
    if Path(mfile).exists():
        try:
            df = pd.read_excel(mfile, sheet_name=0)
            for idx, row in df.iterrows():
                id_tiny = int(row.get('Id'))
                titulo = row.get('Título')
                sku = row.get('Produto (SKU)')
                mapping_data[id_tiny] = {'titulo': titulo, 'sku': sku}
            print(f"✅ {Path(mfile).name}: {len(df)} itens")
        except Exception as e:
            print(f"⚠️  {Path(mfile).name}: {e}")

print(f"   Total: {len(mapping_data)} produtos mapeados")

# 2. CARREGAR PLANILHA BASE ML
print("\n📂 Carregando planilha base ML...")

ml_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
wb_ml = openpyxl.load_workbook(ml_file, data_only=True)
ws_ml = wb_ml.active

header = {}
for col_idx in range(1, ws_ml.max_column + 1):
    val = ws_ml.cell(1, col_idx).value
    if val:
        header[str(val).strip()] = col_idx

title_col = header.get("TITLE", 5)
price_col = header.get("PRICE", 9)

print(f"✅ {ml_file.name}: {ws_ml.max_row} linhas")

# 3. CARREGAR PRODUTOS TINY (para dados do cadastro)
print("\n📥 Carregando produtos Tiny...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

if resp.status_code != 200:
    print(f"❌ Erro: {resp.status_code}")
    exit(1)

products_tiny = {p.get('id'): p for p in resp.json().get("itens", [])}
print(f"✅ Carregados {len(products_tiny)} produtos")

# 4. CRIAR NOVO ARQUIVO
print("\n📝 Criando relatório...")

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Preços Comparativos"

# CABEÇALHO
headers_out = [
    "Linha ML",
    "Título ML",
    "ID Tiny",
    "SKU Tiny",
    "Descrição Tiny",
    "Preço ML",
    "Preço Cadastro",
    "Preço PDV",
    "Preço Shopee",
    "Preço Amazon",
    "Preço TikTok",
    "Status"
]

for col_idx, header_text in enumerate(headers_out, 1):
    cell = ws_out.cell(1, col_idx, header_text)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 5. PROCESSAR DADOS
print("\n🔄 Processando 89 produtos...")

stats = {"total": 0, "found": 0, "with_prices": 0}
out_row = 2

for in_row in range(6, min(6 + 89, ws_ml.max_row + 1)):
    title_ml = ws_ml.cell(in_row, title_col).value
    price_ml = ws_ml.cell(in_row, price_col).value

    if not title_ml or not price_ml:
        continue

    title_ml = str(title_ml).strip()
    price_ml = float(price_ml)

    if price_ml <= 0:
        continue

    stats["total"] += 1

    # Procurar produto Tiny
    id_tiny = None
    prod_tiny = None

    # Estratégia: procurar por similaridade de título nos mapeamentos
    title_lower = title_ml.lower()
    best_match = -1

    for tid, dados in mapping_data.items():
        titulo_map = (dados.get('titulo') or "").lower()
        if titulo_map and (titulo_lower in titulo_map or titulo_map in titulo_lower):
            match_len = len(titulo_map)
            if match_len > best_match:
                best_match = match_len
                id_tiny = tid

    # Se encontrou, buscar dados do Tiny
    if id_tiny and id_tiny in products_tiny:
        stats["found"] += 1
        prod_tiny = products_tiny[id_tiny]
        sku_tiny = prod_tiny.get("sku", "")
        desc_tiny = prod_tiny.get("descricao", "")
        price_cadastro = prod_tiny.get("precos", {}).get("preco")

        # Buscar preços das tabelas
        precos_tabelas = {}

        for nome_tabela, id_tabela in TABELAS.items():
            try:
                url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}?produto_id={id_tiny}"
                resp = requests.get(url, headers=headers, timeout=10)

                if resp.status_code == 200:
                    data = resp.json()
                    excecoes = data.get("excecoes", [])

                    for exc in excecoes:
                        if exc.get("idProduto") == id_tiny:
                            preco = exc.get("preco")
                            if preco:
                                precos_tabelas[nome_tabela] = preco
                                stats["with_prices"] += 1
                            break

            except:
                pass

            time.sleep(0.15)

        final_status = "✅ OK"
        color = "E2EFDA"

    else:
        sku_tiny = ""
        desc_tiny = ""
        price_cadastro = None
        precos_tabelas = {}
        final_status = "❌ Não encontrado"
        color = "FCE4D6"

    # Preencher linha
    ws_out.cell(out_row, 1, in_row)
    ws_out.cell(out_row, 2, title_ml[:50])
    ws_out.cell(out_row, 3, id_tiny)
    ws_out.cell(out_row, 4, sku_tiny)
    ws_out.cell(out_row, 5, desc_tiny[:50] if desc_tiny else "")
    ws_out.cell(out_row, 6, price_ml)
    ws_out.cell(out_row, 7, price_cadastro)
    ws_out.cell(out_row, 8, precos_tabelas.get("PDV"))
    ws_out.cell(out_row, 9, precos_tabelas.get("Shopee"))
    ws_out.cell(out_row, 10, precos_tabelas.get("Amazon"))
    ws_out.cell(out_row, 11, precos_tabelas.get("TikTok"))
    ws_out.cell(out_row, 12, final_status)

    # Formatar
    for col_idx in range(1, 13):
        cell = ws_out.cell(out_row, col_idx)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx in [6, 7, 8, 9, 10, 11]:
            cell.number_format = 'R$ #,##0.00'

    if stats["total"] % 20 == 0:
        print(f"  [{stats['total']:2d}] {title_ml[:40]}")

    out_row += 1

# Ajustar colunas
widths = [10, 35, 12, 12, 35, 14, 14, 14, 14, 14, 14, 15]
for col_idx, width in enumerate(widths, 1):
    ws_out.column_dimensions[get_column_letter(col_idx)].width = width

ws_out.freeze_panes = "A2"

# SALVAR COM NOVO NOME
output_file = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_FINAL_COMPLETO.xlsx")
print(f"\n💾 Salvando: {output_file.name}")
wb_out.save(output_file)

wb_ml.close()

# RELATÓRIO
print(f"\n{'='*100}")
print(f"✅ RELATÓRIO GERADO COM SUCESSO!")
print(f"{'='*100}")
print(f"\n📊 ESTATÍSTICAS:")
print(f"  Total de anúncios: {stats['total']}")
print(f"  ✅ Encontrados no Tiny: {stats['found']}")
print(f"  💰 Preços de tabelas encontrados: {stats['with_prices']}")
print(f"\n📁 Arquivo: {output_file}")
print(f"\n✅ Pronto para download!")
