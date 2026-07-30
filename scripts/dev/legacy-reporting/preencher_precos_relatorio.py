#!/usr/bin/env python3
"""
Preencher os preços das tabelas no RELATORIO_6_PRECOS_COMPLETO.xlsx
"""
import os
import json
import time
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv
from openpyxl.styles import PatternFill, Alignment

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n╔════════════════════════════════════════════════════════════════════════════════╗")
print("║           PREENCHENDO PREÇOS DAS TABELAS NO RELATÓRIO                         ║")
print("╚════════════════════════════════════════════════════════════════════════════════╝")

# IDs das tabelas
TABELAS = {
    "PDV": 982,
    "Shopee": 989,
    "Amazon": 991,
    "TikTok": 990
}

# Abrir arquivo
print("\n📂 Abrindo relatório...")
rel_file = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COMPLETO.xlsx")
wb = openpyxl.load_workbook(rel_file)
ws = wb.active

print(f"✅ Arquivo aberto: {rel_file.name}")
print(f"   Linhas: {ws.max_row}")

# Processar cada linha
print("\n🔄 Buscando preços das tabelas...")
print("="*100)

total = 0
sucesso = 0
erro = 0

for row_idx in range(2, ws.max_row + 1):
    # Ler dados
    id_tiny = ws.cell(row_idx, 3).value  # Coluna C: ID Tiny
    sku = ws.cell(row_idx, 4).value      # Coluna D: SKU

    if not id_tiny or str(id_tiny).strip() == "":
        continue

    total += 1

    try:
        # Converter para int
        id_tiny = int(id_tiny)

        # Buscar preços de cada tabela
        precos_tabelas = {}

        for nome_tabela, id_tabela in TABELAS.items():
            url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}?produto_id={id_tiny}"

            try:
                resp = requests.get(url, headers=headers, timeout=10)

                if resp.status_code == 200:
                    data = resp.json()

                    # Procurar o preço nas exceções
                    preco = None
                    excecoes = data.get("excecoes", [])

                    for exc in excecoes:
                        if exc.get("idProduto") == id_tiny:
                            preco = exc.get("preco")
                            break

                    precos_tabelas[nome_tabela] = preco
                else:
                    precos_tabelas[nome_tabela] = None

            except Exception as e:
                precos_tabelas[nome_tabela] = None
                print(f"  ⚠️  Erro ao buscar {nome_tabela} para produto {id_tiny}: {str(e)[:50]}")

            time.sleep(0.2)  # Rate limit

        # Preencher as colunas de preço
        ws.cell(row_idx, 8, precos_tabelas.get("PDV"))      # Coluna H: Preço PDV
        ws.cell(row_idx, 9, precos_tabelas.get("Shopee"))   # Coluna I: Preço Shopee
        ws.cell(row_idx, 10, precos_tabelas.get("Amazon"))  # Coluna J: Preço Amazon
        ws.cell(row_idx, 11, precos_tabelas.get("TikTok"))  # Coluna K: Preço TikTok

        # Formatar como moeda
        for col in [8, 9, 10, 11]:
            ws.cell(row_idx, col).number_format = 'R$ #,##0.00'

        sucesso += 1

        if total % 10 == 0:
            print(f"  ✅ [{total:2d}] Produto {id_tiny} - PDV: {precos_tabelas.get('PDV')}, Shopee: {precos_tabelas.get('Shopee')}")

    except Exception as e:
        erro += 1
        print(f"  ❌ Erro na linha {row_idx}: {str(e)[:80]}")

# Salvar arquivo atualizado
print(f"\n💾 Salvando arquivo atualizado...")
wb.save(rel_file)
wb.close()

# Relatório
print(f"\n{'='*100}")
print(f"✅ PREENCHIMENTO CONCLUÍDO!")
print(f"{'='*100}")
print(f"\n📊 ESTATÍSTICAS:")
print(f"  Total processado: {total}")
print(f"  ✅ Sucesso: {sucesso}")
print(f"  ❌ Erro: {erro}")
print(f"\n📁 Arquivo atualizado: {rel_file}")
print(f"\n✅ Pronto!")
