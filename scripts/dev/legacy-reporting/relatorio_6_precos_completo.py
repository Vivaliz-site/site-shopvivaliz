#!/usr/bin/env python3
"""
RELATÓRIO FINAL: 6 Preços (ML + Cadastro + PDV + Shopee + Amazon + TikTok)
Com ID do Tiny mapeado
"""
import os
import json
import time
from pathlib import Path
from datetime import datetime
from difflib import SequenceMatcher

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# CARREGAR ENV
for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n╔════════════════════════════════════════════════════════════════════════════════╗")
print("║  RELATÓRIO FINAL: 6 PREÇOS (ML + CADASTRO + PDV + SHOPEE + AMAZON + TIKTOK)   ║")
print("║  " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "                                     ║")
print("╚════════════════════════════════════════════════════════════════════════════════╝")

# IDs das tabelas de preço
TABELAS = {
    "PDV": 982,
    "Shopee": 989,
    "Amazon": 991,
    "TikTok": 990
}

print("\n✅ Tabelas de preço identificadas:")
for nome, id_tab in TABELAS.items():
    print(f"   {nome:15s}: ID {id_tab}")

# 1. CARREGAR PRODUTOS DO TINY
print("\n📥 Carregando produtos do Tiny...")

url_produtos = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url_produtos, headers=headers, timeout=30)

if resp.status_code != 200:
    print(f"❌ Erro ao carregar produtos: {resp.status_code}")
    exit(1)

products_tiny = resp.json().get("itens", [])
print(f"✅ Carregados {len(products_tiny)} produtos do Tiny")

# 2. ABRIR PLANILHA ML
print("\n📂 Abrindo planilha do Mercado Livre...")

input_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
wb_in = openpyxl.load_workbook(input_file, data_only=True)
ws_in = wb_in.active

# Encontrar colunas
header = []
for col_idx in range(1, ws_in.max_column + 1):
    header.append(ws_in.cell(1, col_idx).value)

title_col = price_col = None
for idx, col_name in enumerate(header, 1):
    if col_name == "TITLE":
        title_col = idx
    elif col_name == "PRICE":
        price_col = idx

print(f"✅ Colunas: Título={title_col}, Preço ML={price_col}")

# 3. CRIAR ARQUIVO DE SAÍDA
print("\n📝 Criando relatório...")

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Preços Comparativos"

# CABEÇALHO
headers_out = [
    "Linha ML",
    "Título ML",
    "ID Tiny",
    "SKU Tiny",
    "Descrição Tiny",
    "Preço ML",
    "Preço Cadastro",
    "Preço PDV",
    "Preço Shopee",
    "Preço Amazon",
    "Preço TikTok",
    "Status"
]

for col_idx, header_text in enumerate(headers_out, 1):
    cell = ws_out.cell(1, col_idx, header_text)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 4. PROCESSAR DADOS
print("\n🔄 Processando anúncios...")

stats = {"total": 0, "found": 0, "errors": 0}
out_row = 2

for in_row in range(6, min(6 + 89, ws_in.max_row + 1)):
    title_ml = ws_in.cell(in_row, title_col).value
    price_ml = ws_in.cell(in_row, price_col).value

    if not title_ml or not price_ml:
        continue

    title_ml = str(title_ml).strip()
    price_ml = float(price_ml)

    if price_ml <= 0:
        continue

    stats["total"] += 1

    # Encontrar produto Tiny por similaridade
    title_lower = title_ml.lower()
    best_prod = None
    best_score = 0.4

    for prod in products_tiny:
        desc = prod.get("descricao", "").lower()
        score = SequenceMatcher(None, title_lower, desc).ratio()

        if score > best_score:
            best_prod = prod
            best_score = score

    if not best_prod:
        # Não encontrou
        ws_out.cell(out_row, 1, in_row)
        ws_out.cell(out_row, 2, title_ml[:40])
        ws_out.cell(out_row, 6, price_ml)
        ws_out.cell(out_row, 12, "❌ Não encontrado")
        stats["errors"] += 1
        out_row += 1
        continue

    stats["found"] += 1

    pid = best_prod.get("id")
    sku = best_prod.get("sku", "")
    desc = best_prod.get("descricao", "")
    price_cadastro = best_prod.get("precos", {}).get("preco")

    # Preços das tabelas
    precos_tabelas = {}

    for nome_tabela, id_tabela in TABELAS.items():
        # GET /produtos/{id}/listas-precos/{tabela_id}
        url_tabela = f"https://api.tiny.com.br/public-api/v3/produtos/{pid}/lista-preco/{id_tabela}"

        try:
            resp_tab = requests.get(url_tabela, headers=headers, timeout=10)
            if resp_tab.status_code == 200:
                data_tab = resp_tab.json()
                # Pode vir como "preco" ou dentro de um objeto
                if isinstance(data_tab, dict):
                    preco_tab = data_tab.get("preco") or data_tab.get("preco_tabela")
                else:
                    preco_tab = None
                precos_tabelas[nome_tabela] = preco_tab
            else:
                precos_tabelas[nome_tabela] = None
        except:
            precos_tabelas[nome_tabela] = None

        time.sleep(0.1)  # Rate limit

    # Preencher linha
    ws_out.cell(out_row, 1, in_row)
    ws_out.cell(out_row, 2, title_ml[:40])
    ws_out.cell(out_row, 3, pid)
    ws_out.cell(out_row, 4, sku)
    ws_out.cell(out_row, 5, desc[:40])
    ws_out.cell(out_row, 6, price_ml)
    ws_out.cell(out_row, 7, price_cadastro)
    ws_out.cell(out_row, 8, precos_tabelas.get("PDV"))
    ws_out.cell(out_row, 9, precos_tabelas.get("Shopee"))
    ws_out.cell(out_row, 10, precos_tabelas.get("Amazon"))
    ws_out.cell(out_row, 11, precos_tabelas.get("TikTok"))
    ws_out.cell(out_row, 12, "✅ OK")

    if stats["total"] % 20 == 0:
        print(f"  [{in_row}] {title_ml[:35]}")

    out_row += 1

# Ajustar largura das colunas
for col in range(1, len(headers_out) + 1):
    ws_out.column_dimensions[get_column_letter(col)].width = 15

# SALVAR
output_file = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COMPLETO.xlsx")
print(f"\n💾 Salvando: {output_file.name}")
wb_out.save(output_file)

wb_in.close()

# RELATÓRIO
print(f"\n{'='*80}")
print(f"✅ RELATÓRIO GERADO COM SUCESSO")
print(f"{'='*80}")
print(f"\n📊 ESTATÍSTICAS:")
print(f"  Total processado: {stats['total']}")
print(f"  ✅ Encontrados: {stats['found']}")
print(f"  ❌ Erros: {stats['errors']}")
print(f"\n📁 Arquivo: {output_file}")
print(f"\n✅ Pronto!")
