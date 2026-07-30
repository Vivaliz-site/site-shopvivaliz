#!/usr/bin/env python3
import openpyxl
from pathlib import Path

file_path = Path(r"C:\Users\user\Downloads\exemplo.xlsx")
wb = openpyxl.load_workbook(file_path)

print("\n" + "="*100)
print("ANALISE DO ARQUIVO EXEMPLO")
print("="*100)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    print(f"\n[ABA: {sheet_name}]")
    print(f"Colunas: {ws.max_column}, Linhas: {ws.max_row}")
    print("\n[HEADERS]")
    print("="*100)

    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value

        if header:
            print(f"  Col {col:2d}: {header}")

    print("\n[DADOS]")
    print("="*100)

    for row in range(2, ws.max_row + 1):
        print(f"\nLinha {row}:")

        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            value = ws.cell(row, col).value

            if value is not None:
                valor_str = str(value)[:60]
                header_str = str(header) if header else f"Col{col}"

                print(f"  {header_str:30s}: {valor_str}")

print("\n" + "="*100)
