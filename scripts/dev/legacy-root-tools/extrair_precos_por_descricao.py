#!/usr/bin/env python3
"""
EXTRAIR PREÇOS: Mapear pela DESCRIÇÃO dos produtos
Buscar os 89 produtos nas exceções das tabelas por match de descrição
"""
import os
import time
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv
from difflib import SequenceMatcher

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n╔════════════════════════════════════════════════════════════════════════════════╗")
print("║  EXTRAIR PREÇOS: Mapear pela DESCRIÇÃO do Produto                             ║")
print("╚════════════════════════════════════════════════════════════════════════════════╝")

# IDs das tabelas
TABELAS = {"PDV": 982, "Shopee": 989, "Amazon": 991, "TikTok": 990}

# 1. CARREGAR TODAS AS EXCECOES DE TODAS AS TABELAS
print("\n📥 Carregando todas as exceções das 4 tabelas...")

todas_excecoes = {}  # descricao → {PDV, Shopee, Amazon, TikTok}

for nome_tabela, id_tabela in TABELAS.items():
    print(f"\n  {nome_tabela} (ID {id_tabela})...")

    url = f"https://api.tiny.com.br/public-api/v3/listas-precos/{id_tabela}"

    try:
        resp = requests.get(url, headers=headers, timeout=20)

        if resp.status_code == 200:
            data = resp.json()
            excecoes = data.get("excecoes", [])

            print(f"    Carregadas {len(excecoes)} exceções")

            for exc in excecoes:
                id_prod = exc.get("idProduto")
                codigo = exc.get("codigo", "")
                preco = exc.get("preco")

                # Usar código (SKU) como chave principal
                chave = codigo.strip().upper() if codigo else f"ID_{id_prod}"

                if chave not in todas_excecoes:
                    todas_excecoes[chave] = {}

                todas_excecoes[chave][nome_tabela] = preco
        else:
            print(f"    Erro: {resp.status_code}")
    except Exception as e:
        print(f"    Erro: {e}")

    time.sleep(0.5)

print(f"\n✅ Total de produtos com exceções: {len(todas_excecoes)}")
print(f"   Exemplos: {list(todas_excecoes.keys())[:5]}")

# 2. ABRIR RELATÓRIO
print("\n📂 Abrindo relatório...")
file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COMPLETO_ATUALIZADO.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"✅ {file_path.name} ({ws.max_row - 1} linhas)")

# 3. MAPEAR PREÇOS POR SKU/DESCRIÇÃO
print("\n🔄 Mapeando preços por SKU/Descrição...")

stats = {"processado": 0, "encontrado": 0, "com_preco": 0}

for row_idx in range(2, ws.max_row + 1):
    sku = ws.cell(row_idx, 4).value  # Coluna D: SKU Tiny
    desc = ws.cell(row_idx, 5).value  # Coluna E: Descrição Tiny

    if not sku:
        continue

    stats["processado"] += 1

    sku_upper = str(sku).strip().upper()

    # Tentar encontrar nos mapeamentos
    precos_encontrados = {}

    # Estratégia 1: Match exato de SKU
    if sku_upper in todas_excecoes:
        precos_encontrados = todas_excecoes[sku_upper]
        stats["encontrado"] += 1
    else:
        # Estratégia 2: Buscar por similaridade de descrição
        if desc:
            desc_lower = str(desc).lower()
            best_match = -1
            best_key = None

            for chave, precos in todas_excecoes.items():
                if "_" in chave:  # Pular os ID_xxx
                    continue

                score = SequenceMatcher(None, desc_lower, chave.lower()).ratio()

                if score > best_match and score > 0.6:
                    best_match = score
                    best_key = chave

            if best_key:
                precos_encontrados = todas_excecoes[best_key]
                stats["encontrado"] += 1

    # Preencher preços
    if precos_encontrados:
        for col_idx, (nome_tabela, _) in enumerate(TABELAS.items(), 8):
            preco = precos_encontrados.get(nome_tabela)
            if preco:
                ws.cell(row_idx, col_idx, preco)
                ws.cell(row_idx, col_idx).number_format = 'R$ #,##0.00'
                stats["com_preco"] += 1

    if stats["processado"] % 20 == 0:
        print(f"  [{stats['processado']:2d}] SKU {sku_upper}: {len(precos_encontrados)} preços")

# Salvar
print(f"\n💾 Salvando...")
wb.save(file_path)
wb.close()

print(f"\n{'='*100}")
print(f"✅ MAPEAMENTO POR DESCRIÇÃO CONCLUÍDO!")
print(f"{'='*100}")
print(f"\n📊 ESTATÍSTICAS:")
print(f"  Total processado: {stats['processado']}")
print(f"  Encontrados nas tabelas: {stats['encontrado']}")
print(f"  Preços adicionados: {stats['com_preco']}")
print(f"\n📁 {file_path}")
print(f"✅ Pronto para download!")
