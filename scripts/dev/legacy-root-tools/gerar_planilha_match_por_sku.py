#!/usr/bin/env python3
"""
Usar SKU como chave de match (mais confiavel que titulo)
"""
import os
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv
import requests

try:
    import xlrd
except:
    os.system("pip install xlrd -q")
    import xlrd

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("GERAR PLANILHA - MATCH POR SKU")
print("="*100)

# 1. EXTRAIR ANUNCIOS E INDEXAR POR SKU + INTEGRACAO
print("\n[1] Extraindo e indexando anuncios por SKU...")

downloads_path = Path(r"C:\Users\user\Downloads")
anuncios_files = sorted(downloads_path.glob("anuncios_2026-07-26-*.xls"))

# Estrutura: {(sku, integracao_lower): anuncio}
anuncios_por_sku = {}

for file_path in anuncios_files:
    try:
        workbook = xlrd.open_workbook(str(file_path))
        worksheet = workbook.sheet_by_index(0)

        for row in range(1, worksheet.nrows):
            id_anuncio = worksheet.cell_value(row, 0)
            integracao = worksheet.cell_value(row, 1)
            identificador = worksheet.cell_value(row, 2)
            titulo = worksheet.cell_value(row, 3)
            sku = worksheet.cell_value(row, 4)

            if sku and integracao:
                sku_str = str(sku).strip().upper()
                integ_str = str(integracao).strip().lower()

                key = (sku_str, integ_str)

                anuncios_por_sku[key] = {
                    'id_anuncio': id_anuncio,
                    'integracao': integracao,
                    'identificador': identificador,
                    'titulo': str(titulo) if titulo else '',
                    'sku': sku
                }

    except Exception as e:
        if "Workbook corruption" not in str(e):
            pass

print(f"    Indexados {len(anuncios_por_sku)} anuncios por SKU + Integracao")

# 2. LER RELATORIO
print("\n[2] Lendo relatorio...")

file_path = Path(r"C:\Users\user\Downloads\RELATORIO_6_PRECOS_COM_ALERTA.xlsx")
wb_orig = openpyxl.load_workbook(file_path)
ws_orig = wb_orig.active

# 3. CARREGAR DADOS DO TINY
print("\n[3] Carregando dados do Tiny...")

produtos_tiny = {}
offset = 0

while True:
    url_produtos = f"https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset={offset}"
    resp = requests.get(url_produtos, headers=headers, timeout=30)

    itens = resp.json().get("itens", [])

    if not itens:
        break

    for prod in itens:
        id_prod = prod.get("id")
        produtos_tiny[id_prod] = {
            'sku': prod.get("sku"),
            'descricao': prod.get("descricao"),
            'preco': prod.get("precos", {}).get("preco")
        }

    offset += 100

print(f"    Carregados {len(produtos_tiny)} produtos")

# 4. CRIAR PLANILHA
print("\n[4] Criando planilha...")

wb_new = openpyxl.Workbook()
ws_new = wb_new.active

headers_corretos = [
    "Id",
    "Integração",
    "Identificador",
    "Título",
    "Produto (SKU)",
    "Preço de custo",
    "Preço",
    "Preço promocional"
]

# Escrever headers
style_header = Font(bold=True, color="FFFFFF", size=11)
fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

for col_idx, header in enumerate(headers_corretos, 1):
    cell = ws_new.cell(1, col_idx)
    cell.value = header
    cell.font = style_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Preencher dados
linha_saida = 2
row_count = 0

MAPA_COL = {
    8: "pdv",
    9: "shopee",
    10: "amazon",
    11: "tiktok"
}

# Processar relatorio
for row_orig in range(2, ws_orig.max_row + 1):
    titulo_ml = ws_orig.cell(row_orig, 2).value
    id_tiny = ws_orig.cell(row_orig, 3).value
    preco_ml = ws_orig.cell(row_orig, 6).value

    if not id_tiny or not preco_ml or not titulo_ml:
        continue

    try:
        id_tiny = int(id_tiny)
        preco_ml = float(preco_ml)
    except:
        continue

    # Verificar se tem vermelho
    colunas_com_vermelho = []

    for col_idx in [8, 9, 10, 11]:
        cell = ws_orig.cell(row_orig, col_idx)

        if cell.fill and cell.fill.start_color:
            color_str = str(cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else '')

            if 'FF0000' in color_str:
                colunas_com_vermelho.append(col_idx)

    if not colunas_com_vermelho:
        continue

    # Tem vermelho!
    preco_esperado = round(preco_ml + 0.01, 2)
    preco_cadastro = ws_orig.cell(row_orig, 7).value
    sku_tiny = produtos_tiny.get(id_tiny, {}).get('sku')

    if not sku_tiny:
        continue

    sku_tiny_upper = str(sku_tiny).strip().upper()

    # Para cada coluna com vermelho
    for col_idx in colunas_com_vermelho:
        integracao_lower = MAPA_COL[col_idx]

        # Procurar anuncio por SKU + INTEGRACAO
        key = (sku_tiny_upper, integracao_lower)

        if key in anuncios_por_sku:
            anuncio = anuncios_por_sku[key]
            row_count += 1

            # PREENCHER COM DADOS DO ANUNCIO
            ws_new.cell(linha_saida, 1).value = row_count
            ws_new.cell(linha_saida, 2).value = anuncio['integracao']
            ws_new.cell(linha_saida, 3).value = anuncio['identificador']
            ws_new.cell(linha_saida, 4).value = anuncio['titulo']
            ws_new.cell(linha_saida, 5).value = anuncio['sku']
            ws_new.cell(linha_saida, 6).value = None

            ws_new.cell(linha_saida, 7).value = preco_cadastro
            ws_new.cell(linha_saida, 7).number_format = 'R$ #,##0.00'

            # COLUNA AMARELA: Preco promocional
            cell_promo = ws_new.cell(linha_saida, 8)
            cell_promo.value = preco_esperado
            cell_promo.number_format = 'R$ #,##0.00'
            cell_promo.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell_promo.font = Font(color="000000", bold=True)

            # Alinhamento
            ws_new.cell(linha_saida, 1).alignment = Alignment(horizontal="center")
            for col in range(2, 9):
                ws_new.cell(linha_saida, col).alignment = Alignment(horizontal="left", vertical="center")

            linha_saida += 1

print(f"\n✅ Planilha criada com {row_count} linhas")

# Ajustar larguras
ws_new.column_dimensions['A'].width = 8
ws_new.column_dimensions['B'].width = 20
ws_new.column_dimensions['C'].width = 25
ws_new.column_dimensions['D'].width = 45
ws_new.column_dimensions['E'].width = 18
ws_new.column_dimensions['F'].width = 16
ws_new.column_dimensions['G'].width = 16
ws_new.column_dimensions['H'].width = 18

ws_new.freeze_panes = "A2"

# Salvar
output_path = Path(r"C:\Users\user\Downloads\PLANILHA_FINAL.xlsx")
wb_new.save(output_path)

print(f"\n" + "="*100)
print("RESUMO FINAL")
print("="*100)

print(f"\n  Total de anuncios para importar: {row_count}")
print(f"  Arquivo: {output_path.name}")
print(f"\n  Estrutura:")
print(f"  - Colunas 1-7: Dados originais dos anuncios (SEM ALTERACAO)")
print(f"  - Coluna 8 (AMARELA): Preco promocional (NOVO = ML + 0.01)")

print(f"\n✅ PRONTO PARA IMPORTAR NO TINY!")
