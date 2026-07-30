#!/usr/bin/env python3
"""
Sincronização v2: ML → Tiny (Com rate limiting e cache)
Busca produtos por título, identifica preços mais baratos, atualiza automaticamente
"""
import os
import json
import sys
import time
from pathlib import Path
from typing import Dict, Optional, List
from datetime import datetime
from difflib import SequenceMatcher

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

# ═══════════════════════════════════════════════════════════════════════════════
# CARREGAMENTO
# ═══════════════════════════════════════════════════════════════════════════════

env_files = [
    Path(r"C:\site-shopvivaliz\.env.local"),
    Path(r"C:\site-shopvivaliz\.env"),
    Path("/home/ubuntu/site-shopvivaliz/.env"),
]

for f in env_files:
    if f.exists():
        load_dotenv(f)
        break

# ═══════════════════════════════════════════════════════════════════════════════
# CLIENT COM RATE LIMIT
# ═══════════════════════════════════════════════════════════════════════════════

class TinyAPIClientV2:
    def __init__(self):
        self.access_token = os.getenv("OLIST_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")
        if not self.access_token:
            print("❌ ERRO: Token não configurado")
            sys.exit(1)

        self.base_url = "https://api.tiny.com.br/public-api/v3"
        self.session = requests.Session()
        retry = Retry(total=3, backoff_factor=2, status_forcelist=[429, 500, 502, 503, 504])
        adapter = HTTPAdapter(max_retries=retry)
        self.session.mount("https://", adapter)

        self.cache_file = Path(r"C:\site-shopvivaliz\tiny_products_cache.json")
        self.all_products = None
        self.request_count = 0
        self.last_request_time = 0
        self.min_interval = 1.5  # Mínimo de 1.5s entre requisições

        print(f"✅ Cliente API inicializado (com rate limit)")

    def _wait_rate_limit(self):
        """Respeita rate limit"""
        elapsed = time.time() - self.last_request_time
        if elapsed < self.min_interval:
            time.sleep(self.min_interval - elapsed)

    def _make_request(self, method: str, endpoint: str, params: Dict = None, json_data: Dict = None) -> Optional[Dict]:
        """Faz requisição com rate limit"""
        self._wait_rate_limit()
        self.request_count += 1
        self.last_request_time = time.time()

        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        }

        url = f"{self.base_url}{endpoint}"

        try:
            if method == "GET":
                response = self.session.get(url, headers=headers, params=params, timeout=20)
            elif method == "PUT":
                response = self.session.put(url, headers=headers, json=json_data, timeout=20)
            else:
                response = self.session.request(method, url, headers=headers, timeout=20)

            # Verificar rate limit
            remaining = response.headers.get("x-ratelimit-remaining")
            reset = response.headers.get("x-ratelimit-reset")
            if remaining:
                print(f"     (Rate Limit: {remaining} requisições restantes, reset em {reset}s)", end="\r")

            if response.status_code == 429:
                print(f"\n⚠️  Rate limit atingido! Aguardando...")
                time.sleep(int(reset or 60))
                return self._make_request(method, endpoint, params, json_data)

            if response.status_code in [200, 204]:
                return response.json() if response.text else {"status": "success"}
            else:
                return None

        except Exception as e:
            print(f"\n❌ Erro: {e}")
            return None

    def load_all_products_cached(self):
        """Carrega produtos com cache em arquivo"""
        if self.all_products is not None:
            return self.all_products

        # Tentar carregar do cache
        if self.cache_file.exists():
            print("\n🔄 Carregando produtos do cache...")
            try:
                with open(self.cache_file, "r", encoding="utf-8") as f:
                    self.all_products = json.load(f)
                print(f"✅ Carregados do cache: {len(self.all_products)} produtos")
                return self.all_products
            except:
                pass

        # Carregar da API
        print("\n🔄 Carregando produtos da API (primeira vez)...")
        all_products = []
        offset = 0
        limit = 500

        while True:
            print(f"\n  Requisição {self.request_count + 1}: offset={offset}, limit={limit}")
            result = self._make_request("GET", "/produtos", params={"limit": limit, "offset": offset})
            if not result or "itens" not in result:
                print(f"  Nenhum resultado retornado")
                break

            items = result.get("itens", [])
            all_products.extend(items)
            print(f"  ✅ Carregados {len(items)} itens, total agora: {len(all_products)}")

            pagination = result.get("paginacao", {})
            total = pagination.get("total", 0)
            print(f"  Paginação: offset={offset}, limit={limit}, total={total}")

            if offset + limit >= total:
                print(f"  Fim da paginação atingido")
                break

            offset += limit

        print(f"\n✅ Total carregado: {len(all_products)} produtos")

        # Salvar em cache
        with open(self.cache_file, "w", encoding="utf-8") as f:
            json.dump(all_products, f, ensure_ascii=False, indent=2)
        print(f"💾 Cache salvo em {self.cache_file.name}")

        self.all_products = all_products
        return all_products

    def search_product_by_title(self, title: str, threshold: float = 0.5) -> Optional[Dict]:
        """Busca produto por similaridade de título"""
        if self.all_products is None:
            self.load_all_products_cached()

        if not self.all_products:
            return None

        title_lower = title.lower().strip()
        best_match = None
        best_score = threshold

        for product in self.all_products:
            prod_desc = product.get("descricao", "").lower()
            # Remover pontuação para melhor match
            title_clean = "".join(c for c in title_lower if c.isalnum() or c.isspace())
            desc_clean = "".join(c for c in prod_desc if c.isalnum() or c.isspace())

            score = SequenceMatcher(None, title_clean, desc_clean).ratio()

            if score > best_score:
                best_score = score
                best_match = (product, score)

        if best_match:
            return best_match[0]
        return None

    def update_product_price(self, product_id: int, price: float) -> bool:
        """Atualiza preço do produto"""
        payload = {
            "precos": {
                "preco": float(price)
            }
        }
        result = self._make_request("PUT", f"/produtos/{product_id}", json_data=payload)
        return result is not None

# ═══════════════════════════════════════════════════════════════════════════════
# PROCESSAMENTO
# ═══════════════════════════════════════════════════════════════════════════════

def process_ml_and_sync():
    """Processa ML, encontra produtos, identifica discrepâncias e sincroniza"""

    input_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
    if not input_file.exists():
        print(f"❌ Arquivo não encontrado: {input_file}")
        sys.exit(1)

    print(f"\n📂 Abrindo planilha: {input_file.name}")
    wb_in = openpyxl.load_workbook(input_file, data_only=True)
    ws_in = wb_in.active

    # Análise de colunas
    header = []
    for col_idx in range(1, ws_in.max_column + 1):
        header.append(ws_in.cell(1, col_idx).value)

    # Identificar colunas
    price_col_idx = None
    title_col_idx = None

    for idx, col_name in enumerate(header, 1):
        col_lower = str(col_name).lower() if col_name else ""
        if "price" in col_lower and price_col_idx is None:
            price_col_idx = idx
        if "title" in col_lower:
            title_col_idx = idx

    print(f"✅ Planilha carregada")
    print(f"  Título: coluna {title_col_idx}")
    print(f"  Preço: coluna {price_col_idx}")

    # Inicializar cliente
    client = TinyAPIClientV2()
    client.load_all_products_cached()

    # Criar nova planilha de resultado
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Preços Mais Baratos"

    # Cabeçalho
    result_columns = [
        "Linha ML", "Título ML", "Preço ML", "ID Tiny", "Descrição Tiny",
        "Preço Cadastro Tiny", "Diferença", "% Menor", "Novo Preço", "Status"
    ]

    for col_idx, col_name in enumerate(result_columns, 1):
        ws_out.cell(1, col_idx, col_name)

    # Estilo
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(result_columns) + 1):
        cell = ws_out.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font

    # Processamento
    print(f"\n🔍 Processando anúncios...")

    data_start_row = 6
    stats = {
        "total": 0,
        "found": 0,
        "cheaper": 0,
        "updated": 0,
    }

    out_row = 2

    for in_row_idx in range(data_start_row, min(data_start_row + 89, ws_in.max_row + 1)):
        title_ml = ws_in.cell(in_row_idx, title_col_idx).value if title_col_idx else None
        price_ml = ws_in.cell(in_row_idx, price_col_idx).value if price_col_idx else None

        if not title_ml or not price_ml:
            continue

        title_ml = str(title_ml).strip()
        price_ml = float(price_ml) if isinstance(price_ml, (int, float)) else 0

        if price_ml <= 0:
            continue

        stats["total"] += 1

        # Buscar no Tiny
        tiny_product = client.search_product_by_title(title_ml, threshold=0.4)

        if not tiny_product:
            continue

        stats["found"] += 1
        product_id = tiny_product.get("id")
        product_desc = tiny_product.get("descricao", "")[:60]
        price_tiny = tiny_product.get("precos", {}).get("preco", 0)

        # Verificar se é mais barato
        if price_tiny > 0 and price_tiny < price_ml:
            stats["cheaper"] += 1

            diff = price_ml - price_tiny
            pct_diff = (diff / price_ml) * 100
            new_price = price_ml + 0.01

            # Atualizar
            updated = client.update_product_price(product_id, new_price)
            if updated:
                stats["updated"] += 1
                status = "✅ ATUALIZADO"
            else:
                status = "❌ FALHA"

            # Preencher linha
            ws_out.cell(out_row, 1, in_row_idx)
            ws_out.cell(out_row, 2, title_ml)
            ws_out.cell(out_row, 3, price_ml)
            ws_out.cell(out_row, 4, product_id)
            ws_out.cell(out_row, 5, product_desc)
            ws_out.cell(out_row, 6, price_tiny)
            ws_out.cell(out_row, 7, diff)
            ws_out.cell(out_row, 8, pct_diff)
            ws_out.cell(out_row, 9, new_price)
            ws_out.cell(out_row, 10, status)

            print(f"⚠️  [{in_row_idx}] {title_ml[:40]}")
            print(f"    ML R$ {price_ml:.2f} > Tiny R$ {price_tiny:.2f} | Corrigido: R$ {new_price:.2f} {status}")

            out_row += 1

        print(f"   Processados: {stats['total']}/{89} | Encontrados: {stats['found']} | Mais baratos: {stats['cheaper']}", end="\r")

    # Salvar
    output_file = Path(r"C:\Users\user\Downloads\ML_Precos_Mais_Baratos.xlsx")
    print(f"\n\n💾 Salvando: {output_file.name}")
    wb_out.save(output_file)

    wb_in.close()
    wb_out.close()

    # Relatório
    print(f"\n" + "=" * 80)
    print(f"✅ SINCRONIZAÇÃO CONCLUÍDA")
    print(f"=" * 80)
    print(f"\n📊 RESULTADOS:")
    print(f"  Anúncios processados: {stats['total']}")
    print(f"  Encontrados no Tiny: {stats['found']}")
    print(f"  COM PREÇO MAIS BARATO: {stats['cheaper']} ⚠️")
    print(f"  Atualizados com sucesso: {stats['updated']}")
    print(f"\n📁 Relatório: {output_file}")
    print(f"🔗 Requisições API: {client.request_count}")


if __name__ == "__main__":
    print("╔════════════════════════════════════════════════════════════════╗")
    print("║  Sincronização v2: ML → Tiny (Com Cache)                     ║")
    print("║  Data: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "                              ║")
    print("╚════════════════════════════════════════════════════════════════╝")

    process_ml_and_sync()
