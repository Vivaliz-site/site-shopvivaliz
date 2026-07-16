#!/usr/bin/env python3
import csv
import logging
import sys
from pathlib import Path
from typing import Dict, List, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    print('Missing dependency: openpyxl. Install with: pip install openpyxl')
    sys.exit(1)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

SOURCE_FILES = [
    Path('mass_update_media_info.xlsx'),
    Path('mass_update_media_info_604371761_20260629183550.xlsx'),
]
MAPPING_FILE = Path('storage/uploaded_urls.csv')
OUTPUT_FILE = Path('planilhas/shopee_import.xlsx')


def get_source_file() -> Path:
    for file_path in SOURCE_FILES:
        if file_path.exists():
            return file_path
    raise FileNotFoundError('No source workbook found. Expected one of: ' + ', '.join(str(p) for p in SOURCE_FILES))


def load_mapping() -> Dict[str, Dict[str, str]]:
    if not MAPPING_FILE.exists():
        raise FileNotFoundError(f'Mapping file not found: {MAPPING_FILE}')
    mapping: Dict[str, Dict[str, str]] = {}
    with MAPPING_FILE.open('r', encoding='utf-8', newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            sku = str(row.get('sku', '') or '').strip()
            if not sku:
                continue
            mapping[sku] = {
                'image_url_1': str(row.get('image_url_1', '') or '').strip(),
                'image_url_2': str(row.get('image_url_2', '') or '').strip(),
                'image_url_3': str(row.get('image_url_3', '') or '').strip(),
                'image_url_4': str(row.get('image_url_4', '') or '').strip(),
            }
    return mapping


def normalize_header(value: Optional[str]) -> str:
    if value is None:
        return ''
    return str(value).strip().lower()


def find_columns(sheet) -> Dict[str, Optional[int]]:
    headers = [normalize_header(cell) for cell in next(sheet.iter_rows(values_only=True), [])]
    cols = {
        'sku': None,
        'item_id': None,
        'cover': None,
        'images': [],
    }
    for index, header in enumerate(headers, start=1):
        if not header:
            continue
        if cols['sku'] is None and 'sku' in header and 'image' not in header:
            cols['sku'] = index
        if cols['item_id'] is None and any(token in header for token in ('product_id', 'item_id', 'item id', 'id')) and 'image' not in header:
            cols['item_id'] = index
        if header == 'ps_item_cover_image':
            cols['cover'] = index
        if header.startswith('ps_item_image'):
            cols['images'].append(index)
    return cols


def update_sheet(sheet, mapping: Dict[str, Dict[str, str]], cols: Dict[str, Optional[int]]) -> int:
    sku_col = cols['sku']
    item_id_col = cols['item_id']
    cover_col = cols['cover']
    image_cols = cols['images']
    if not cover_col or len(image_cols) < 3:
        raise ValueError('Could not detect the Shopee image columns in the source workbook')

    updated = 0
    for row_idx in range(2, sheet.max_row + 1):
        sku_value = ''
        item_id_value = ''
        if sku_col:
            sku_value = str(sheet.cell(row=row_idx, column=sku_col).value or '').strip()
        if item_id_col:
            item_id_value = str(sheet.cell(row=row_idx, column=item_id_col).value or '').strip()
        if not sku_value and not item_id_value:
            continue
        lookup = sku_value or item_id_value
        mapping_row = mapping.get(lookup)
        if not mapping_row:
            continue
        sheet.cell(row=row_idx, column=cover_col).value = mapping_row['image_url_1']
        for idx, image_col in enumerate(image_cols[:3], start=2):
            sheet.cell(row=row_idx, column=image_col).value = mapping_row.get(f'image_url_{idx}', '')
        updated += 1
    return updated


def main(argv=None) -> int:
    try:
        source_file = get_source_file()
    except FileNotFoundError as exc:
        logger.error(exc)
        return 1
    try:
        mapping = load_mapping()
    except FileNotFoundError as exc:
        logger.error(exc)
        return 1
    workbook = load_workbook(filename=source_file)
    sheet = workbook.active
    cols = find_columns(sheet)
    logger.info(f'Detected columns: {cols}')
    try:
        updated = update_sheet(sheet, mapping, cols)
    except Exception as exc:
        logger.error(f'Failed to update sheet: {exc}')
        return 1
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(OUTPUT_FILE)
        logger.info(f'Wrote updated Shopee import spreadsheet to {OUTPUT_FILE}')
    except PermissionError as exc:
        alternate_output = OUTPUT_FILE.with_name(f'{OUTPUT_FILE.stem}.updated{OUTPUT_FILE.suffix}')
        workbook.save(alternate_output)
        logger.warning(
            f'Could not overwrite {OUTPUT_FILE} due to permission error. '
            f'Saved updated workbook to {alternate_output} instead.'
        )
    logger.info(f'Updated rows: {updated}')
    return 0


if __name__ == '__main__':
    sys.exit(main())