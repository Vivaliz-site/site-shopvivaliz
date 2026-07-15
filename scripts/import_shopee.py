#!/usr/bin/env python3
import concurrent.futures
import csv
import logging
import os
import re
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional
from urllib.parse import urlparse

try:
    import requests
except ImportError:
    print("Missing dependency: requests. Install with: pip install requests")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

DEFAULT_INPUT_FILE = Path('mass_update_media_info.xlsx')
OUTPUT_EXCEL = Path('planilhas/produtos.xlsx')
OUTPUT_RAW_ROOT = Path('storage/raw')

IMAGE_COLUMN_KEY = 'image'
MAX_WORKERS = 20
DOWNLOAD_TIMEOUT = 10
MAX_RETRIES = 3
SKU_MAPPING_FILE = Path('storage/sku_mapping.csv')
WINDOWS_RESERVED_NAMES = {
    'CON', 'PRN', 'AUX', 'NUL',
    'COM1', 'COM2', 'COM3', 'COM4', 'COM5', 'COM6', 'COM7', 'COM8', 'COM9',
    'LPT1', 'LPT2', 'LPT3', 'LPT4', 'LPT5', 'LPT6', 'LPT7', 'LPT8', 'LPT9',
}


def sanitize_folder_name(value: str) -> str:
    safe_value = re.sub(r'[\\/*:?"<>|]', '_', value or '')
    safe_value = re.sub(r'\s+', '_', safe_value)
    safe_value = safe_value.strip().rstrip('. ')
    safe_value = safe_value or 'unknown_sku'
    if safe_value.upper() in WINDOWS_RESERVED_NAMES:
        safe_value = f'_{safe_value}'
    return safe_value[:120]


def make_unique_folder_name(base_name: str, existing_names: set[str]) -> str:
    candidate = base_name
    suffix = 1
    while candidate in existing_names:
        candidate = f'{base_name}_{suffix}'
        suffix += 1
    existing_names.add(candidate)
    return candidate


def is_valid_url(value: Optional[str]) -> bool:
    if not value or not isinstance(value, str):
        return False
    value = value.strip()
    if not value:
        return False
    parsed = urlparse(value)
    return parsed.scheme in ('http', 'https') and bool(parsed.netloc)


def load_rows_from_xlsx(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(f'Input file not found: {path}')

    if OPENPYXL_AVAILABLE:
        logger.info('Using openpyxl to read the workbook')
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = []
        rows_iter = sheet.iter_rows(values_only=True)
        header_values = next(rows_iter, None)
        if header_values is None:
            return []
        header = [str(cell).strip() if cell is not None else '' for cell in header_values]
        for row in rows_iter:
            row_data = {header[idx].strip(): str(value).strip() if value is not None else '' for idx, value in enumerate(row) if idx < len(header)}
            if any(row_data.values()):
                rows.append(row_data)
        return rows

    logger.info('openpyxl not installed; using fallback XLSX reader')
    return load_rows_from_xlsx_fallback(path)


def load_rows_from_xlsx_fallback(path: Path) -> List[Dict[str, str]]:
    import zipfile
    import xml.etree.ElementTree as ET

    def parse_shared_strings(zf):
        strings = []
        try:
            with zf.open('xl/sharedStrings.xml') as shared:
                root = ET.parse(shared).getroot()
                for si in root.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                    text_fragments = []
                    for t in si.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t'):
                        if t.text:
                            text_fragments.append(t.text)
                    strings.append(''.join(text_fragments))
        except KeyError:
            pass
        return strings

    def get_sheet_path(zf):
        with zf.open('xl/workbook.xml') as workbook_xml:
            root = ET.parse(workbook_xml).getroot()
            sheets = root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet')
            if not sheets:
                raise ValueError('No sheets found in workbook.xml')
            first_sheet = sheets[0].attrib.get('name')
        with zf.open('xl/_rels/workbook.xml.rels') as rels_xml:
            root = ET.parse(rels_xml).getroot()
            for rel in root.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                if rel.attrib.get('Id') == sheets[0].attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'):
                    target = rel.attrib.get('Target')
                    return Path('xl') / target.replace('..', '')
        return Path('xl/worksheets/sheet1.xml')

    def parse_cell_value(cell, shared_strings):
        cell_type = cell.attrib.get('t')
        value_node = cell.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
        if value_node is None or value_node.text is None:
            inline_node = cell.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}is')
            if inline_node is not None:
                text_node = inline_node.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                return text_node.text if text_node is not None else ''
            return ''
        text = value_node.text
        if cell_type == 's':
            try:
                return shared_strings[int(text)]
            except Exception:
                return text
        return text

    with zipfile.ZipFile(path, 'r') as zf:
        shared_strings = parse_shared_strings(zf)
        sheet_path = get_sheet_path(zf)
        with zf.open(str(sheet_path)) as sheet_xml:
            root = ET.parse(sheet_xml).getroot()
            rows = []
            header: List[str] = []
            for row in root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
                row_index = int(row.attrib.get('r', '0'))
                values = {}
                for cell in row.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                    ref = cell.attrib.get('r', '')
                    col = re.sub(r'\d+$', '', ref)
                    value = parse_cell_value(cell, shared_strings)
                    values[col] = value.strip() if isinstance(value, str) else str(value)
                if row_index == 1:
                    max_col = max((len(col) for col in values.keys()), default=0)
                    header = [values.get(re.sub(r'\d+$', '', f'{chr(65 + i)}1'), '') for i in range(max_col)]
                else:
                    if not header:
                        continue
                    row_data = {}
                    for idx, col_name in enumerate(header):
                        if not col_name:
                            continue
                        col_letter = chr(65 + idx)
                        row_data[str(col_name).strip()] = values.get(col_letter, '')
                    if any(val for val in row_data.values()):
                        rows.append(row_data)
            return rows


def detect_columns(rows: List[Dict[str, str]]) -> Dict[str, Optional[str]]:
    if not rows:
        raise ValueError('Empty workbook: no rows to inspect')

    first_row = rows[0]
    headers = [str(key).strip() for key in first_row.keys()]
    sku_column = None
    item_id_column = None
    image_columns = []

    for header in headers:
        normalized = header.lower().strip()
        if not normalized:
            continue
        if 'sku' == normalized or normalized.endswith('sku') or ' sku' in normalized or normalized.startswith('sku'):
            sku_column = header
        if any(token in normalized for token in ('item_id', 'item id', 'itemid', 'id')) and 'image' not in normalized:
            if item_id_column is None:
                item_id_column = header
        if IMAGE_COLUMN_KEY in normalized:
            image_columns.append(header)

    if sku_column is None and item_id_column is None:
        raise ValueError('Could not detect item_id or sku column in the header row')
    if not image_columns:
        raise ValueError('Could not detect any image column in the header row')

    logger.info(f'Detected columns: sku={sku_column or "<none>"}, item_id={item_id_column or "<none>"}, images={image_columns}')
    return {
        'sku': sku_column,
        'item_id': item_id_column,
        'image_columns': image_columns,
    }


def choose_sku(row: Dict[str, str], sku_column: Optional[str], item_id_column: Optional[str]) -> str:
    sku_value = ''
    if sku_column and row.get(sku_column):
        sku_value = str(row.get(sku_column)).strip()
    if not sku_value and item_id_column:
        sku_value = str(row.get(item_id_column) or '').strip()
    return sku_value


def find_first_image_url(row: Dict[str, str], image_columns: List[str]) -> Optional[str]:
    for column in image_columns:
        value = row.get(column)
        if not value:
            continue
        if isinstance(value, str) and is_valid_url(value):
            return value.strip()
        candidate = str(value).strip()
        if is_valid_url(candidate):
            return candidate
    return None


def download_image(url: str, sku: str, folder_name: str) -> Optional[Path]:
    try:
        response = requests.get(url.strip(), timeout=DOWNLOAD_TIMEOUT, stream=True)
        if response.status_code != 200:
            logger.error(f'  [SKU {sku}] HTTP {response.status_code} for URL: {url}')
            return None

        sku_folder = OUTPUT_RAW_ROOT / folder_name
        sku_folder.mkdir(parents=True, exist_ok=True)
        image_path = sku_folder / '1.jpg'

        with open(image_path, 'wb') as image_file:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    image_file.write(chunk)

        logger.info(f'  [SKU {sku}] Downloaded image: {image_path}')
        return image_path
    except requests.exceptions.RequestException as exc:
        logger.error(f'  [SKU {sku}] Failed to download {url}: {exc}')
        return None
    except OSError as exc:
        logger.error(f'  [SKU {sku}] Failed to write image file: {exc}')
        return None


def download_image_with_retries(url: str, sku: str, folder_name: str) -> Optional[Path]:
    for attempt in range(1, MAX_RETRIES + 1):
        logger.info(f'  [SKU {sku}] Download attempt {attempt} for URL: {url}')
        result = download_image(url, sku, folder_name)
        if result is not None:
            return result
        if attempt < MAX_RETRIES:
            logger.warning(f'  [SKU {sku}] Retry {attempt} failed, retrying...')
            time.sleep(1)
    logger.error(f'  [SKU {sku}] All {MAX_RETRIES} download attempts failed for URL: {url}')
    return None


def save_sku_mapping(mapping: Dict[str, str]) -> None:
    SKU_MAPPING_FILE.parent.mkdir(parents=True, exist_ok=True)
    with SKU_MAPPING_FILE.open('w', encoding='utf-8', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=['sanitized_folder', 'sku'])
        writer.writeheader()
        for sanitized_folder, sku in sorted(mapping.items()):
            writer.writerow({'sanitized_folder': sanitized_folder, 'sku': sku})


def save_skus_excel(skus: List[str], output_path: Path) -> None:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError('openpyxl is required to write the output Excel file. Install it with: pip install openpyxl')
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'produtos'
    sheet.append(['sku'])
    for sku in sorted(k for k in skus if k):
        sheet.append([sku])
    workbook.save(output_path)
    logger.info(f'Wrote SKU list to {output_path}')


def main(argv: Optional[List[str]] = None) -> int:
    argv = argv or sys.argv[1:]
    input_path = Path(argv[0]) if argv else DEFAULT_INPUT_FILE
    logger.info(f'Input Excel file: {input_path}')

    try:
        rows = load_rows_from_xlsx(input_path)
    except Exception as exc:
        logger.error(f'Failed to read input file: {exc}')
        return 1

    if not rows:
        logger.error('No rows found in the input workbook')
        return 1

    try:
        columns = detect_columns(rows)
    except Exception as exc:
        logger.error(f'Column detection failed: {exc}')
        return 1

    skus = []
    skipped_count = 0
    rows_to_download = []
    folder_mapping: Dict[str, str] = {}
    used_folders: set[str] = set()

    for idx, row in enumerate(rows, start=1):
        sku = choose_sku(row, columns['sku'], columns['item_id'])
        if not sku:
            logger.warning(f'Row {idx}: no SKU/item_id value; skipping')
            skipped_count += 1
            continue

        image_url = find_first_image_url(row, columns['image_columns'])
        safe_folder = make_unique_folder_name(sanitize_folder_name(sku), used_folders)
        if image_url is None:
            logger.warning(f'  [SKU {sku}] No valid image URL found; skipping download')
            skus.append(sku)
            continue

        rows_to_download.append((idx, sku, safe_folder, image_url))
        folder_mapping[safe_folder] = sku
        skus.append(sku)

    total_images = len(rows_to_download)
    success_count = 0
    failed_count = 0

    if rows_to_download:
        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_task = {
                executor.submit(download_image_with_retries, image_url, sku, safe_folder): (idx, sku, safe_folder, image_url)
                for idx, sku, safe_folder, image_url in rows_to_download
            }
            for future in concurrent.futures.as_completed(future_to_task):
                idx, sku, safe_folder, image_url = future_to_task[future]
                try:
                    result = future.result()
                    if result is not None:
                        success_count += 1
                    else:
                        failed_count += 1
                except Exception as exc:
                    failed_count += 1
                    logger.error(f'  [SKU {sku}] Unexpected error downloading {image_url}: {exc}')
    
    try:
        save_skus_excel(skus, OUTPUT_EXCEL)
        if folder_mapping:
            save_sku_mapping(folder_mapping)
    except Exception as exc:
        logger.error(f'Failed to write output files: {exc}')
        return 1

    logger.info(f'Total images attempted: {total_images}; success: {success_count}; failed: {failed_count}; skipped rows: {skipped_count}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
