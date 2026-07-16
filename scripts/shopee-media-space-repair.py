#!/usr/bin/env python3
"""
Repara atualizacoes de imagem Shopee que falharam ao usar URL externa.

Fluxo:
1. Le a planilha de resultado da Shopee.
2. Relaciona cada linha com a planilha local de mapeamento e localiza o arquivo.
3. Faz upload do arquivo local via v2.media_space.upload_image.
4. Atualiza o anuncio com a nova capa, preservando os image_id_list atuais.

Uso local:
  python scripts/shopee-media-space-repair.py ^
    --result-xlsx "C:\\Users\\FRED\\Downloads\\Result_shopee_mass_update_media_info_604371761_20260702124508.xlsx" ^
    --mapping-xlsx "C:\\Users\\FRED\\Downloads\\mapeamento_shopee_geradas.xlsx" ^
    --dry-run
"""
from __future__ import annotations

import argparse
import json
import ntpath
import re
import unicodedata
from collections import Counter
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import load_workbook

from utils.shopee_client import ShopeeClient

load_dotenv()

FAILED_REASON_FRAGMENT = "falhou em carregar algumas imagens"
ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png"}
MAX_TOTAL_IMAGES = 9


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Repara imagens Shopee via Media Space API")
    parser.add_argument("--result-xlsx", required=True, help="Planilha Result_*.xlsx exportada pela Shopee")
    parser.add_argument("--mapping-xlsx", required=True, help="Planilha mapeamento_shopee_geradas.xlsx")
    parser.add_argument("--image-root", default="", help="Pasta alternativa com as imagens locais para runner/Actions")
    parser.add_argument("--output-json", default="", help="Caminho do relatorio JSON")
    parser.add_argument("--only-failed", action="store_true", help="Processa somente linhas com motivo de falha")
    parser.add_argument("--limit", type=int, default=0, help="Limita o numero de linhas processadas")
    parser.add_argument("--dry-run", action="store_true", help="Nao chama a API; apenas gera o plano")
    return parser.parse_args()


def normalize_text(text: Any) -> str:
    value = "" if text is None else str(text)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def normalize_sku(text: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_text(text))


def text_tokens(text: Any) -> set[str]:
    return {token for token in normalize_text(text).split() if len(token) >= 2}


def code_tokens(text: Any) -> set[str]:
    raw = re.findall(r"[A-Za-z0-9][A-Za-z0-9\-/*]{2,}", "" if text is None else str(text))
    tokens = set()
    for value in raw:
        normalized = normalize_sku(value)
        if len(normalized) >= 3:
            tokens.add(normalized)
    return tokens


def score_mapping(result_row: dict[str, Any], mapping_row: dict[str, Any]) -> dict[str, Any]:
    result_sku = str(result_row.get("sku") or "").strip()
    mapping_sku = str(mapping_row.get("sku") or "").strip()
    result_sku_norm = normalize_sku(result_sku)
    mapping_sku_norm = normalize_sku(mapping_sku)
    result_name = str(result_row.get("name") or "")
    mapping_name = str(mapping_row.get("name") or "")
    result_name_norm = normalize_text(result_name)
    mapping_name_norm = normalize_text(mapping_name)

    score = 0.0
    reasons: list[str] = []

    if result_sku and mapping_sku and result_sku.upper() == mapping_sku.upper():
        score += 120
        reasons.append("exact_sku")
    if result_sku_norm and mapping_sku_norm and result_sku_norm == mapping_sku_norm:
        score += 90
        reasons.append("normalized_sku")

    name_similarity = SequenceMatcher(None, result_name_norm, mapping_name_norm).ratio() * 100
    score += name_similarity * 0.8

    intersection = text_tokens(result_name) & text_tokens(mapping_name)
    union = text_tokens(result_name) | text_tokens(mapping_name)
    token_overlap = (len(intersection) / len(union) * 100) if union else 0.0
    score += token_overlap * 0.8

    result_codes = code_tokens(result_sku + " " + result_name)
    mapping_codes = code_tokens(mapping_sku + " " + mapping_name)
    overlap_codes = sorted(result_codes & mapping_codes)
    if overlap_codes:
        score += min(60, 12 * len(overlap_codes))
        reasons.append("code_overlap:" + "|".join(overlap_codes[:5]))

    if mapping_sku_norm and len(mapping_sku_norm) >= 4 and mapping_sku_norm in result_name_norm.replace(" ", ""):
        score += 50
        reasons.append("catalog_sku_in_mass_name")
    if result_sku_norm and len(result_sku_norm) >= 4 and result_sku_norm in mapping_name_norm.replace(" ", ""):
        score += 35
        reasons.append("mass_sku_in_mapping_name")

    file_name_norm = normalize_text(mapping_row.get("local_path") or "")
    if result_sku_norm and len(result_sku_norm) >= 4 and result_sku_norm in file_name_norm.replace(" ", ""):
        score += 20
        reasons.append("mass_sku_in_file_path")
    if mapping_sku_norm and len(mapping_sku_norm) >= 4 and mapping_sku_norm in file_name_norm.replace(" ", ""):
        score += 20
        reasons.append("mapping_sku_in_file_path")

    result_numbers = set(re.findall(r"\d+[a-z]*", result_name_norm))
    mapping_numbers = set(re.findall(r"\d+[a-z]*", mapping_name_norm))
    overlap_numbers = sorted(result_numbers & mapping_numbers)
    if overlap_numbers:
        score += 8 * len(overlap_numbers)
        reasons.append("numeric_overlap:" + "|".join(overlap_numbers[:5]))

    return {
        "score": round(score, 2),
        "name_similarity": round(name_similarity, 2),
        "token_overlap": round(token_overlap, 2),
        "reasons": reasons,
        "mapping": mapping_row,
    }


def decide_mapping(best: dict[str, Any], second: dict[str, Any] | None, result_name: str) -> tuple[str, float, str]:
    gap = round(best["score"] - (second["score"] if second else 0), 2)
    reason_keys = {value.split(":", 1)[0] for value in best["reasons"]}
    mapping_sku_norm = normalize_sku(best["mapping"].get("sku"))
    result_name_norm = normalize_text(result_name).replace(" ", "")

    if "exact_sku" in reason_keys or "normalized_sku" in reason_keys:
        return "fill", gap, "sku_direto"
    if mapping_sku_norm and len(mapping_sku_norm) >= 4 and mapping_sku_norm in result_name_norm and gap >= 20:
        return "fill", gap, "codigo_no_titulo"
    if best["name_similarity"] >= 85 and gap >= 20:
        return "fill", gap, "descricao_muito_proxima"
    if best["name_similarity"] >= 75 and best["token_overlap"] >= 65 and gap >= 25:
        return "fill", gap, "descricao_tokens_fortes"
    if best["score"] >= 150 and gap >= 15:
        return "review", gap, "revisao_recomendada"
    return "skip", gap, "sem_confianca_suficiente"


def resolve_local_image_path(original: Any, image_root: Path | None) -> Path | None:
    if not original:
        return None

    raw = str(original).strip()
    original_path = Path(raw)
    if original_path.is_file():
        return original_path

    if image_root and image_root.is_dir():
        file_names = [original_path.name, ntpath.basename(raw)]
        for file_name in dict.fromkeys(name for name in file_names if name):
            candidate = image_root / file_name
            if candidate.is_file():
                return candidate

    return original_path


def read_mapping_sheet(path: Path, image_root: Path | None) -> list[dict[str, Any]]:
    worksheet = load_workbook(path, read_only=True, data_only=True).active
    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        local_path = resolve_local_image_path(row[7], image_root)
        rows.append(
            {
                "mapping_id": "" if row[0] is None else str(row[0]).strip(),
                "sku": "" if row[1] is None else str(row[1]).strip(),
                "name": "" if row[2] is None else str(row[2]).strip(),
                "category": "" if row[3] is None else str(row[3]).strip(),
                "base_url": "" if row[4] is None else str(row[4]).strip(),
                "base_file": "" if row[5] is None else str(row[5]).strip(),
                "status": "" if row[6] is None else str(row[6]).strip(),
                "local_path": local_path,
                "status_shopee": "" if row[8] is None else str(row[8]).strip(),
            }
        )
    return rows


def read_result_sheet(path: Path, only_failed: bool, limit: int) -> list[dict[str, Any]]:
    worksheet = load_workbook(path, read_only=True, data_only=True).active
    rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(worksheet.iter_rows(min_row=7, values_only=True), start=7):
        if not any(value not in (None, "") for value in row[:5]):
            continue
        reason = "" if row[40] is None else str(row[40]).strip()
        if only_failed and FAILED_REASON_FRAGMENT not in reason.lower():
            continue
        rows.append(
            {
                "row_excel": row_index,
                "item_id": "" if row[0] is None else str(row[0]).strip(),
                "sku": "" if row[1] is None else str(row[1]).strip(),
                "name": "" if row[2] is None else str(row[2]).strip(),
                "category": "" if row[3] is None else str(row[3]).strip(),
                "cover_url": "" if row[4] is None else str(row[4]).strip(),
                "gallery_urls": [str(value).strip() for value in row[5:13] if value not in (None, "")],
                "reason": reason,
            }
        )
        if limit and len(rows) >= limit:
            break
    return rows


def choose_mapping(result_row: dict[str, Any], mappings: list[dict[str, Any]]) -> dict[str, Any]:
    best: dict[str, Any] | None = None
    second: dict[str, Any] | None = None
    for mapping in mappings:
        candidate = score_mapping(result_row, mapping)
        if best is None or candidate["score"] > best["score"]:
            second = best
            best = candidate
        elif second is None or candidate["score"] > second["score"]:
            second = candidate

    if best is None:
        raise RuntimeError("Nao foi possivel avaliar candidatos de mapeamento.")

    decision, gap, method = decide_mapping(best, second, result_row["name"])
    mapping = best["mapping"]
    local_path = mapping.get("local_path")
    exists = isinstance(local_path, Path) and local_path.is_file()
    extension_ok = exists and local_path.suffix.lower() in ALLOWED_EXTENSIONS

    return {
        "decision": decision,
        "decision_reason": method,
        "confidence_gap": gap,
        "confidence_score": best["score"],
        "name_similarity": best["name_similarity"],
        "token_overlap": best["token_overlap"],
        "mapping_id": mapping.get("mapping_id"),
        "mapping_sku": mapping.get("sku"),
        "mapping_name": mapping.get("name"),
        "mapping_local_path": str(local_path) if local_path else "",
        "mapping_status_shopee": mapping.get("status_shopee"),
        "image_exists": exists,
        "image_extension_ok": extension_ok,
        "reasons": best["reasons"],
    }


def chunked(values: list[int], size: int) -> list[list[int]]:
    return [values[index:index + size] for index in range(0, len(values), size)]


def extract_existing_image_ids(detail: dict[str, Any]) -> list[str]:
    candidates = [
        (((detail.get("image") or {}).get("image_id_list")) if isinstance(detail.get("image"), dict) else None),
        detail.get("image_id_list"),
        (((detail.get("images") or {}).get("image_id_list")) if isinstance(detail.get("images"), dict) else None),
    ]
    for candidate in candidates:
        if isinstance(candidate, list):
            return [str(value).strip() for value in candidate if str(value).strip()]
    return []


def build_remote_detail_map(client: ShopeeClient, item_ids: list[int]) -> dict[str, dict[str, Any]]:
    details: dict[str, dict[str, Any]] = {}
    for batch in chunked(item_ids, 50):
        for item in client.get_product_details(batch):
            item_id = str(item.get("item_id") or "").strip()
            if item_id:
                details[item_id] = item
    return details


def ensure_supported_image(path: Path) -> None:
    if not path.is_file():
        raise FileNotFoundError(f"Arquivo de imagem ausente: {path}")
    if path.suffix.lower() not in ALLOWED_EXTENSIONS:
        raise RuntimeError(f"Extensao nao suportada pela Shopee: {path.suffix}")
    if path.stat().st_size > 10 * 1024 * 1024:
        raise RuntimeError(f"Arquivo acima de 10MB: {path}")


def resolve_output_path(args: argparse.Namespace) -> Path:
    if args.output_json:
        return Path(args.output_json)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    return Path("reports") / f"shopee-media-space-repair-{stamp}.json"


def main() -> int:
    args = parse_args()
    result_xlsx = Path(args.result_xlsx)
    mapping_xlsx = Path(args.mapping_xlsx)
    output_json = resolve_output_path(args)
    output_json.parent.mkdir(parents=True, exist_ok=True)

    image_root = Path(args.image_root) if args.image_root else None
    mappings = read_mapping_sheet(mapping_xlsx, image_root)
    result_rows = read_result_sheet(result_xlsx, only_failed=args.only_failed, limit=args.limit)

    rows_to_fill: list[dict[str, Any]] = []
    report_rows: list[dict[str, Any]] = []
    counters: Counter[str] = Counter()

    for result_row in result_rows:
        match = choose_mapping(result_row, mappings)
        merged = {**result_row, **match}
        report_rows.append(merged)
        counters[merged["decision"]] += 1
        if merged["decision"] == "fill" and merged["image_exists"] and merged["image_extension_ok"]:
            rows_to_fill.append(merged)
            counters["planned_upload"] += 1
        elif merged["decision"] == "fill":
            counters["fill_but_invalid_file"] += 1

    detail_map: dict[str, dict[str, Any]] = {}
    if not args.dry_run and rows_to_fill:
        client = ShopeeClient()
        item_ids = [int(row["item_id"]) for row in rows_to_fill if str(row.get("item_id")).isdigit()]
        detail_map = build_remote_detail_map(client, item_ids)

        for row in rows_to_fill:
            try:
                local_path = Path(row["mapping_local_path"])
                ensure_supported_image(local_path)

                upload = client.upload_image_full(str(local_path))
                row["uploaded_image_id"] = str(upload.get("image_id") or "")
                row["uploaded_image_url"] = str(upload.get("image_url") or "")

                current_detail = detail_map.get(str(row["item_id"]), {})
                current_ids = extract_existing_image_ids(current_detail)
                new_ids = [row["uploaded_image_id"]] + [image_id for image_id in current_ids if image_id != row["uploaded_image_id"]]
                row["new_image_id_list"] = new_ids[:MAX_TOTAL_IMAGES]
                row["existing_image_id_list"] = current_ids

                client.update_product(int(row["item_id"]), image_ids=row["new_image_id_list"])
                row["api_status"] = "updated"
                counters["updated"] += 1
            except Exception as exc:
                row["api_status"] = "error"
                row["api_error"] = str(exc)
                counters["api_error"] += 1
    else:
        for row in rows_to_fill:
            row["api_status"] = "dry_run"

    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "mode": "dry_run" if args.dry_run else "live",
        "result_xlsx": str(result_xlsx),
        "mapping_xlsx": str(mapping_xlsx),
        "image_root": str(image_root) if image_root else "",
        "only_failed": args.only_failed,
        "processed_rows": len(result_rows),
        "rows_selected_for_upload": len(rows_to_fill),
        "counters": dict(counters),
        "rows": report_rows,
    }
    api_error_samples = [
        {
            "row_excel": row.get("row_excel"),
            "item_id": row.get("item_id"),
            "sku": row.get("sku"),
            "api_error": row.get("api_error"),
        }
        for row in report_rows
        if row.get("api_status") == "error"
    ][:5]
    if api_error_samples:
        summary["api_error_samples"] = api_error_samples
    output_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: v for k, v in summary.items() if k != "rows"}, ensure_ascii=False, indent=2))
    print(f"report_saved={output_json}")
    return 0 if not counters.get("api_error") else 2


if __name__ == "__main__":
    raise SystemExit(main())
