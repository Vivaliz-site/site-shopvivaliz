#!/usr/bin/env python3
"""
TESTAR: Processar alguns items um a um para validar a logica
"""
import os
import requests
import openpyxl
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv

for f in [Path(r"C:\site-shopvivaliz\.env.local"), Path(r"C:\site-shopvivaliz\.env")]:
    if f.exists():
        load_dotenv(f)
        break

token = os.getenv("TINY_ACCESS_TOKEN") or os.getenv("TINY_ACCESS_TOKEN")

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

print("\n" + "="*100)
print("TESTE: Processar alguns items um a um")
print("="*100)

# Carregar mapeamento
print("\n[1] Carregando mapeamento...")
mapping_sku = {}

for mfile in [
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-48.xls',
    r'C:\Users\user\Downloads\anuncios_2026-07-26-12-26-50.xls',
]:
    if Path(mfile).exists():
        df = pd.read_excel(mfile, sheet_name=0)
        for idx, row in df.iterrows():
            sku = str(row.get('Produto (SKU)')).strip().upper()
            titulo = row.get('Título')

            if sku and sku != 'NAN':
                mapping_sku[sku] = titulo

print(f"    Total: {len(mapping_sku)} SKUs")

# Carregar produtos Tiny
print("\n[2] Carregando produtos Tiny...")
url = "https://api.tiny.com.br/public-api/v3/produtos?limit=100&offset=0"
resp = requests.get(url, headers=headers, timeout=30)

produtos_tiny_por_sku = {}

for prod in resp.json().get("itens", []):
    sku = str(prod.get("sku", "")).strip().upper()
    if sku:
        produtos_tiny_por_sku[sku] = {
            'id': prod.get("id"),
            'descricao': prod.get("descricao"),
            'preco': prod.get("precos", {}).get("preco")
        }

print(f"    Total: {len(produtos_tiny_por_sku)} produtos")

# Carregar planilha ML
print("\n[3] Carregando planilha base ML...")
ml_file = Path(r"C:\Users\user\Downloads\Anuncios-2026_07_26-09_32(1).xlsx")
wb_ml = openpyxl.load_workbook(ml_file, data_only=True)
ws_ml = wb_ml.active

header = {}
for col_idx in range(1, ws_ml.max_column + 1):
    val = ws_ml.cell(1, col_idx).value
    if val:
        header[str(val).strip()] = col_idx

title_col = header.get("TITLE", 5)
price_col = header.get("PRICE", 9)

# Testar alguns items
print("\n[4] Testando alguns items:\n")

teste_rows = [6, 7, 8, 9, 10]  # Primeiras 5 linhas

for in_row in teste_rows:
    title_ml = ws_ml.cell(in_row, title_col).value
    price_ml = ws_ml.cell(in_row, price_col).value

    if not title_ml or not price_ml:
        continue

    title_ml = str(title_ml).strip()
    price_ml = float(price_ml)

    print(f"Linha {in_row}: {title_ml[:55]}")
    print(f"  Preco ML: R$ {price_ml}")

    # Procurar por SKU
    title_lower = title_ml.lower()
    palavras = [p for p in title_lower.split() if len(p) > 3]

    print(f"  Palavras-chave: {palavras[:3]}")

    # Procurar match
    best_sku = None
    best_matches = 0

    for sku, titulo_map in mapping_sku.items():
        titulo_map_lower = titulo_map.lower()

        matches = sum(1 for palavra in palavras if palavra in titulo_map_lower)

        if matches > best_matches:
            best_matches = matches
            best_sku = sku

    if best_sku and best_sku in produtos_tiny_por_sku:
        prod = produtos_tiny_por_sku[best_sku]
        print(f"  ENCONTRADO!")
        print(f"    SKU: {best_sku}")
        print(f"    ID Tiny: {prod['id']}")
        print(f"    Descricao: {prod['descricao'][:50]}")
        print(f"    Preco Cadastro: R$ {prod['preco']}")
    else:
        print(f"  NAO ENCONTRADO")
        if best_sku:
            print(f"    SKU encontrado: {best_sku}")
            print(f"    Mas nao existe no Tiny")

    print()

wb_ml.close()
